# -*- coding: utf-8 -*-
# Generated from folder: /Users/alysonchen/Downloads/KeDing/#拜訪清單

# ==================================================
# SOURCE FILE: 0.主管前50大客戶_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


top_50 = kd.get_sap_with_relate_company(year_ago_1d)
top_50 = top_50.groupby('公司代號')['未稅本位幣'].sum().reset_index()

# 根據top_50銷貨數據所在公司,並查找公司的各維度數據
# and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
all_company_ids = top_50['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(top_50, account_df, on = '公司代號', how = 'inner')
K_invite = K_invite[~K_invite['資料區域群組名稱'].astype(str).str.contains("TW-Z")]


# 增加區域數據
區域業務 = pd.read_excel("Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/91.台灣業務負責區域表/台灣業務負責區域_2026.xlsx", dtype='object',sheet_name='區域表(報表用)',skiprows=1 ).ffill()
區域業務 = 區域業務[['小區主管','小區','大區']]
區域業務 = 區域業務[~區域業務['小區'].astype(str).str.contains('TW-Z', na=False) &區域業務['小區'].notna()]
K_invite = pd.merge(K_invite, 區域業務, left_on = '資料區域群組名稱', right_on = '小區', how = 'left')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')
 

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')


# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')

target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite = K_invite[K_invite['大區'].notna()].copy()
K_invite['未稅本位幣'] = pd.to_numeric(K_invite['未稅本位幣'], errors='coerce').fillna(0).astype(int)
K_invite['Rank'] = K_invite.groupby('大區')['未稅本位幣'].rank(method='first', ascending=False).astype(int)
K_invite = K_invite[K_invite['Rank'] <= 50].copy()
K_invite = K_invite.sort_values(by=['大區', 'Rank'])
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first').rename(columns={"小區主管": "執行人"})


# 主旨欄位：銷貨50大排行榜-{RANK}
K_invite['主旨'] = '銷貨50大排行榜-' + K_invite['Rank'].astype(str)



path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
K_invite.to_excel(path + "主管前50大客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 1.2今年交易金額差異超百萬.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_7d = (datetime.today() - relativedelta(months=7) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000
jan_first = datetime.today().replace(month=1, day=1).date().strftime("%Y/%m/%d")
last_jan_first = (datetime.today().replace(month=1, day=1) - relativedelta(years=1)).strftime("%Y/%m/%d")



def get_sap_with_relate_company(year_ago_3d):
    # 取得 SAP 銷售資料
    sap = kd.get_data_from_MSSQL(f'''
        SELECT buyer as SAP公司代號, taxfree_basecurr as 未稅本位幣,  planned_shipping_date as 預計發貨日期 
        FROM sap_sales_data 
        WHERE  buyer LIKE 'TW%'
        AND planned_shipping_date >= '{year_ago_3d}'
    ''')
    sap['預計發貨日期'] = pd.to_datetime(sap['預計發貨日期'], errors='coerce')

    # 取得公司對應表，只取需要欄位，避免名稱衝突
    company_map = kd.get_data_from_MSSQL('''
                        SELECT  company_id 公司代號
                            ,sap_company_id SAP公司代號
                            ,company_id_parent 關聯公司
                        FROM [raw_data].[dbo].[crm_related_company]
    ''')

    # 合併後取代公司代號
    merged = pd.merge(sap, company_map, on='SAP公司代號', how='left')
    merged['公司代號'] = merged['關聯公司'].fillna(merged['公司代號'])

    account_type = kd.get_data_from_MSSQL('''
                        SELECT  distinct buyer SAP公司代號
                            ,industry 公司型態
                            FROM sap_sales_data 
                         where industry like '%C%' or industry like '%D%'
    ''')

    merged = pd.merge(merged, account_type, on='SAP公司代號', how='inner')


    # 選擇最終欄位（只保留最終的公司代號）
    result = merged[['公司代號','公司型態', '未稅本位幣', '預計發貨日期']].dropna(subset=['公司代號'])

    return result



# 銷貨數據
total_sales= get_sap_with_relate_company(last_jan_first)


sales_last_year = total_sales[(total_sales['預計發貨日期'] >= last_jan_first) &(total_sales['預計發貨日期'] < jan_first)]
sales_this_year = total_sales[total_sales['預計發貨日期'] >= jan_first]

target_ly = sales_last_year.groupby(['公司代號','公司型態'])['未稅本位幣'].sum().reset_index()
target_ty = sales_this_year.groupby(['公司代號','公司型態'])['未稅本位幣'].sum().reset_index()

# 合併今年與去年的銷售
merged = pd.merge(
    target_ly, target_ty,
    on='公司代號',
    how='inner',
    suffixes=('_ly', '_ty')  # ly = last year, ty = this year
)
merged = merged.drop(columns=['公司型態_ty']).rename(columns={'公司型態_ly': '公司型態'})

# 計算銷售差異（去年 - 今年）
merged['差異'] = merged['未稅本位幣_ly'] - merged['未稅本位幣_ty']

# 篩選出差異大於 100 萬的公司
diff_over_1m = merged[merged['差異'].abs() > 1_000_000]

# 根據diff_over_1m銷貨數據所在公司,並查找公司的各維度數據
all_company_ids = diff_over_1m['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' 
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(diff_over_1m, account_df, on = '公司代號', how = 'inner')



#  排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts, source_type='拜訪')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人

            from customEntity22__c 
            where customItem37__c  like '%TW%'
            '''
            )

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}")
filename = f"兩年交易差異百萬客戶_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print('數據導出成功~')

# ==================================================
# SOURCE FILE: 1.3北中南區前20大客戶中為W類_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000
jan_first = datetime.today().replace(month=1, day=1).date().strftime("%Y/%m/%d")
last_jan_first = (datetime.today().replace(month=1, day=1) - relativedelta(years=1)).strftime("%Y/%m/%d")

top_20 = kd.get_sap_with_relate_company(year_ago_1d)
top_20 = top_20.groupby('公司代號')['未稅本位幣'].sum().reset_index()

# 根據top_20銷貨數據所在公司,並查找公司的各維度數據
# and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
all_company_ids = top_20['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' 
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(top_20, account_df, on = '公司代號', how = 'inner')
K_invite = K_invite[~K_invite['資料區域群組名稱'].astype(str).str.contains("TW-Z")]


# 增加區域數據
區域業務 = pd.read_excel("Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/91.台灣業務負責區域表/台灣業務負責區域_2026.xlsx", dtype='object',sheet_name='區域表(報表用)',skiprows=1 ).ffill()
區域業務 = 區域業務[['小區主管','小區','大區']]
區域業務 = 區域業務[~區域業務['小區'].astype(str).str.contains('TW-Z', na=False) &區域業務['小區'].notna()]
K_invite = pd.merge(K_invite, 區域業務, left_on = '資料區域群組名稱', right_on = '小區', how = 'left')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')
 

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

#   排除近3個月拜訪, K大過
pass_visited =  kd.last_connected(month_ago_3ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')

# target = ['經營客戶']
# K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite = K_invite[K_invite['大區'].notna()].copy()
K_invite['未稅本位幣'] = pd.to_numeric(K_invite['未稅本位幣'], errors='coerce').fillna(0).astype(int)
K_invite['Rank'] = K_invite.groupby('大區')['未稅本位幣'].rank(method='first', ascending=False).astype(int)
K_invite = K_invite[K_invite['Rank'] <= 20].copy()
K_invite = K_invite.sort_values(by=['大區', 'Rank'])
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')

K_invite = K_invite[K_invite['公司型態'] == 'W'].copy()




# 主旨欄位：銷貨50大排行榜-{RANK}
K_invite['主旨'] = '(指標)北中南區前20大客戶中為W類'



path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
K_invite.to_excel(path + "北中南區前20大客戶中為W類_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 1.5近半年未叫貨_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_7d = (datetime.today() - relativedelta(months=7) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000

# 銷貨數據
total_sales= kd.get_sap_with_relate_company(month_ago_7d)

total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

sales_month_7_to_6 = total_sales[(total_sales['預計發貨日期'] >= month_ago_7d) &(total_sales['預計發貨日期'] < month_ago_6d)]
sales_last_6_months = total_sales[total_sales['預計發貨日期'] >= month_ago_6d]

target_summary = sales_month_7_to_6.groupby('公司代號')['未稅本位幣'].sum().reset_index()
sap_ids_last6 = set(sales_last_6_months['公司代號'].unique())
sales_nobuy_after = target_summary[~target_summary['公司代號'].isin(sap_ids_last6)]


# 根據sales_month_7_to_6銷貨數據所在公司,並查找公司的各維度數據
all_company_ids = sales_nobuy_after['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%' or customItem199__c like '%D%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(sales_nobuy_after, account_df, on = '公司代號', how = 'inner')



#  排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts, source_type='拜訪')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人

            from customEntity22__c 
            where customItem37__c  like '%TW%'
            '''
            )

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}")
filename = f"近半年未叫貨_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print('數據導出成功~')

# ==================================================
# SOURCE FILE: 1.1客訴後未交易客戶_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

year_ago_3d = (datetime.today() - relativedelta(years=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(years=3) + relativedelta(days=1)).date()).timestamp()*1000


# select from trackingrecord  -->  客訴數據,最近3年
complaint = kd.get_data_from_CRM(
            f'''
                    select
                    name 客訴代號,
                    customItem40__c 資料日期,
                    customItem4__c 工作類別,
                    accountCode__c 公司代號
                    from customEntity15__c
                    where customItem118__c like '%TW%' and (customItem4__c = 12 or customItem168__c in ('C3','C3-1','C3-2'))
                    and customItem40__c >= {year_ago_3ts}
            ''')

kd.convert_to_date(complaint,"資料日期")
complaint = complaint.sort_values('資料日期', ascending=False).drop_duplicates(subset=['公司代號'], keep='first')

# 關聯公司
complaint = kd.merge_company_to_parent(complaint)

# 根據客訴數據找到所在公司,並查找公司的各維度數據
all_company_ids = complaint['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱,customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%' or customItem199__c like '%D%')
        and accountCode__c in {target_customer_str}
    '''

    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(complaint, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')


#   排除近2個月拜訪, K大過
pass_visited =  kd.last_connected(month_ago_2ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]


# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)


# from 銷貨
sap = kd.get_sap_with_relate_company(year_ago_3d)
K_invite['資料日期'] = pd.to_datetime(K_invite['資料日期'])
sap['預計發貨日期'] = pd.to_datetime(sap['預計發貨日期'])


# 每家公司最晚投訴日
complaint_latest = K_invite.groupby('公司代號')['資料日期'].max().reset_index()
complaint_latest.columns = ['公司代號', '最近投訴日期']

# inner join 獲取投訴後下單的
sap_merge = pd.merge(sap, complaint_latest, left_on='公司代號', right_on='公司代號', how='inner')

# 篩選出在投訴日期之後還有下單的, 反向篩選那些沒有再回購的公司
came_back = sap_merge[sap_merge['預計發貨日期'] > sap_merge['最近投訴日期']]['公司代號'].unique()
no_return = complaint_latest[~complaint_latest['公司代號'].isin(came_back)]


K_invite = K_invite[K_invite['公司代號'].isin(no_return['公司代號'])]

# # 關聯公司
# K_invite = kd.merge_company_to_parent(K_invite, kd)
# K_invite = K_invite[K_invite['公司代號'].isin(account_df['公司代號'])]

K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}")
filename = f"客訴後未交易客戶_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 2-新建CRM客戶_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000


new_company = kd.get_data_from_CRM(f'''
        select accountCode__c 公司代號
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%' or customItem199__c like '%D%')
        and customItem226__c >= {month_ago_1ts}
''')


new_company = kd.merge_company_to_parent(new_company)






# 根據sales_month_7_to_6銷貨數據所在公司,並查找公司的各維度數據
all_company_ids = new_company['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%' or customItem199__c like '%D%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(new_company, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#  排除近1個月拜訪過
pass_visited =  kd.last_connected(month_ago_1ts, source_type='拜訪')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 近一年有交易
sap = kd.get_sap_with_relate_company(year_ago_1d)
K_invite = K_invite[~K_invite['公司代號'].isin(sap['公司代號'].dropna())]

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}")
filename = f"新建CRM客戶_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 4.五年內有交易之客戶_wenbin.py
# ==================================================


###############################    contact類型,不找關聯公司,只要最佳聯絡人
import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

year_ago_5d = (datetime.today() - relativedelta(years=5) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_3d = (datetime.today() - relativedelta(years=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
day_ago_45ts = pd.to_datetime((datetime.today() - relativedelta(days=45) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


# 銷貨數據
#total_sales = kd.get_sap_with_relate_company(year_ago_5d)

total_sales = kd.get_data_from_MSSQL(f'''
        SELECT buyer as SAP公司代號, taxfree_basecurr as 未稅本位幣,  planned_shipping_date as 預計發貨日期 
        FROM sap_sales_data 
        WHERE  buyer LIKE 'TW%' 
        AND planned_shipping_date >= '{year_ago_5d}'
    ''')

target_summary = total_sales.groupby('SAP公司代號')['未稅本位幣'].sum().reset_index()


# 根據最近5年的銷貨數據,找到所在公司的各維度數據
all_company_ids = target_summary['SAP公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,customItem202__c 公司地址,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
        and SAP_CompanyID__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(target_summary, account_df, on = 'SAP公司代號', how = 'inner')



# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

# 只保留開發客戶
target = ['經營客戶']
K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

# 排除近45天拜訪過,K大過
pass_visited =  kd.last_connected(day_ago_45ts, first='all',merge_type="alone")
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts,merge_type="alone")
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 剔除配合的關係聯絡人
K_invite = K_invite[~K_invite['關係狀態'].astype(str).str.contains("配合")]

target = ['經營客戶']
K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite,data_type="contact")

# 篩選最佳聯絡人
K_invite = kd.best_contact(K_invite)
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


company_map = kd.get_data_from_MSSQL(f'''SELECT  company_id ,company_id_parent 關聯公司代號
           FROM [raw_data].[dbo].[crm_related_company]''')

K_invite = pd.merge(K_invite, company_map, left_on = '公司代號', right_on  = 'company_id', how='left')

# 建偏好旗標：公司代號 等於 關聯公司代號 → True（優先保留）
K_invite = K_invite.assign(
    _prefer=(K_invite['公司代號'].fillna('').astype(str).str.strip()
             == K_invite['關聯公司代號'].fillna('').astype(str).str.strip())
)

# 按照手機號碼排序（先偏好、再日期）
K_sorted = K_invite.sort_values(by=['手機號碼', '_prefer', '日期'],
                                ascending=[True, False, True],
                                na_position='last')

# 找到每個手機號碼保留的第一筆
keep_index = K_sorted.drop_duplicates(subset=['手機號碼'], keep='first').index

# 建立一個布林欄位：True=被保留, False=被排除
K_sorted['_是否保留'] = K_sorted.index.isin(keep_index)

# 分出被排除的資料
excluded = K_sorted[~K_sorted['_是否保留']]

print(f"總共有 {len(excluded)} 筆被排除")
print(excluded[['手機號碼', '公司代號', '關聯公司代號', '日期']].head(20))

included = K_sorted[K_sorted['_是否保留']]



path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
included.to_excel(path + "五年內有交易之客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')

# ==================================================
# SOURCE FILE: 5.C類冷卻期客戶_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_2d = (datetime.today() - relativedelta(years=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000

# 銷貨數據
total_sales = kd.get_sap_with_relate_company(year_ago_2d)
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])


sales_year_ago_2 = total_sales[ (total_sales['預計發貨日期'] >= year_ago_2d) & (total_sales['預計發貨日期'] <  year_ago_1d)]
sales_year_ago_1 = total_sales[total_sales['預計發貨日期'] >= year_ago_1d]

target_summary_1 = sales_year_ago_2.groupby('公司代號')['未稅本位幣'].sum().reset_index()
target_summary_2 = sales_year_ago_1.groupby('公司代號')['未稅本位幣'].sum().reset_index()
target_summary_1 = target_summary_1[target_summary_1['未稅本位幣'] > 200000]
target_summary_2 = target_summary_2[target_summary_2['未稅本位幣'] < 50000]

summary_merged = pd.merge(
    target_summary_1.rename(columns={'未稅本位幣': '銷售_13_24月'}),
    target_summary_2.rename(columns={'未稅本位幣': '銷售_0_12月'}), on='公司代號',how='inner')



# 根據summary_merged銷貨數據所在公司,並查找公司的各維度數據
# 特定C類
all_company_ids = summary_merged['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%')
        and accountCode__c in {target_customer_str}
    '''

    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(summary_merged, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            '''
            )

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')


#  排除近2個月拜訪, K大過
pass_visited =  kd.last_connected(month_ago_2ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 剔除近兩個月有銷貨記錄
sales_month_2 = kd.get_sap_with_relate_company(month_ago_2d)
K_invite = K_invite[~K_invite['公司代號'].isin(sales_month_2['公司代號'])]


# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
K_invite.to_excel(path + "C類冷卻期客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')




# ==================================================
# SOURCE FILE: 6.50萬客戶三個月未拜訪_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_3d = (datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_2d = (datetime.today() - relativedelta(years=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000

# 銷貨數據
total_sales = kd.get_sap_with_relate_company(year_ago_1d)
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()
target_summary = target_summary[target_summary['未稅本位幣'] > 200000]



# 根據summary_merged銷貨數據所在公司,並查找公司的各維度數據
all_company_ids = target_summary['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(target_summary, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')


# 排除"近3個月內有拜訪、K大"
pass_visited =  kd.last_connected(month_ago_3ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')


# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


# 近一年業績區間切分
invite_50up = K_invite[K_invite['未稅本位幣'] > 500000].copy()
invite_20_50 = K_invite[(K_invite['未稅本位幣'] >= 200000) & (K_invite['未稅本位幣'] <= 500000)].copy()



# path C:\Users\TW0002.TPTWKD\Desktop
path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
invite_50up.to_excel(path + "近1年50萬以上客戶三個月未拜訪_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
invite_20_50.to_excel(path + "近1年20-50萬客戶三個月未拜訪_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 8.歷史銷貨排行_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000



top_20 = kd.get_sap_with_relate_company(year_ago_1d)
top_20 = top_20.groupby('公司代號')['未稅本位幣'].sum().reset_index()


# 根據top_20銷貨數據所在公司,並查找公司的各維度數據
# 特定C類
all_company_ids = top_20['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
        and accountCode__c in {target_customer_str}
    '''

    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(top_20, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            '''
            )

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')


# 排除"近2個月內有拜訪、K大"
pass_visited =  kd.last_connected(month_ago_2ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')





# 關聯到大公司,並增加相對應的大小群組說明
group_info = kd.get_data_from_MSSQL('''select  buyer SAP公司代號, s_group_desc 小群組說明, l_group_desc 大群組說明  ,planned_shipping_date
                                    from [clean_data].[dbo].[sap_sales_data_processed]
                                    where buyer like 'TW%'
                                    ''')
group_info.sort_values(by=['SAP公司代號', 'planned_shipping_date'], ascending=[True, False], inplace=True)
group_info = kd.add_relate_company(group_info,"SAP")
group_info = group_info.drop_duplicates(subset=['SAP公司代號'], keep='first')

K_invite = pd.merge(K_invite, group_info[['SAP公司代號','小群組說明','大群組說明']],on ='SAP公司代號',how='left' )



# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()

K_invite['排名'] = K_invite.groupby('小群組說明')['未稅本位幣']\
                          .rank(method='first', ascending=False)
K_invite_top20 = K_invite[K_invite['排名'] <= 20].copy()
K_invite_top20.drop(columns=['排名'], inplace=True)

K_invite_top20 = K_invite_top20.sort_values(by='日期', ascending=True, na_position='first')

path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
K_invite_top20.to_excel(path + "歷史銷貨排行_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 9.高資本額公司_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


target_account = kd.get_data_from_CRM('''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,customItem204__c 資本額,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%' or customItem199__c like '%D%')
        ''')

K_invite = kd.merge_company_to_parent(target_account)

K_invite = (
    K_invite.assign(資本額=pd.to_numeric(K_invite['資本額'], errors='coerce'))
    .dropna(subset=['資本額']).query('資本額 > 0')
    .astype({'資本額': int}).sort_values('資本額', ascending=False))



# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近2個月拜訪, K大過
pass_visited =  kd.last_connected(month_ago_2ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
K_invite.to_excel(path + "高資本額客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')




# ==================================================
# SOURCE FILE: 10.近兩個月未拜訪_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


target_account = kd.get_data_from_CRM('''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%' or customItem199__c like '%D%')
        ''')

K_invite = kd.merge_company_to_parent(target_account)




# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近2個月拜訪, K大過
pass_visited =  kd.last_connected(month_ago_2ts, first='all')
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['開發客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/"
K_invite.to_excel(path + "近兩個月未拜訪_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')

# ==================================================
# SOURCE FILE: 11.備份與去重.py
# ==================================================

import pandas as pd
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta

# 設定今天日期與資料夾名稱
today_str = datetime.today().strftime('%Y.%m.%d')
month_folder = (datetime.today() + relativedelta(months=1)).strftime('%Y.%m')

#    f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{month_folder}"
# 定位到指定資料夾
folder = Path(    
   f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{month_folder}"
)



from pathlib import Path
import shutil

# === 先備份：把今天的所有檔案複製到「文斌備份用」資料夾 ===
backup_dir = folder / "文斌備份用"
backup_dir.mkdir(parents=True, exist_ok=True)

copied = 0
for src in folder.glob(f"*{today_str}.xlsx"):
    if src.is_file():
        dst = backup_dir / src.name
        # 若已存在就覆蓋；想保留舊檔可改成：if not dst.exists(): 再 copy2
        shutil.copy2(src, dst)
        copied += 1
print(f" 已備份 {copied} 個檔案到：{backup_dir}")



# 1. 讀取主管前50大客戶檔案，保留公司代號
top_customer_path = folder / f"主管前50大客戶_{today_str}.xlsx"
df_top = pd.read_excel(top_customer_path, dtype=str)  # 保留格式
target_company = df_top['公司代號'].dropna().unique().tolist()
print(f" 抓到的公司代號共 {len(target_company)} 筆")

# 2. 遍歷其他檔案，排除主管檔案
for file in folder.glob(f"*{today_str}.xlsx"):
    if "主管前50大客戶" in file.name:
        continue  # 跳過主管那份

    print(f" 處理中: {file.name}")
    df = pd.read_excel(file, dtype=str)  # 讀入所有欄位為字串，保留原始格式

    if '公司代號' not in df.columns:
        print(f" 檔案中沒有 '公司代號' 欄位，略過: {file.name}")
        continue

    # 3. 過濾掉重複公司
    before = len(df)
    df_filtered = df[~df['公司代號'].isin(target_company)]
    after = len(df_filtered)
    print(f" 剔除 {before - after} 筆公司代號，剩下 {after} 筆")

    # 4. 儲存過濾後的檔案
    save_path = file.with_name(file.stem + ".xlsx")
    df_filtered.to_excel(save_path, index=False)
    print(f" 已儲存: {save_path.name}")



# ###########
# # 只處理特定檔案
# import pandas as pd
# from pathlib import Path
# from datetime import datetime
# from dateutil.relativedelta import relativedelta

# # 設定今天日期與資料夾名稱
# today_str = datetime.today().strftime('%Y.%m.%d')  # 今天日期（用來找 4/28的檔案）
# month_folder = (datetime.today() + relativedelta(months=1)).strftime('%Y.%m')

# # 指定資料夾
# folder = Path(
#     f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{month_folder}"
# )

# # 1. 固定抓 2025/04/24 的主管前50大客戶
# top_customer_path = folder / "主管前50大客戶_2025.04.24.xlsx"
# df_top = pd.read_excel(top_customer_path, dtype=str)  # 保留格式
# target_company = df_top['公司代號'].dropna().unique().tolist()
# print(f" 抓到的公司代號共 {len(target_company)} 筆")

# # 2. 只處理指定的兩個檔案
# file_names = [
#     f"新建CRM客戶_{today_str}.xlsx",
#     f"近半年未叫貨_{today_str}.xlsx"
# ]

# for file_name in file_names:
#     file = folder / file_name
#     if not file.exists():
#         print(f" 找不到檔案: {file.name}，跳過")
#         continue

#     print(f" 處理中: {file.name}")
#     df = pd.read_excel(file, dtype=str)  # 全部讀成字串

#     if '公司代號' not in df.columns:
#         print(f" 檔案中沒有 '公司代號' 欄位，略過: {file.name}")
#         continue

#     # 過濾掉主管前50大客戶
#     before = len(df)
#     df_filtered = df[~df['公司代號'].isin(target_company)]
#     after = len(df_filtered)
#     print(f" 剔除 {before - after} 筆公司代號，剩下 {after} 筆")

#     # 儲存過濾後的檔案（覆蓋原檔）
#     save_path = file.with_name(file.stem + ".xlsx")
#     df_filtered.to_excel(save_path, index=False)
#     print(f" 已儲存: {save_path.name}")




# import pandas as pd
# from pathlib import Path

# # 1. 設定檔案路徑
# folder = Path(r"C:\Users\TW0002.TPTWKD\Desktop\0620")

# # 2. 主管前50大客戶的檔案（6/20 的版本）
# top_customer_path = folder / "主管前50大客戶_2025.06.20.xlsx"
# df_top = pd.read_excel(top_customer_path, dtype=str)
# target_company = df_top['公司代號'].dropna().unique().tolist()
# print(f" 抓到的公司代號共 {len(target_company)} 筆")

# # 3. 只處理這兩個檔案（6/25 的）
# file_names = [
#     "CD類儲值金客戶_2025.06.25.xlsx",
#     "CD類儲值金未滿10萬客戶_2025.06.25.xlsx"
# ]

# for file_name in file_names:
#     file = folder / file_name
#     if not file.exists():
#         print(f" 找不到檔案: {file.name}，跳過")
#         continue

#     print(f" 處理中: {file.name}")
#     df = pd.read_excel(file, dtype=str)

#     if '公司代號' not in df.columns:
#         print(f" 檔案中沒有 '公司代號' 欄位，略過: {file.name}")
#         continue

#     # 過濾掉主管前50大客戶
#     before = len(df)
#     df_filtered = df[~df['公司代號'].isin(target_company)]
#     after = len(df_filtered)
#     print(f" 剔除 {before - after} 筆公司代號，剩下 {after} 筆")

#     # 覆蓋儲存
#     save_path = file.with_name(file.stem + ".xlsx")
#     df_filtered.to_excel(save_path, index=False)
#     print(f" 已儲存: {save_path.name}")


# ==================================================
# SOURCE FILE: 12.儲值金邀約-達10萬_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

# 模擬今天為 2025 年 6 月 20 日,如果超過20號, 需要回溯到20號
# mock_today = datetime(2025, 9, 20)
mock_today = datetime.today()

month_ago_2d = (mock_today - relativedelta(months=2) ).date().strftime("%Y/%m/%d")
month_ago_1d = (mock_today - relativedelta(months=1) ).date().strftime("%Y/%m/%d")
year_ago_1d = (mock_today - relativedelta(years=1) ).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((mock_today - relativedelta(months=3) ).date()).timestamp() * 1000
year_ago_1ts = pd.to_datetime((mock_today - relativedelta(years=1) ).date()).timestamp() * 1000

# 銷貨數據
#total_sales= kd.get_sap_with_relate_company(month_ago_2d)
total_sales = kd.get_data_from_MSSQL(f'''
        SELECT buyer as SAP公司代號, taxfree_basecurr as 未稅本位幣,  planned_shipping_date as 預計發貨日期 
        FROM sap_sales_data 
        WHERE  buyer LIKE 'TW%' 
            AND planned_shipping_date >= '{month_ago_2d}'  AND planned_shipping_date < '{mock_today}'
    ''')

total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

target_summary = total_sales.groupby('SAP公司代號')['未稅本位幣'].sum().reset_index()
target_summary = target_summary[target_summary['未稅本位幣'] / 2 >= 100000]



# 根據target_summary銷貨數據所在公司,並查找公司的各維度數據
all_company_ids = target_summary['SAP公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
        and SAP_CompanyID__c in {target_customer_str}
    '''

    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(target_summary, account_df, on = 'SAP公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')


# 剩餘儲值金大於0的都排除
stored_value_file = r"Z:\18_各部門共享區\03_台灣事業部\6.訊息公佈欄\1.每日帳款查詢"
stored_value_keyword = "儲值金餘額"
stored_value_path = kd.get_latest_excel(stored_value_file, stored_value_keyword)
stored_value_df = pd.read_excel(stored_value_path, sheet_name="業務觀看用", header=1)
stored_value_df = stored_value_df[stored_value_df['剩餘儲值金\n(含稅)']>0]
# 找到關聯公司, 排除---更改
stored_value_df = stored_value_df.rename(columns={'客戶編號': 'SAP公司代號'})
# stored_value_df = kd.add_relate_company(stored_value_df,"SAP")
K_invite = K_invite[~K_invite['SAP公司代號'].isin(set(stored_value_df['SAP公司代號']))]


# # 欄位"專案到期日"中的日期，大於當前日期時，排除
# project_file = r"Z:\18_各部門共享區\01_會計課\01_會計部報表\06_專案彙總表(製表人-銷管助理)"
# project_keyword = "專案價彙總表"
# project_path = kd.get_latest_excel(project_file, project_keyword)
# project_df = pd.read_excel(project_path, sheet_name="專案價總表")
# today = datetime.today().date()
# def parse_date_or_future(val):
#     val_str = str(val).strip()
#     if val_str == '無期限':
#         return date(2999, 12, 31)
#     try:
#         return pd.to_datetime(val_str).date()
#     except:
#         return None  
# project_df['專案到期日_parsed'] = project_df['專案到期日'].apply(parse_date_or_future)
# project_df = project_df[project_df['專案到期日_parsed'] > today].copy()
# project_df = pd.merge(project_df, company_map,left_on='客代', right_on='SAP公司代號', how='left')
# K_invite = K_invite[~K_invite['公司代號'].isin(set(stored_value_df['關聯公司']) )]


# 欄位"專案到期日"中的日期，大於當前日期時，排除
# 也就是只保留「專案已到期」的公司
today = datetime.today().date()
start_date = (today.replace(day=1) - relativedelta(months=1)).replace(day=21)
end_date = (today.replace(day=1) + relativedelta(months=1)).replace(day=1)
project_file = r"Z:\18_各部門共享區\01_會計課\01_會計部報表\06_專案彙總表(製表人-銷管助理)"
project_keyword = "專案價彙總表"
project_path = kd.get_latest_excel(project_file, project_keyword)
project_df = pd.read_excel(project_path, sheet_name="專案價總表")
today = datetime.today().date()
def parse_date_or_future(val):
    val_str = str(val).strip()
    if val_str == '無期限':
        return date(2999, 12, 31)
    try:
        return pd.to_datetime(val_str).date()
    except:
        return None  

project_df['專案到期日_parsed'] = project_df['專案到期日'].apply(parse_date_or_future)
future_project = project_df[project_df['專案到期日_parsed'] >= end_date].copy()
project_df = project_df[(project_df['專案到期日_parsed'] >= start_date) &(project_df['專案到期日_parsed'] < end_date)  ].copy()
project_df = project_df.rename(columns={'客代': 'SAP公司代號'})
# project_df = kd.add_relate_company(project_df,"SAP")



latest_project = project_df.sort_values('專案到期日_parsed', ascending=False).drop_duplicates('SAP公司代號')
K_invite = pd.merge(K_invite, latest_project[['SAP公司代號', '專案到期日_parsed']], on='SAP公司代號', how='left')
K_invite = K_invite.rename(columns={'專案到期日_parsed': '專案到期日'})
K_invite = K_invite[~K_invite['SAP公司代號'].isin(future_project['客代'])].copy()


# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts,merge_type='alone')
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite,data_type="stored_value")
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')

path = fr"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\儲值金\儲值金名單\{str((datetime.today()).date().strftime("%Y-%m"))}\\"
os.makedirs(path, exist_ok=True) # ← 若路徑不存在就建立（含所有父層）
K_invite.to_excel(path + "CD類儲值金客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')


# ==================================================
# SOURCE FILE: 13.儲值金邀約-當月3案場以上_wenbin.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

# 模擬今天為 2025 年 6 月 20 日,如果超過20號, 需要回溯到20號
# mock_today = datetime(2025, 9, 20)
mock_today = datetime.today()

month_ago_2d = (mock_today - relativedelta(months=2) ).date().strftime("%Y/%m/%d")
month_ago_1d = (mock_today - relativedelta(months=1) ).date().strftime("%Y/%m/%d")
year_ago_1d = (mock_today - relativedelta(years=1) ).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((mock_today - relativedelta(months=3) ).date()).timestamp() * 1000
year_ago_1ts = pd.to_datetime((mock_today - relativedelta(years=1) ).date()).timestamp() * 1000


# 銷貨數據
#total_sales= kd.get_sap_with_relate_company(month_ago_2d)
total_sales = kd.get_data_from_MSSQL(f'''
        SELECT buyer as SAP公司代號, taxfree_basecurr as 未稅本位幣,  planned_shipping_date as 預計發貨日期 
        FROM sap_sales_data 
        WHERE  buyer LIKE 'TW%' 
            AND planned_shipping_date >= '{month_ago_2d}'  AND planned_shipping_date < '{mock_today}'
    ''')

total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

target_summary = total_sales.groupby('SAP公司代號')['未稅本位幣'].sum().reset_index()
target_summary = target_summary[
    (target_summary['未稅本位幣'] / 2 <= 100000) & (target_summary['未稅本位幣'] / 2 >= 0)]



# 大於等於三個案場
project = kd.get_data_from_MSSQL(f'''
        SELECT 
            buyer AS SAP公司代號, 
            COUNT(DISTINCT delivery_party) AS 案場數
        FROM 
            [raw_data].[dbo].[sap_sales_data]
        WHERE 
            buyer LIKE 'TW%' 
            AND TRY_CAST(taxfree_basecurr AS FLOAT) > 0
            AND planned_shipping_date >= '{month_ago_2d}'  AND planned_shipping_date < '{mock_today}'
     AND NOT (
      (LEN(delivery_party) = 10 AND RIGHT(delivery_party, 3) IN ('001', '002', '003', '004', '005', '006', '007', '008', '009', '010'))
      OR LEN(delivery_party) = 7)
        GROUP BY buyer
    ''')

company_map = kd.get_data_from_MSSQL('''
            SELECT sap_company_id AS SAP公司代號,  company_id AS 公司代號 
            FROM [raw_data].[dbo].[crm_related_company]
            where sap_company_id is not null
        ''')
project = pd.merge(project[['SAP公司代號','案場數']], company_map, on = 'SAP公司代號', how = 'inner')
project_above_3 = project.groupby('SAP公司代號')['案場數'].sum().reset_index().query('案場數 >= 3')

target_summary = pd.merge(target_summary, project_above_3, on = 'SAP公司代號', how = 'inner')



# 根據target_summary銷貨數據所在公司,並查找公司的各維度數據
# 特定C類
all_company_ids = target_summary['SAP公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%C%'  or customItem199__c like '%D%' )
        and SAP_CompanyID__c in {target_customer_str}
    '''

    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(target_summary, account_df, on = 'SAP公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            '''
            )

K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

# 剩餘儲值金大於0的都排除
stored_value_file = r"Z:\18_各部門共享區\03_台灣事業部\6.訊息公佈欄\1.每日帳款查詢"
stored_value_keyword = "儲值金餘額"
stored_value_path = kd.get_latest_excel(stored_value_file, stored_value_keyword)
stored_value_df = pd.read_excel(stored_value_path, sheet_name="業務觀看用", header=1)
stored_value_df = stored_value_df[stored_value_df['剩餘儲值金\n(含稅)']>0]
# 找到關聯公司, 排除---更改, 不找關聯公司

stored_value_df = pd.merge(stored_value_df, company_map,left_on='客戶編號', right_on='SAP公司代號', how='left')
K_invite = K_invite[~K_invite['公司代號'].isin(set(stored_value_df['公司代號']) )]

output_path = r"C:\Users\TW0002.TPTWKD\Desktop\不滿10萬3案場CD.xlsx"
K_invite.to_excel(output_path, index=False)
today = mock_today.date()
start_date = (today.replace(day=1) - relativedelta(months=1)).replace(day=21)
end_date = (today.replace(day=1) + relativedelta(months=1)).replace(day=1)


project_file = r"Z:\18_各部門共享區\01_會計課\01_會計部報表\06_專案彙總表(製表人-銷管助理)"
project_keyword = "專案價彙總表"
project_path = kd.get_latest_excel(project_file, project_keyword)
project_df = pd.read_excel(project_path, sheet_name="專案價總表")
today = datetime.today().date()
def parse_date_or_future(val):
    val_str = str(val).strip()
    if val_str == '無期限':
        return date(2999, 12, 31)
    try:
        return pd.to_datetime(val_str).date()
    except:
        return None  

project_df['專案到期日_parsed'] = project_df['專案到期日'].apply(parse_date_or_future)
# project_df = project_df[(project_df['專案到期日_parsed'] >= start_date) &(project_df['專案到期日_parsed'] < end_date)  ].copy()
project_df = pd.merge(project_df, company_map, left_on='客代', right_on='SAP公司代號', how='left')
latest_project = project_df.sort_values('專案到期日_parsed', ascending=False).drop_duplicates('公司代號')
future_project = project_df[project_df['專案到期日_parsed'] >= end_date].copy()
K_invite = pd.merge(K_invite, latest_project[['公司代號', '專案到期日_parsed']], on='公司代號', how='left')
K_invite = K_invite.rename(columns={'專案到期日_parsed': '專案到期日'})
K_invite = K_invite[~K_invite['公司代號'].isin(future_project['公司代號'])].copy()


# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts,merge_type='alone')
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite = kd.clean_invalid_entries_visit(K_invite,data_type="stored_value")
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')


path = fr"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\儲值金\儲值金名單\{str((datetime.today()).date().strftime("%Y-%m"))}\\"
os.makedirs(path, exist_ok=True) # ← 若路徑不存在就建立（含所有父層）
K_invite.to_excel(path + "CD類儲值金未滿10萬客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')





# ==================================================
# SOURCE FILE: 14.儲值金數據標準化.py
# ==================================================


import pandas as pd
import numpy as np
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path
from openpyxl import load_workbook


path_old = fr"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\儲值金\儲值金名單\{str((datetime.today()+relativedelta(months=-1)).date().strftime("%Y-%m"))}\\"
path_old =  kd.get_latest_excel(path_old,'儲值金名單(模板)')

template_path = path_old 
dest_folder = fr"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\儲值金\儲值金名單\{datetime.today().strftime('%Y-%m')}\\"
os.makedirs(dest_folder, exist_ok=True)

# month_cn = f"{datetime.today().month}月"
# dest_path = os.path.join(dest_folder, f"{month_cn}儲值金名單(模板).xlsx")

# wb = load_workbook(template_path, data_only=False, keep_vba=False)

# for ws in wb.worksheets:
#     max_row, max_col = ws.max_row, ws.max_column
#     if max_row <= 1:
#         continue  
#     for r in range(2, max_row + 1):
#         for c in range(1, max_col + 1):
#             cell = ws.cell(row=r, column=c)
#             cell.value = None
#             cell.hyperlink = None
#             cell.comment = None
# wb.save(dest_path)
# print(f"已建立：{dest_path}")



# 獲取儲值金達10萬的名單
path_over10 =  kd.get_latest_excel(dest_folder,'CD類儲值金客戶')
over10 = pd.read_excel(path_over10,  dtype=str)
over10 = over10[['公司代號', '資料區域群組名稱', '公司型態', '專案到期日']]
over10['主旨'] = '(經營)儲值金邀約-達10萬'

# 獲取儲值金未達10萬的名單
path_below10 =  kd.get_latest_excel(dest_folder,'CD類儲值金未滿10萬客戶')
below10 = pd.read_excel(path_below10,  dtype=str)
below10 = below10[['公司代號', '資料區域群組名稱', '公司型態', '專案到期日']]
below10['主旨'] = '(經營)儲值金邀約-當月3案場以上'

# 合併兩個DataFrame
K_invite = pd.concat([over10, below10], ignore_index=True).drop_duplicates(['公司代號'])

# 獲取所有客戶資料
account = kd.get_data_from_CRM("""SELECT accountCode__c,accountName 公司名稱, dimDepart.departName 資料區域名稱 FROM account  where dimDepart.departName like '%TW%'  """)


# 合併客戶資料
K_invite_total = pd.merge(K_invite, account, left_on='公司代號', right_on='accountCode__c', how='left')

K_invite_total_ok = K_invite_total.loc[
    K_invite_total["資料區域群組名稱"].eq(K_invite_total["資料區域名稱"])
    & ~K_invite_total["資料區域群組名稱"].astype(str).str.contains("Z", case=False, na=False)
    & ~K_invite_total["資料區域名稱"].astype(str).str.contains("Z", case=False, na=False)]
   
K_invite_total_ok.to_excel(dest_folder + "儲值金數據留存_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)


K_invite_total_ok = (K_invite_total_ok [['公司代號','資料區域名稱','專案到期日','公司名稱']]
      .assign(簡稱=lambda d: d["公司名稱"].fillna("").str[:2],處理人員="", 處理進度="",異動時間="")
    [["公司代號","簡稱","資料區域名稱","專案到期日","處理人員","處理進度","異動時間","公司名稱"]])

K_invite_total_ok.to_clipboard(index=False, header=False, excel=True)
print("已將數據複製到剪貼簿，大人! 請去企微文檔中建立新的表格並將數據 粘貼到Excel中, 記得選擇保持原樣。")
# 點擊上個月sheet, 選擇創建副本,點擊確定, 修改名稱與到期日

# ==================================================
# SOURCE FILE: 20.菲律賓拜訪清單.py
# ==================================================



import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path
import re

month_ago_7d = (datetime.today() - relativedelta(months=7) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3d = (datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_5d = (datetime.today() - relativedelta(years=5) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000




##### 信用管制 #####
invalid_company = kd.get_data_from_MSSQL('select *   FROM [clean_data].[dbo].[crm_account_invalid] ')

##### 客戶 #####
select_query = f'''
SELECT accountCode__c 公司代號
, accountName 公司全名
, SAP_CompanyID__c sap公司代號
, dimDepart.departName 資料區域名稱
, customItem202__c 公司地址
, region__c.customItem9__c 是否主營
, customItem322__c 目標客戶類型
, customItem198__c.name 公司型態
, SAP_CompanyID__c sap公司代號
, customItem278__c 倒閉無效
, customItem291__c 勿擾選項
, createdAt 創建日期
FROM account
where dimDepart.departName like 'PH%'
'''
account = kd.get_data_from_CRM(select_query)
data_account = account.copy()

cols = [  '公司全名', '目標客戶類型', '倒閉無效', '勿擾選項', '是否主營']
data_account[cols] = data_account[cols].astype(str) .replace(['None', 'nan'], '')
kd.convert_to_date(data_account,'創建日期')
data_account = data_account.drop_duplicates('公司代號')

company_name = ['倒閉', '歇業', '停業', '轉行', '退休', '過世', '燈箱', '群組', '支援', '留守', '教育訓練', '無效拜訪', '資料不全', '搬遷', '廢止', 
                '解散', '管制', '非營業中', 'c>']

data_account = data_account.loc[(
    (~data_account['公司全名'].str.contains('|'.join(company_name), na=False)) &
    (~data_account['倒閉無效'].str.contains('是')) &
    (~data_account['資料區域名稱'].str.contains('INV|Others')) &
    ~data_account['勿擾選項'].str.contains('勿拜訪', na=False) &
    ~data_account['公司地址'].str.contains(r'\(x\)', na=False) &
     data_account['公司型態'].astype(str).str.contains(r'[CD]', na=False) 
         & (~data_account['是否主營'].str.contains('否'))
    )]

data_account = data_account.loc[~data_account['sap公司代號'].isin(invalid_company['company_id'])]

########## 客戶關係聯絡人 ##########
select_query = f'''
SELECT customItem8__c 公司代號
, customItem24__c 關係狀態
, customItem50__c 空號
, customItem51__c 停機
, customItem42__c 聯絡人資料無效
, customItem89__c 號碼錯誤非本人
, dimDepart.departName 資料區域名稱
, customItem95__c 職務類別
FROM customEntity22__c
where dimDepart.departName like 'PH%'
'''
rel_contact = kd.get_data_from_CRM((select_query))
data_rel_contact = rel_contact.copy()

def clean_str(s):
    return re.sub(r"[\[\]'\"\(\)]", "", s)

cols = ['關係狀態','號碼錯誤非本人','聯絡人資料無效','職務類別']
data_rel_contact[cols] = (  data_rel_contact[cols] .fillna('').astype(str) .applymap(clean_str))
data_rel_contact[['空號','停機']] = (  data_rel_contact[['空號','停機']] .fillna(0) .astype(int))

data_rel_contact = data_rel_contact.loc[
    data_rel_contact['關係狀態'].str.contains('在職') &
    (data_rel_contact['空號'] == 0) &
    (data_rel_contact['停機'] == 0) &
    (~data_rel_contact['聯絡人資料無效'].str.contains('是')) &
    (~data_rel_contact['號碼錯誤非本人'].str.contains('是'))]
data_rel_contact = data_rel_contact.drop_duplicates('公司代號')


########## 近兩個月拜訪 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
where visit_date >= '{two_months_ago}' 
and (region like 'PH-%')
and (customers_type = 'A1 拜訪' OR customers_type = 'C2 視訊拜訪')'''
track = kd.get_data_from_MSSQL(cont_query)
data_track = track.copy()

########## 近兩個月K大 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, region, visit_date, list_type, present_time
FROM clean_data.dbo.crm_K_3M 
WHERE region LIKE 'PH-%'
    AND list_type = 'K大預約表'
    AND visit_date >= '{two_months_ago}' 
    AND present_time >= 8'''
data_k_2M_orgi = kd.get_data_from_MSSQL(cont_query)
data_k_2M = data_k_2M_orgi.copy()

# ########## 近三個月電訪 ##########
# now = pd.Timestamp.now(tz="UTC")
# three_months_ago = now - pd.DateOffset(months=3)

# cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
# where visit_date >= '{three_months_ago}' 
# and (region like 'PH-%')
# and (customers_type = 'B2 電訪-無效')'''
# tel_track = kd.get_data_from_MSSQL(cont_query)
# tel_track = tel_track.copy()





# 近半年未叫貨  剔除 兩個月內 拜訪 K大
total_sales= kd.get_sap_with_relate_company(month_ago_7d,location = 'PH')
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
sales_month_7_to_6 = total_sales[(total_sales['預計發貨日期'] >= month_ago_7d) &(total_sales['預計發貨日期'] < month_ago_6d)]
sales_last_6_months = total_sales[total_sales['預計發貨日期'] >= month_ago_6d]
target_summary = sales_month_7_to_6.groupby('公司代號')['未稅本位幣'].sum().reset_index()
sap_ids_last6 = set(sales_last_6_months['公司代號'].unique())
sales_nobuy_after = target_summary[~target_summary['公司代號'].isin(sap_ids_last6)]

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
                                                             set(data_k_2M['company_id'].dropna()) )],   # K大
                                data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
K_invite = pd.merge(sales_nobuy_after, K_invite, on = '公司代號', how = 'inner')

K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"PH_近半年未叫貨_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')





# # 5年有交易  剔除 兩個月內 拜訪 K大
# total_sales= kd.get_sap_with_relate_company(year_ago_5d,location = 'PH')
# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

# target_contact = data_rel_contact[~data_rel_contact['關係狀態'].astype(str).str.contains("配合|離職")]
# target_contact = target_contact[target_contact['職務類別'].astype(str).str.contains("001|002|003|004|005|006|007|010|011|015", na=False)]
# target_contact['職務類別'] = target_contact['職務類別'].astype('string')
# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
#                                                              set(data_k_2M['company_id'].dropna()) )],   # K大
#                      target_contact[['公司代號','關係狀態','職務類別']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(total_sales, K_invite, on = '公司代號', how = 'inner')

# K_invite = (  K_invite .sort_values('職務類別', ascending=True) .drop_duplicates('公司代號', keep='first'))
# target = ['經營客戶']
# K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)].drop(columns=['未稅本位幣','未稅金額_台幣','預計發貨日期'], errors='ignore')

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"PH_5年有交易的開發_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')



# # 高交易額客戶二個月未拜訪   剔除 兩個月內 拜訪
# total_sales= kd.get_sap_with_relate_company(month_ago_6d,location = 'PH')
# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
# target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()
# target_summary = target_summary[target_summary['未稅本位幣'] > 9000]

# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) )],  # 拜訪
#                      data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')

# K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
# target = ['經營客戶']
# K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"PH_高交易額客戶二個月未拜訪_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')



# 曆史銷貨排行 剔除 兩個月內 拜訪 K大 交易
sales_in_2month = kd.get_sap_with_relate_company_os(month_ago_2d,location = 'PH')

total_sales= kd.get_sap_with_relate_company_os(year_ago_1d,location = 'PH')
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |    # 拜訪
                                                            set(data_k_2M['company_id'].dropna()) |        # K大
                                                            set(sales_in_2month['公司代號'].dropna())  )],  # 交易
                                    data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')

K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite = ( K_invite .sort_values('未稅本位幣', ascending=False) .groupby('資料區域名稱', as_index=False) .head(20)) # 每區前20

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"PH_曆史銷貨排行_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')



# ==================================================
# SOURCE FILE: 21.泰國拜訪清單.py
# ==================================================



import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path
import re

month_ago_7d = (datetime.today() - relativedelta(months=7) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3d = (datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_5d = (datetime.today() - relativedelta(years=5) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000




##### 信用管制 #####
invalid_company = kd.get_data_from_MSSQL('select *   FROM [clean_data].[dbo].[crm_account_invalid] ')

##### 客戶 #####
select_query = f'''
SELECT accountCode__c 公司代號
, accountName 公司全名
, SAP_CompanyID__c sap公司代號
, dimDepart.departName 資料區域名稱
, customItem202__c 公司地址
, region__c.customItem9__c 是否主營
, customItem322__c 目標客戶類型
, customItem198__c.name 公司型態
, SAP_CompanyID__c sap公司代號
, customItem278__c 倒閉無效
, customItem291__c 勿擾選項
, createdAt 創建日期
FROM account
where dimDepart.departName like 'TH%'
'''
account = kd.get_data_from_CRM(select_query)
data_account = account.copy()

cols = [  '公司全名', '目標客戶類型', '倒閉無效', '勿擾選項', '是否主營']
data_account[cols] = data_account[cols].astype(str) .replace(['None', 'nan'], '')
kd.convert_to_date(data_account,'創建日期')
data_account = data_account.drop_duplicates('公司代號')

company_name = ['倒閉', '歇業', '停業', '轉行', '退休', '過世', '燈箱', '群組', '支援', '留守', '教育訓練', '無效拜訪', '資料不全', '搬遷', '廢止', 
                '解散', '管制', '非營業中', 'c>']

data_account = data_account.loc[(
    (~data_account['公司全名'].str.contains('|'.join(company_name), na=False)) &
    (~data_account['倒閉無效'].str.contains('是')) &
    (~data_account['資料區域名稱'].str.contains('INV|Others')) &
    ~data_account['勿擾選項'].str.contains('勿拜訪', na=False) &
    ~data_account['公司地址'].str.contains(r'\(x\)', na=False) &
     data_account['公司型態'].astype(str).str.contains(r'[CD]', na=False) 
         & (~data_account['是否主營'].str.contains('否'))
    )]

data_account = data_account.loc[~data_account['sap公司代號'].isin(invalid_company['company_id'])]

########## 客戶關係聯絡人 ##########
select_query = f'''
SELECT customItem8__c 公司代號
, customItem24__c 關係狀態
, customItem50__c 空號
, customItem51__c 停機
, customItem42__c 聯絡人資料無效
, customItem89__c 號碼錯誤非本人
, dimDepart.departName 資料區域名稱
, customItem95__c 職務類別
FROM customEntity22__c
where dimDepart.departName like 'TH%'
'''
rel_contact = kd.get_data_from_CRM((select_query))
data_rel_contact = rel_contact.copy()

def clean_str(s):
    return re.sub(r"[\[\]'\"\(\)]", "", s)

cols = ['關係狀態','號碼錯誤非本人','聯絡人資料無效','職務類別']
data_rel_contact[cols] = (  data_rel_contact[cols] .fillna('').astype(str) .applymap(clean_str))
data_rel_contact[['空號','停機']] = (  data_rel_contact[['空號','停機']] .fillna(0) .astype(int))

data_rel_contact = data_rel_contact.loc[
    data_rel_contact['關係狀態'].str.contains('在職') &
    (data_rel_contact['空號'] == 0) &
    (data_rel_contact['停機'] == 0) &
    (~data_rel_contact['聯絡人資料無效'].str.contains('是')) &
    (~data_rel_contact['號碼錯誤非本人'].str.contains('是'))]
data_rel_contact = data_rel_contact.drop_duplicates('公司代號')


########## 近兩個月拜訪 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
where visit_date >= '{two_months_ago}' 
and (region like 'TH-%')
and (customers_type = 'A1 拜訪' OR customers_type = 'C2 視訊拜訪')'''
track = kd.get_data_from_MSSQL(cont_query)
data_track = track.copy()

########## 近兩個月K大 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, region, visit_date, list_type, present_time
FROM clean_data.dbo.crm_K_3M 
WHERE region LIKE 'TH-%'
    AND list_type = 'K大預約表'
    AND visit_date >= '{two_months_ago}' 
    AND present_time >= 8'''
data_k_2M_orgi = kd.get_data_from_MSSQL(cont_query)
data_k_2M = data_k_2M_orgi.copy()

# ########## 近三個月電訪 ##########
# now = pd.Timestamp.now(tz="UTC")
# three_months_ago = now - pd.DateOffset(months=3)

# cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
# where visit_date >= '{three_months_ago}' 
# and (region like 'TH-%')
# and (customers_type = 'B2 電訪-無效')'''
# tel_track = kd.get_data_from_MSSQL(cont_query)
# tel_track = tel_track.copy()





# 近半年未叫貨  剔除 兩個月內 拜訪 K大
total_sales= kd.get_sap_with_relate_company_os(month_ago_7d,location = 'TH')
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
sales_month_7_to_6 = total_sales[(total_sales['預計發貨日期'] >= month_ago_7d) &(total_sales['預計發貨日期'] < month_ago_6d)]
sales_last_6_months = total_sales[total_sales['預計發貨日期'] >= month_ago_6d]
target_summary = sales_month_7_to_6.groupby('公司代號')['未稅本位幣'].sum().reset_index()
sap_ids_last6 = set(sales_last_6_months['公司代號'].unique())
sales_nobuy_after = target_summary[~target_summary['公司代號'].isin(sap_ids_last6)]

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
                                                             set(data_k_2M['company_id'].dropna()) )],   # K大
                                data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
K_invite = pd.merge(sales_nobuy_after, K_invite, on = '公司代號', how = 'inner')

K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"TH_近半年未叫貨_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')





# # 5年有交易  剔除 兩個月內 拜訪 K大
# total_sales= kd.get_sap_with_relate_company(year_ago_5d,location = 'TH')
# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

# target_contact = data_rel_contact[~data_rel_contact['關係狀態'].astype(str).str.contains("配合|離職")]
# target_contact = target_contact[target_contact['職務類別'].astype(str).str.contains("001|002|003|004|005|006|007|010|011|015", na=False)]
# target_contact['職務類別'] = target_contact['職務類別'].astype('string')
# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
#                                                              set(data_k_2M['company_id'].dropna()) )],   # K大
#                      target_contact[['公司代號','關係狀態','職務類別']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(total_sales, K_invite, on = '公司代號', how = 'inner')

# K_invite = (  K_invite .sort_values('職務類別', ascending=True) .drop_duplicates('公司代號', keep='first'))
# target = ['經營客戶']
# K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)].drop(columns=['未稅本位幣','未稅金額_台幣','預計發貨日期'], errors='ignore')

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"TH_5年有交易的開發_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')



# # 高交易額客戶二個月未拜訪   剔除 兩個月內 拜訪
# total_sales= kd.get_sap_with_relate_company(month_ago_6d,location = 'TH')
# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
# target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()
# target_summary = target_summary[target_summary['未稅本位幣'] > 9000]

# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) )],  # 拜訪
#                      data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')

# K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
# target = ['經營客戶']
# K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"TH_高交易額客戶二個月未拜訪_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')



# 曆史銷貨排行 剔除 兩個月內 拜訪 K大 交易
sales_in_2month = kd.get_sap_with_relate_company_os(month_ago_2d,location = 'TH')

total_sales= kd.get_sap_with_relate_company_os(year_ago_1d,location = 'TH')
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |    # 拜訪
                                                            set(data_k_2M['company_id'].dropna()) |        # K大
                                                            set(sales_in_2month['公司代號'].dropna())  )],  # 交易
                                    data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')

K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite = ( K_invite .sort_values('未稅本位幣', ascending=False) .groupby('資料區域名稱', as_index=False) .head(20)) # 每區前20

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"TH_曆史銷貨排行_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')



# ==================================================
# SOURCE FILE: 22.新加坡拜訪清單.py
# ==================================================



import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path
import re

month_ago_7d = (datetime.today() - relativedelta(months=7) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3d = (datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_5d = (datetime.today() - relativedelta(years=5) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000




##### 信用管制 #####
invalid_company = kd.get_data_from_MSSQL('select *   FROM [clean_data].[dbo].[crm_account_invalid] ')

##### 客戶 #####
select_query = f'''
SELECT accountCode__c 公司代號
, accountName 公司全名
, SAP_CompanyID__c sap公司代號
, dimDepart.departName 資料區域名稱
, customItem202__c 公司地址
, region__c.customItem9__c 是否主營
, customItem322__c 目標客戶類型
, customItem198__c.name 公司型態
, SAP_CompanyID__c sap公司代號
, customItem278__c 倒閉無效
, customItem291__c 勿擾選項
, createdAt 創建日期
FROM account
where dimDepart.departName like 'SG%'
'''
account = kd.get_data_from_CRM(select_query)
data_account = account.copy()

cols = [  '公司全名', '目標客戶類型', '倒閉無效', '勿擾選項', '是否主營']
data_account[cols] = data_account[cols].astype(str) .replace(['None', 'nan'], '')
kd.convert_to_date(data_account,'創建日期')
data_account = data_account.drop_duplicates('公司代號')

company_name = ['倒閉', '歇業', '停業', '轉行', '退休', '過世', '燈箱', '群組', '支援', '留守', '教育訓練', '無效拜訪', '資料不全', '搬遷', '廢止', 
                '解散', '管制', '非營業中', 'c>']

data_account = data_account.loc[
    (~data_account['公司全名'].str.contains('|'.join(company_name), na=False)) &
    (~data_account['倒閉無效'].str.contains('是')) &
    (~data_account['資料區域名稱'].str.contains('INV|Others')) &
    data_account['資料區域名稱'].isin(['SG-A1', 'SG-A2', 'SG-A3', 'SG-A4', 'SG-A5']) &
    ~data_account['勿擾選項'].str.contains('勿拜訪', na=False) &
    ~data_account['公司地址'].str.contains(r'\(x\)', na=False) &
     data_account['公司型態'].astype(str).str.contains(r'[CD]', na=False) &
    ~( data_account['資料區域名稱'].isin(['SG-A1', 'SG-A2', 'SG-A3', 'SG-A4']) &
     data_account['公司地址'].str.contains(r'JOHOR|Johor|johor', na=False) 
    # & (~data_account['是否主營'].str.contains('否'))
    )]

data_account = data_account.loc[~data_account['sap公司代號'].isin(invalid_company['company_id'])]

########## 客戶關係聯絡人 ##########
select_query = f'''
SELECT customItem8__c 公司代號
, customItem24__c 關係狀態
, customItem50__c 空號
, customItem51__c 停機
, customItem42__c 聯絡人資料無效
, customItem89__c 號碼錯誤非本人
, dimDepart.departName 資料區域名稱
, customItem95__c 職務類別
FROM customEntity22__c
where dimDepart.departName like 'SG%'
'''
rel_contact = kd.get_data_from_CRM((select_query))
data_rel_contact = rel_contact.copy()


def clean_str(s):
    return re.sub(r"[\[\]'\"\(\)]", "", s)

cols = ['關係狀態','號碼錯誤非本人','聯絡人資料無效','職務類別']
data_rel_contact[cols] = (  data_rel_contact[cols] .fillna('').astype(str) .applymap(clean_str))
data_rel_contact[['空號','停機']] = (  data_rel_contact[['空號','停機']] .fillna(0) .astype(int))

data_rel_contact = data_rel_contact.loc[
    data_rel_contact['關係狀態'].str.contains('在職') &
    (data_rel_contact['空號'] == 0) &
    (data_rel_contact['停機'] == 0) &
    (~data_rel_contact['聯絡人資料無效'].str.contains('是')) &
    (~data_rel_contact['號碼錯誤非本人'].str.contains('是'))]

data_rel_contact = data_rel_contact.drop_duplicates('公司代號')




########## 近兩個月拜訪 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
where visit_date >= '{two_months_ago}' 
and (region like 'SG-%')
and (customers_type = 'A1 拜訪' OR customers_type = 'C2 視訊拜訪')'''
track = kd.get_data_from_MSSQL(cont_query)
data_track = track.copy()

########## 近兩個月K大 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, region, visit_date, list_type, present_time
FROM clean_data.dbo.crm_K_3M 
WHERE region LIKE 'SG-%'
    AND list_type = 'K大預約表'
    AND visit_date >= '{two_months_ago}' 
    AND present_time >= 8'''
data_k_2M_orgi = kd.get_data_from_MSSQL(cont_query)
data_k_2M = data_k_2M_orgi.copy()

# ########## 近三個月電訪 ##########
# now = pd.Timestamp.now(tz="UTC")
# three_months_ago = now - pd.DateOffset(months=3)

# cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
# where visit_date >= '{three_months_ago}' 
# and (region like 'SG-%')
# and (customers_type = 'B2 電訪-無效')'''
# tel_track = kd.get_data_from_MSSQL(cont_query)
# tel_track = tel_track.copy()





# 近半年未叫貨  剔除 兩個月內 拜訪 K大
total_sales_SG= kd.get_sap_with_relate_company_os(month_ago_7d,location = 'SG')
total_sales_MY= kd.get_sap_with_relate_company_os(month_ago_7d,location = 'MY')
total_sales_KL= kd.get_sap_with_relate_company_os(month_ago_7d,location = 'KL')
total_sales = pd.concat([total_sales_SG, total_sales_MY,total_sales_KL], ignore_index=True)

total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
sales_month_7_to_6 = total_sales[(total_sales['預計發貨日期'] >= month_ago_7d) &(total_sales['預計發貨日期'] < month_ago_6d)]
sales_last_6_months = total_sales[total_sales['預計發貨日期'] >= month_ago_6d]
target_summary = sales_month_7_to_6.groupby('公司代號')['未稅金額_台幣'].sum().reset_index()
sap_ids_last6 = set(sales_last_6_months['公司代號'].unique())
sales_nobuy_after = target_summary[~target_summary['公司代號'].isin(sap_ids_last6)]

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
                                                             set(data_k_2M['company_id'].dropna()) )],   # K大
                                data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
K_invite = pd.merge(sales_nobuy_after, K_invite, on = '公司代號', how = 'inner')

K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"SG_近半年未叫貨_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')





# # 5年有交易  剔除 兩個月內 拜訪 K大
# total_sales_SG= kd.get_sap_with_relate_company_os(year_ago_5d,location = 'SG')
# total_sales_MY= kd.get_sap_with_relate_company_os(year_ago_5d,location = 'MY')
# total_sales_KL= kd.get_sap_with_relate_company_os(year_ago_5d,location = 'KL')
# total_sales = pd.concat([total_sales_SG, total_sales_MY,total_sales_KL], ignore_index=True)

# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

# target_contact = data_rel_contact[~data_rel_contact['關係狀態'].astype(str).str.contains("配合|離職")]
# target_contact = target_contact[target_contact['職務類別'].astype(str).str.contains("001|002|003|004|005|006|007|010|011|015", na=False)]
# target_contact['職務類別'] = target_contact['職務類別'].astype('string')
# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
#                                                              set(data_k_2M['company_id'].dropna()) )],   # K大
#                      target_contact[['公司代號','關係狀態','職務類別']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(total_sales, K_invite, on = '公司代號', how = 'inner')

# K_invite = (  K_invite .sort_values('職務類別', ascending=True) .drop_duplicates('公司代號', keep='first'))
# target = ['經營客戶']
# K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)].drop(columns=['未稅本位幣','未稅金額_台幣','預計發貨日期'], errors='ignore')

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"SG_5年有交易的開發_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')














# 高交易額客戶二個月未拜訪   剔除 兩個月內 拜訪
total_sales_SG= kd.get_sap_with_relate_company_os(month_ago_6d,location = 'SG')
total_sales_MY= kd.get_sap_with_relate_company_os(month_ago_6d,location = 'MY')
total_sales_KL= kd.get_sap_with_relate_company_os(month_ago_6d,location = 'KL')
total_sales = pd.concat([total_sales_SG, total_sales_MY,total_sales_KL], ignore_index=True)

total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
target_summary = total_sales.groupby('公司代號')['未稅金額_台幣'].sum().reset_index()
target_summary = target_summary[target_summary['未稅金額_台幣'] > 500000]

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) )],  # 拜訪
                     data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')

K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"SG_高交易額客戶二個月未拜訪_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')



# 曆史銷貨排行 剔除 兩個月內 拜訪 K大 交易
sales_in_2month_SG = kd.get_sap_with_relate_company_os(month_ago_2d,location = 'SG')
sales_in_2month_MY = kd.get_sap_with_relate_company_os(month_ago_2d,location = 'MY')
sales_in_2month_KL = kd.get_sap_with_relate_company_os(month_ago_2d,location = 'KL')
sales_in_2month = pd.concat([sales_in_2month_SG, sales_in_2month_MY,sales_in_2month_KL], ignore_index=True)


total_sales_SG= kd.get_sap_with_relate_company_os(year_ago_1d,location = 'SG')
total_sales_MY= kd.get_sap_with_relate_company_os(year_ago_1d,location = 'MY')
total_sales_KL= kd.get_sap_with_relate_company_os(year_ago_1d,location = 'KL')
total_sales = pd.concat([total_sales_SG, total_sales_MY,total_sales_KL], ignore_index=True)
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
target_summary = total_sales.groupby('公司代號')['未稅金額_台幣'].sum().reset_index()

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |    # 拜訪
                                                            set(data_k_2M['company_id'].dropna()) |        # K大
                                                            set(sales_in_2month['公司代號'].dropna())  )],  # 交易
                                    data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')

K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite = ( K_invite .sort_values('未稅金額_台幣', ascending=False) .groupby('資料區域名稱', as_index=False) .head(20)) # 每區前20

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"SG_曆史銷貨排行_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')



# ==================================================
# SOURCE FILE: 23.越南拜訪清單.py
# ==================================================



import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path
import re

month_ago_7d = (datetime.today() - relativedelta(months=7) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_6d = (datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3d = (datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_5d = (datetime.today() - relativedelta(years=5) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_6ts = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000




##### 信用管制 #####
invalid_company = kd.get_data_from_MSSQL('select *   FROM [clean_data].[dbo].[crm_account_invalid] ')

##### 客戶 #####
select_query = f'''
SELECT accountCode__c 公司代號
, accountName 公司全名
, SAP_CompanyID__c sap公司代號
, dimDepart.departName 資料區域名稱
, customItem202__c 公司地址
, region__c.customItem9__c 是否主營
, customItem322__c 目標客戶類型
, customItem198__c.name 公司型態
, SAP_CompanyID__c sap公司代號
, customItem278__c 倒閉無效
, customItem291__c 勿擾選項
, createdAt 創建日期
FROM account
where dimDepart.departName like 'VN%'
'''
account = kd.get_data_from_CRM(select_query)
data_account = account.copy()

cols = [  '公司全名', '目標客戶類型', '倒閉無效', '勿擾選項', '是否主營']
data_account[cols] = data_account[cols].astype(str) .replace(['None', 'nan'], '')
kd.convert_to_date(data_account,'創建日期')
data_account = data_account.drop_duplicates('公司代號')

company_name = ['倒閉', '歇業', '停業', '轉行', '退休', '過世', '燈箱', '群組', '支援', '留守', '教育訓練', '無效拜訪', '資料不全', '搬遷', '廢止', 
                '解散', '管制', '非營業中', 'c>']

data_account = data_account.loc[(
    (~data_account['公司全名'].str.contains('|'.join(company_name), na=False)) &
    (~data_account['倒閉無效'].str.contains('是')) &
    (~data_account['資料區域名稱'].str.contains('INV|Others')) &
    ~data_account['勿擾選項'].str.contains('勿拜訪', na=False) &
    ~data_account['公司地址'].str.contains(r'\(x\)', na=False) &
     data_account['公司型態'].astype(str).str.contains(r'[CD]', na=False) 
         & (~data_account['是否主營'].str.contains('否'))
    )]
data_account = data_account.loc[~data_account['sap公司代號'].isin(invalid_company['company_id'])]

########## 客戶關係聯絡人 ##########
select_query = f'''
SELECT customItem8__c 公司代號
, customItem24__c 關係狀態
, customItem50__c 空號
, customItem51__c 停機
, customItem42__c 聯絡人資料無效
, customItem89__c 號碼錯誤非本人
, dimDepart.departName 資料區域名稱
, customItem95__c 職務類別
FROM customEntity22__c
where dimDepart.departName like 'VN%'
'''
rel_contact = kd.get_data_from_CRM((select_query))
data_rel_contact = rel_contact.copy()

def clean_str(s):
    return re.sub(r"[\[\]'\"\(\)]", "", s)

cols = ['關係狀態','號碼錯誤非本人','聯絡人資料無效','職務類別']
data_rel_contact[cols] = (  data_rel_contact[cols] .fillna('').astype(str) .applymap(clean_str))
data_rel_contact[['空號','停機']] = (  data_rel_contact[['空號','停機']] .fillna(0) .astype(int))

data_rel_contact = data_rel_contact.loc[
    data_rel_contact['關係狀態'].str.contains('在職') &
    (data_rel_contact['空號'] == 0) &
    (data_rel_contact['停機'] == 0) &
    (~data_rel_contact['聯絡人資料無效'].str.contains('是')) &
    (~data_rel_contact['號碼錯誤非本人'].str.contains('是'))]
data_rel_contact = data_rel_contact.drop_duplicates('公司代號')


########## 近兩個月拜訪 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
where visit_date >= '{two_months_ago}' 
and (region like 'VN-%')
and (customers_type = 'A1 拜訪' OR customers_type = 'C2 視訊拜訪')'''
track = kd.get_data_from_MSSQL(cont_query)
data_track = track.copy()

########## 近兩個月K大 ##########
now = pd.Timestamp.now(tz="UTC")
two_months_ago = now - pd.DateOffset(months=2)

cont_query = f''' SELECT company_id, region, visit_date, list_type, present_time
FROM clean_data.dbo.crm_K_3M 
WHERE region LIKE 'VN-%'
    AND list_type = 'K大預約表'
    AND visit_date >= '{two_months_ago}' 
    AND present_time >= 8'''
data_k_2M_orgi = kd.get_data_from_MSSQL(cont_query)
data_k_2M = data_k_2M_orgi.copy()

# ########## 近三個月電訪 ##########
# now = pd.Timestamp.now(tz="UTC")
# three_months_ago = now - pd.DateOffset(months=3)

# cont_query = f''' SELECT company_id, customers_type, region, visit_date FROM clean_data.dbo.crm_track_1year 
# where visit_date >= '{three_months_ago}' 
# and (region like 'VN-%')
# and (customers_type = 'B2 電訪-無效')'''
# tel_track = kd.get_data_from_MSSQL(cont_query)
# tel_track = tel_track.copy()





# 近半年未叫貨  剔除 兩個月內 拜訪 K大
total_sales= kd.get_sap_with_relate_company(month_ago_7d,location = 'VN')
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
sales_month_7_to_6 = total_sales[(total_sales['預計發貨日期'] >= month_ago_7d) &(total_sales['預計發貨日期'] < month_ago_6d)]
sales_last_6_months = total_sales[total_sales['預計發貨日期'] >= month_ago_6d]
target_summary = sales_month_7_to_6.groupby('公司代號')['未稅本位幣'].sum().reset_index()
sap_ids_last6 = set(sales_last_6_months['公司代號'].unique())
sales_nobuy_after = target_summary[~target_summary['公司代號'].isin(sap_ids_last6)]

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
                                                             set(data_k_2M['company_id'].dropna()) )],   # K大
                                data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
K_invite = pd.merge(sales_nobuy_after, K_invite, on = '公司代號', how = 'inner')

K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"VN_近半年未叫貨_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')





# # 5年有交易  剔除 兩個月內 拜訪 K大
# total_sales= kd.get_sap_with_relate_company(year_ago_5d,location = 'VN')
# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])

# target_contact = data_rel_contact[~data_rel_contact['關係狀態'].astype(str).str.contains("配合|離職")]
# target_contact = target_contact[target_contact['職務類別'].astype(str).str.contains("001|002|003|004|005|006|007|010|011|015", na=False)]
# target_contact['職務類別'] = target_contact['職務類別'].astype('string')
# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |  # 拜訪
#                                                              set(data_k_2M['company_id'].dropna()) )],   # K大
#                      target_contact[['公司代號','關係狀態','職務類別']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(total_sales, K_invite, on = '公司代號', how = 'inner')

# K_invite = (  K_invite .sort_values('職務類別', ascending=True) .drop_duplicates('公司代號', keep='first'))
# target = ['經營客戶']
# K_invite = K_invite[~K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)].drop(columns=['未稅本位幣','未稅金額_台幣','預計發貨日期'], errors='ignore')

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"VN_5年有交易的開發_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')



# # 高交易額客戶二個月未拜訪   剔除 兩個月內 拜訪
# total_sales= kd.get_sap_with_relate_company(month_ago_6d,location = 'VN')
# total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
# target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()
# target_summary = target_summary[target_summary['未稅本位幣'] > 9000]

# K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) )],  # 拜訪
#                      data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')
# K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')

# K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
# target = ['經營客戶']
# K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]

# folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
# filename = f"VN_高交易額客戶二個月未拜訪_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
# K_invite.to_excel(folder / filename, index=False)
# print(f'{filename}數據導出成功~')



# 曆史銷貨排行 剔除 兩個月內 拜訪 K大 交易
sales_in_2month = kd.get_sap_with_relate_company_os(month_ago_2d,location = 'VN')

total_sales= kd.get_sap_with_relate_company_os(year_ago_1d,location = 'VN')
total_sales['預計發貨日期'] = pd.to_datetime(total_sales['預計發貨日期'])
target_summary = total_sales.groupby('公司代號')['未稅本位幣'].sum().reset_index()

K_invite = pd.merge(data_account[~data_account['公司代號'].isin(set(data_track['company_id'].dropna()) |    # 拜訪
                                                            set(data_k_2M['company_id'].dropna()) |        # K大
                                                            set(sales_in_2month['公司代號'].dropna())  )],  # 交易
                                    data_rel_contact[['公司代號','關係狀態']], on = '公司代號', how = 'inner')

K_invite = pd.merge(target_summary, K_invite, on = '公司代號', how = 'inner')
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
target = ['經營客戶']
K_invite = K_invite[K_invite['目標客戶類型'].astype(str).str.contains('|'.join(target), na=False)]
K_invite = ( K_invite .sort_values('未稅本位幣', ascending=False) .groupby('資料區域名稱', as_index=False) .head(20)) # 每區前20

folder = Path(f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today() + relativedelta(months=1)).strftime('%Y.%m')}/raw")
folder.mkdir(parents=True, exist_ok=True)
filename = f"VN_曆史銷貨排行_{datetime.today().strftime('%Y.%m.%d')}.xlsx"
K_invite.to_excel(folder / filename, index=False)
print(f'{filename}數據導出成功~')



# ==================================================
# SOURCE FILE: 31專案_久未聯繫F類客戶.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


# 優先選擇關聯公司
f_company = kd.get_data_from_CRM(f'''
        select accountCode__c 公司代號
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%FA%' or customItem199__c like '%FB%' or customItem199__c like '%FD%')
''')
f_company = kd.merge_company_to_parent(f_company)


# 根據account所在公司,並查找公司的各維度數據
all_company_ids = f_company['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%FA%' or customItem199__c like '%FB%' or customItem199__c like '%FD%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(f_company, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts,"拜訪")
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite,excluded_df = kd.clean_invalid_entries_project(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')
K_invite['公司型態'].value_counts()
K_invite['主旨'] = ' 拜訪3:久未聯繫F類客戶'

path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/專案/"
K_invite.to_excel(path + "久未聯繫F類客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')








# ==================================================
# SOURCE FILE: 32專案_久未聯繫G類客戶.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


# 優先選擇關聯公司
se_company = kd.get_data_from_CRM(f'''
        select accountCode__c 公司代號
        from account
        where dimDepart.departName like '%TW%'  and (customItem199__c like '%GD%' or customItem199__c like '%GZ%')
''')
se_company = kd.merge_company_to_parent(se_company)


# 根據account所在公司,並查找公司的各維度數據
all_company_ids = se_company['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%GD%' or customItem199__c like '%GZ%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(se_company, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts,"拜訪")
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite,excluded_df = kd.clean_invalid_entries_project(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')
K_invite['主旨'] = ' 拜訪6:久未聯繫G類客戶'


path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/專案/"
K_invite.to_excel(path + "久未聯繫G類客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')








# ==================================================
# SOURCE FILE: 33專案_久未聯繫K類客戶.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000



# 優先選擇關聯公司
kz_company = kd.get_data_from_CRM(f'''
        select accountCode__c 公司代號
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%KZ%' or customItem199__c like '%KD%')
''')
kz_company = kd.merge_company_to_parent(kz_company)


# 根據account所在公司,並查找公司的各維度數據
all_company_ids = kz_company['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%KZ%' or customItem199__c like '%KD%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(kz_company, account_df, on = '公司代號', how = 'inner')





# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts,"拜訪")
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite,excluded_df = kd.clean_invalid_entries_project(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')
K_invite['主旨'] = ' 拜訪5:久未聯繫K類客戶'

path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{(datetime.today()+relativedelta(months=1)).strftime('%Y.%m')}/專案/"

os.makedirs(path, exist_ok=True)  # ← 若路徑不存在就建立（含所有父層）

K_invite.to_excel(path + f"久未聯繫K類客戶_{datetime.today().strftime('%Y.%m.%d')}.xlsx", index=False)
print("數據導出成功~")


# ==================================================
# SOURCE FILE: 34專案_久未聯繫SE類客戶.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


# 優先選擇關聯公司
se_company = kd.get_data_from_CRM(f'''
        select accountCode__c 公司代號
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%SE%')
''')
se_company = kd.merge_company_to_parent(se_company)


# 根據account所在公司,並查找公司的各維度數據
all_company_ids = se_company['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%SE%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(se_company, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts,"拜訪")
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite,excluded_df = kd.clean_invalid_entries_project(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')
K_invite['主旨'] = ' 拜訪4:久未聯繫SE類客戶'


path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/專案/"
K_invite.to_excel(path + "久未聯繫SE類客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')








# ==================================================
# SOURCE FILE: 35專案_久未聯繫SF類客戶.py
# ==================================================


import pandas as pd
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path

month_ago_2d = (datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_1d = (datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
year_ago_1d = (datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date().strftime("%Y/%m/%d")
month_ago_3ts = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_2ts = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
year_ago_1ts = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000


# 優先選擇關聯公司
se_company = kd.get_data_from_CRM(f'''
        select accountCode__c 公司代號
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%SF%')
''')
se_company = kd.merge_company_to_parent(se_company)


# 根據account所在公司,並查找公司的各維度數據
all_company_ids = se_company['公司代號'].dropna().unique().tolist()
batch_size = 100
account_df = pd.DataFrame()

for i in range(0, len(all_company_ids), batch_size):
    batch_ids = all_company_ids[i:i + batch_size]
    target_customer_str = "(" + ",".join(f"'{c}'" for c in batch_ids) + ")"

    xoql = f'''
        select accountCode__c 公司代號, customItem202__c 公司地址, dimDepart.departName 資料區域群組名稱,
        customItem197__c 公司簡稱, SAP_CompanyID__c SAP公司代號, accountName 公司名稱, customItem322__c 目標客戶類型,
        customItem199__c 公司型態, customItem291__c 公司勿擾選項,  Phone__c 公司電話, customItem278__c 倒閉
        from account
        where dimDepart.departName like '%TW%' and (customItem199__c like '%SF%')
        and accountCode__c in {target_customer_str}
    '''
    try:
        df = kd.get_data_from_CRM(xoql)
        account_df = pd.concat([account_df, df], ignore_index=True, sort=False)
    except Exception as e:
        print(f"Error on batch {i // batch_size + 1}: {e}")

K_invite = pd.merge(se_company, account_df, on = '公司代號', how = 'inner')


# select related_contact 客關連數據
contact_related = kd.get_data_from_CRM(
            f'''
            select name 客戶關係聯絡人代號, customItem2__c.contactName 連絡人, contactCode__c__c 連絡人代號, 
            customItem8__c 公司代號,contactPhone__c__c 手機號碼,
            id 客戶關係連絡人 ,customItem74__c LINEID,customItem95__c 職務類別, customItem109__c 聯絡人勿擾選項, 
            customItem42__c 連絡人資料無效, customItem24__c 關係狀態 ,customItem50__c 空號,
            customItem51__c 停機,customItem52__c 號碼錯誤非本人
            from customEntity22__c 
            where customItem37__c  like '%TW%'
            ''')
K_invite = pd.merge(K_invite, contact_related, on = '公司代號', how = 'inner')

#   排除近3個月拜訪過
pass_visited =  kd.last_connected(month_ago_3ts,"拜訪")
K_invite = K_invite[~K_invite['公司代號'].isin(set(pass_visited['公司代號']) )]

# 獲取最近聯係日期
last_connected = kd.last_connected(year_ago_1ts)
K_invite = pd.merge(K_invite, last_connected, on = '公司代號', how = 'left')

# 數據清洗
K_invite,excluded_df = kd.clean_invalid_entries_project(K_invite)
K_invite = K_invite.drop_duplicates('公司代號',keep= 'first')
K_invite['公司代號'].value_counts()
K_invite = K_invite.sort_values(by='日期', ascending=True, na_position='first')
K_invite['主旨'] = ' 拜訪2:久未聯繫SF類客戶'


path = f"Z:/02_台灣事業部/1.北區/13.業務管理組/CRM共用資料夾/業助工作資料/業助資料夾/●皓皓●/★每月業務電拜訪★/共用CRM資訊/Temp/電拜訪清單資料/{str((datetime.today()+relativedelta(months=1)).date().strftime("%Y.%m"))}/專案/"
K_invite.to_excel(path + "久未聯繫SF類客戶_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)
print('數據導出成功~')








# ==================================================
# SOURCE FILE: 99.清單執行.py
# ==================================================


import pandas as pd
import subprocess
import sys
import os
import time
from datetime import datetime
from zoneinfo import ZoneInfo

TZ_TW = ZoneInfo("Asia/Taipei")

# ================= 基本設定 =================
BASE_PATH = r"C:\Users\TW0002.TPTWKD\Desktop\Wenbin\外勤业务\拜訪清單_更新"
EXCEL_PATH = os.path.join(BASE_PATH, "py檔案清單.xlsx")
LOG_DIR = os.path.join(BASE_PATH, "Log")
os.makedirs(LOG_DIR, exist_ok=True)

# 要執行的分類
# RUN_CATEGORIES = ["台灣拜訪", "台灣儲值金",  "台灣專案", "海外拜訪",]
RUN_CATEGORIES = ["台灣拜訪", "台灣專案",]

SLEEP_SECONDS = 5
MAX_RETRY = 3

# ================= Log 設定 =================
today_str = datetime.now(TZ_TW).strftime("%Y%m%d")
log_path = os.path.join(LOG_DIR, f"run_log_{today_str}.log")

def write_log(msg: str):
    ts = datetime.now(TZ_TW).strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")

# ================= 讀取設定檔 =================
df = pd.read_excel(EXCEL_PATH)

# ================= 預計執行清單 =================
to_run = []

for _, row in df.iterrows():
    file_path = row.get("完整路徑")
    is_run = bool(row.get("是否執行", False))
    category = row.get("分類")

    if (
        category in RUN_CATEGORIES
        and is_run
        and isinstance(file_path, str)
        and os.path.exists(file_path)
    ):
        to_run.append(file_path)

print("=== 本次預計執行的 py 檔案 ===")
for i, p in enumerate(to_run, 1):
    print(f"[{i}/{len(to_run)}] {p}")
print("================================")

write_log("=== 本次執行開始 ===")

# ================= 主流程 =================
total = len(to_run)
current = 0

for _, row in df.iterrows():
    file_path = row.get("完整路徑")
    is_run = bool(row.get("是否執行", False))
    category = row.get("分類")

    if category not in RUN_CATEGORIES or not is_run:
        continue

    if not isinstance(file_path, str) or not os.path.exists(file_path):
        write_log(f"[NOT FOUND] 分類={category} | {file_path}")
        continue

    current += 1
    print(f"\n>>> 執行進度 {current}/{total}")
    print(f">>> 分類: {category}")
    print(f">>> 檔案: {file_path}")

    for attempt in range(1, MAX_RETRY + 1):
        print(f">>> 嘗試第 {attempt} 次")
        start_time = time.time()
        write_log(f"[START] 分類={category} | 檔案={file_path} | 第 {attempt} 次")

        try:
            result = subprocess.run(
                [sys.executable, file_path],
                text=True,
                capture_output=True
            )

            if result.returncode != 0:
                raise subprocess.CalledProcessError(
                    result.returncode,
                    result.args,
                    output=result.stdout,
                    stderr=result.stderr
                )

            duration = round(time.time() - start_time, 2)
            print(f">>> 成功 完成 {duration}s")
            write_log(
                f"[SUCCESS] 分類={category} | 檔案={file_path} | 耗時={duration}s"
            )
            break

        except subprocess.CalledProcessError as e:
            duration = round(time.time() - start_time, 2)
            stdout = (e.output or "").strip()
            stderr = (e.stderr or "").strip()

            print(f">>> 失敗 耗時={duration}s")

            write_log(
                f"[FAIL] 分類={category} | 檔案={file_path} | "
                f"第 {attempt} 次失敗 | 耗時={duration}s\n"
                f"===== STDOUT =====\n{stdout}\n"
                f"===== STDERR =====\n{stderr}\n"
                f"=================="
            )

            if attempt == MAX_RETRY:
                print(">>> 連續失敗 放棄此檔案")
                write_log(
                    f"[GIVE UP] 分類={category} | 檔案={file_path} | 連續失敗 {MAX_RETRY} 次"
                )

        time.sleep(SLEEP_SECONDS)

    time.sleep(SLEEP_SECONDS)

write_log("=== 本次執行結束 ===")







# import os
# import pandas as pd

# # 目標資料夾
# base_path = r"C:\Users\TW0002.TPTWKD\Desktop\Wenbin\外勤业务\拜訪清單_更新"

# rows = []

# for root, _, files in os.walk(base_path):
#     for file in files:
#         if file.lower().endswith(".py"):
#             full_path = os.path.join(root, file)
#             rows.append({
#                 "檔名": file,
#                 "完整路徑": full_path,
#                 "是否執行": True
#             })

# df = pd.DataFrame(rows)

# # 輸出 Excel
# output_path = os.path.join(base_path, "py檔案清單.xlsx")
# df.to_excel(output_path, index=False)

# print(f"已產生檔案：{output_path}")




# ==================================================
# SOURCE FILE: XXX拜訪清單資料區域匯總.py
# ==================================================

import pandas as pd
from pathlib import Path
folder_path = Path(r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\2025.11\文斌備份用")
excel_files = list(folder_path.glob("*.xls*"))
result = []
for file in excel_files:
    df = pd.read_excel(file)
    if '資料區域群組名稱' not in df.columns:
        print(f"欄位不存在，跳過：{file.name}")
        continue
    cnt = ( df['資料區域群組名稱'].value_counts(dropna=False).rename(file.stem) )
    result.append(cnt)
summary_df = pd.concat(result, axis=1).fillna(0).astype(int)
summary_df.to_excel("拜訪清單資料區域匯總202511.xlsx")

summary_df


