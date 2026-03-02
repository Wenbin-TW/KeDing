# -*- coding: utf-8 -*-
# Generated from folder: /Users/alysonchen/Downloads/KeDing/自動化月結

# ==================================================
# SOURCE FILE: 1_自動化月結.py
# ==================================================

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
CUSTOM_PATH = Path(r"C:\Users\TW0002.TPTWKD\Desktop\項目整理")
sys.path.append(str(CUSTOM_PATH))
import common as kd 
BASE_DIR = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度"
FILE_CONFIG = {
    "業管客資": r"Z:\18_各部門共享區\03_台灣事業部\TO 圻馨\每日新增ZSD31",
    "呆帳": r"Z:\18_各部門共享區\01_會計課\11.海外子公司共用\呆帳",
    "逾期": r"Z:\18_各部門共享區\03_台灣事業部\6.訊息公佈欄\4.逾期收款明細表",
    "五項指標": r"Z:\18_各部門共享區\03_台灣事業部\TO 圻馨\每日新增ZSD31"
}

def clean_client_id(df, col='客戶編號'):
    return df[col].astype(str).str.strip()

def load_data():
    path = kd.get_latest_excel(FILE_CONFIG["業管客資"], "業管-客資(自動化月結)")
    df_zsd31 = pd.read_excel(path, sheet_name="SAP_Data", header=None)
    df_zsd31 = pd.DataFrame(df_zsd31.values[1:-3], columns=df_zsd31.iloc[0])
    df_zsd31 = df_zsd31[['客戶編號', '名稱1', '搜尋條件1', '信用管制說明', '大群組說明', '小群組說明', '區域', '交易額度', '客戶價格群組']]
    df_zsd31.columns = ['客戶編號', '名稱1', '搜尋條件1', '信用管制說明', '大區', '小區', '區域', '前月交易額度', '類別']
    df_zsd31['客戶編號'] = clean_client_id(df_zsd31)
    df_zsd31['前月交易額度'] = pd.to_numeric(df_zsd31['前月交易額度'], errors='coerce').fillna(0).astype(int)
    zsd31 = df_zsd31[
        (~df_zsd31['客戶編號'].str.contains('P')) & 
        (~df_zsd31['搜尋條件1'].str.contains('建檔錯誤', na=False)) & 
        (df_zsd31['大區'].notna())
    ].copy()
    path_bad_debt = kd.get_latest_excel(FILE_CONFIG["呆帳"], "1號交呆帳明細20")
    df_bad_debt = pd.read_excel(path_bad_debt, sheet_name='呆帳明細', usecols=['客代']).drop_duplicates()
    df_bad_debt['客代'] = clean_client_id(df_bad_debt, '客代')
    path_shared = kd.get_latest_excel(BASE_DIR, "共用額度登記")
    df_shared = pd.read_excel(path_shared, sheet_name="工作表1", usecols=['客戶編號', '共用組數', '分配額度', '額度分配比例'])
    df_shared.columns = ['客戶編號', '共用額度組別', '分配額度', '額度分配比例']
    df_shared['客戶編號'] = clean_client_id(df_shared)
    path_overdue = kd.get_latest_excel(FILE_CONFIG["逾期"], "逾期帳款(製表人陳宗卿 2025")
    df_overdue = pd.concat(pd.read_excel(path_overdue, sheet_name=['北', '中', '南'], header=3, usecols=['客戶', '合計未收']).values())
    df_overdue = df_overdue[df_overdue['客戶'].str.contains('TW', na=False)].copy()
    df_overdue['客戶編號'] = clean_client_id(df_overdue, '客戶')
    df_overdue['當月逾期金額'] = (df_overdue['合計未收'] / 1.05).clip(lower=0).round(0).astype(int)
    
    return zsd31, df_bad_debt, df_shared, df_overdue

def process_logic(zsd31, df_bad_debt, df_shared, df_overdue):
    df = pd.merge(zsd31, df_bad_debt, left_on="客戶編號", right_on="客代", how='left')
    df = pd.merge(df, df_shared, on="客戶編號", how='left')
    df = pd.merge(df, df_overdue[['客戶編號', '當月逾期金額']], on="客戶編號", how='left').fillna({'當月逾期金額': 0})
    two_years_str = (datetime.now() - relativedelta(years=2)).strftime('%Y/%m/%d')
    sql_sales = f"SELECT buyer as 客戶編號, taxfree_basecurr FROM sap_sales_data WHERE buyer LIKE 'TW%' AND planned_shipping_date >= '{two_years_str}'"
    df_sap_sales = kd.get_data_from_MSSQL(sql_sales).groupby('客戶編號')['taxfree_basecurr'].sum().reset_index()
    
    sql_cash = "SELECT KUNNR as 客戶編號, sum(DMBTR) as '已兌現金額_稅後' FROM [SAPdb].[dbo].[ZFI66] WHERE BUKRS = '1000' GROUP BY KUNNR"
    df_cash = kd.get_data_from_MSSQL(sql_cash)
    df_cash['已兌現金額_未稅'] = (df_cash['已兌現金額_稅後'] / 1.05).round(0)
    df = pd.merge(df, df_sap_sales, on='客戶編號', how='left')
    df = pd.merge(df, df_cash[['客戶編號', '已兌現金額_未稅']], on='客戶編號', how='left').fillna(0)
    df['近2年有交易'] = np.where(df['taxfree_basecurr'] > 0, 'V', 'X')
    path_rank = kd.get_latest_excel(BASE_DIR, "級別及額度")
    df_rank = pd.read_excel(path_rank, header=1).dropna(subset=['累計兌現金額(元)'])
    
    def get_rank_credit(money, mode='credit'):
        bins = df_rank['累計兌現金額(元)'].tolist()
        labels = df_rank['信用額度(元)'].tolist() if mode == 'credit' else df_rank['級別'].tolist()
        idx = np.digitize(money, bins) - 1
        return labels[max(0, idx)]
    df['兌現_信用額度'] = df['已兌現金額_未稅'].apply(lambda x: get_rank_credit(x, 'credit') if x > 0 else 250000)
    df['兌現_級別'] = df['已兌現金額_未稅'].apply(lambda x: get_rank_credit(x, 'rank') if x >= 50000 else '現金')
    group_cash = df.groupby('共用額度組別')['已兌現金額_未稅'].sum().reset_index(name='組別總兌現')
    df = pd.merge(df, group_cash, on='共用額度組別', how='left')
    df['組別判定額度'] = df['組別總兌現'].apply(lambda x: get_rank_credit(x, 'credit') if pd.notna(x) else np.nan)
    df['最新交易額度'] = df['兌現_信用額度']
    mask_freeze = (df['已兌現金額_未稅'] <= 50000) | (df['信用管制說明'] == '呆賬管制') | (df['當月逾期金額'] > 0) | (df['近2年有交易'] == 'X')
    df.loc[mask_freeze, '最新交易額度'] = df['前月交易額度']
    
    return df

def export_results(df):
    today = datetime.now().strftime("%Y%m%d")
    output_path = Path(BASE_DIR) / "自動化月結_輸出" / today
    output_path.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path / f"新交易額度客戶明細_{today}.xlsx", index=False)
    with pd.ExcelWriter(output_path / f"sap大批匯入_{today}.xlsx") as writer:
        df[df['最新交易額度'] != df['前月交易額度']].to_excel(writer, sheet_name='額度變更', index=False)

    print(f"處理完成，檔案儲存於: {output_path}")

if __name__ == "__main__":
    zsd31, bad_debt, shared, overdue = load_data()
    final_df = process_logic(zsd31, bad_debt, shared, overdue)
    export_results(final_df)

# ==================================================
# SOURCE FILE: 2_自動化月結SG.py
# ==================================================



import os
import glob
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from datetime import datetime
import pymysql
from dateutil.relativedelta import relativedelta
from pathlib import Path
import sys
custom_path = Path(r"C:\Users\TW0002.TPTWKD\Desktop\項目整理")
sys.path.append(str(custom_path))
import common as kd
import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


業管_客資_df = kd.get_data_from_MSSQL('''
            SELECT iNo	流水編號 ,KUNNR	客戶編號,KNAME	客戶名稱,SORT1	搜尋條件1,STREET	街道1,BEZEI	'客戶群組1-說明'
            ,STR_SUPPL2	街道3,TEL_NUMBER	電話,VTEX8	信用管制說明
            ,VTEX3	稅分類說明,STCD1	統一編號,VTEX10	性能分類說明
            ,GFORM	專案價,TDLIN3	行銷註記,PANAME	員工姓名
            ,VTEX08	小群組說明,FAX_NUMBER	傳真,TDLIN1	銷售註記
            ,SORT2	搜尋條件2,ERDAT	建檔日期
            ,ERNAM	建立者,TDLIN2	出貨指示,COUNTRY	國家碼
            ,ZZCRM_CUSTOMER	銷售易公司代號
            ,ZZCRM_COSHORTNAME	銷售易公司名稱
            ,VKORG	銷售組織,VTEX1	銷售組織說明
            ,BZIRK	銷售地區,BZTXT	銷售地區說明
            ,VKBUR	銷售據點,BEZE1	銷售據點說明
            ,WAERS	幣別,KTOKD	客戶科目群組
            ,TXT30	客戶科目群組說明,NAME1	名稱1
            ,ZTERM	'付款條件碼(公司)',VTEX9	'付款條件說明(公司)'
            ,ZTERM_1	'付款條件碼(銷售)'
            ,KONDA	客戶價格群組,VTEX2	客戶價格群組說明
            ,VSBED	出貨條件,VTEXT1	出貨條件說明,INCO1	國貿條件
            ,BEZE2	國貿條件說明,INCO2	國貿條件2,PERNR	員工編號
            ,ADNAME	聯絡人,TEL_NO	聯絡人手機,KTGRD	科目指派群組
            ,VTEX6	科目指派群組說明,BRTXT	產業說明,KATR1	結帳日
            ,VTEX01	結帳日說明,KATR2	寄單日,VTEX02	寄單日說明
            ,VTEX03	放款日說明,KATR4	帳單寄送方式
            ,VTEX04	帳單寄送方式說明,KATR5	付款方式,VTEX05	付款方式說明
            ,KATR6	'地區(附加)',VTEX06	'地區說明(附加)',KATR7	大群組
            ,VTEX07	大群組說明,KATR8	小群組,KATR9	區域,BANKL	銀行代號
            ,BANKN	銀行帳號,KOINH	銀行戶名,REMARK	備註
            ,VTEXT2	專案價說明,GRUPP	客戶信用群組,KLIMK	交易額度
            ,POST_CODE1	郵遞區號,REGION	'地區(一般)'
            ,BEZE4	'地區說明(一般)',PERNR_ZM	主任秘書代號
            ,PERNR_ZM_NAME	主任秘書
            ,EMAIL	Email,ZZBANKAC	銀行專屬繳款帳號,BANKA	銀行名稱
            ,dCreateDate	資料建立日期
            FROM [SAPdb].[dbo].[ZSD31B]
            where vkorg= 'SG00' and KTOKD = 'YB01' and (KUNNR like 'SG%' or KUNNR like 'MY%')

            ''')


業管_客資_df = 業管_客資_df[['客戶編號', '名稱1', '搜尋條件1', '信用管制說明','國家碼','大群組說明', '小群組說明', '區域', '交易額度', '客戶價格群組'
                     ]].rename(columns={'大群組說明': '大區','小群組說明': '小區','交易額度': '前月交易額度','客戶價格群組': '類別'})
業管_客資_df = 業管_客資_df[~業管_客資_df['類別'].astype(str).str.contains('z', case=False, na=False)]

業管_客資_df['客戶編號'] = 業管_客資_df['客戶編號'].astype(str)
業管_客資_df['前月交易額度'] = pd.to_numeric(業管_客資_df['前月交易額度'], errors='coerce').fillna(0).astype(int)
ZSD31 = 業管_客資_df.copy()
呆帳_file = r"Z:\18_各部門共享區\01_會計課\11.海外子公司共用\呆帳"
呆帳_keyword = "SG呆帳明細"
呆帳_path = kd.get_latest_excel(呆帳_file, 呆帳_keyword)

呆帳_df = pd.read_excel(呆帳_path, sheet_name='呆帳明細', usecols=['客代']).drop_duplicates(subset=['客代'], keep='first')
呆帳_merge = pd.merge(ZSD31, 呆帳_df, left_on="客戶編號", right_on="客代", how='left').rename(columns={'客代': '呆帳客戶'})
共額_file = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度"
共額_keyword = "共用額度登記"
共額_path = kd.get_latest_excel(共額_file, 共額_keyword)

共額_sg = pd.read_excel(共額_path,sheet_name="SG",usecols=['客戶編號', '共用組數', '額度分配比例'])
共額_df = pd.concat([共額_sg], ignore_index=True)

共額_df = 共額_df.rename(columns={'共用組數': '共用額度組別'})
共額_df['客戶編號'] = 共額_df['客戶編號'].str.strip()
def classify_group(df):
    my_groups = (
        df[df['客戶編號'].str.startswith('MY')]
        ['共用額度組別']
        .unique()
    )
    df['共用額度組別分類'] = df['共用額度組別'].apply(
        lambda x: 'MY' if x in my_groups else 'SG'
    )
    return df

共額_df = classify_group(共額_df)


共額_df['共用額度組別'] = 共額_df['共用額度組別'].astype('Int64')  # 注意 Int64 是 pandas 的 nullable int 型別
共額_merge = pd.merge(呆帳_merge, 共額_df, on="客戶編號", how='left')
共額_merge = 共額_merge.drop_duplicates(subset=['客戶編號'], keep='first')
Overdue_file = r"Z:\18_各部門共享區\06_海外業務部\$$每月逾期貨款\逾期帳款報表-每月更新"
Overdue_keyword_sg = "月逾期帳款-SG("
Overdue_sheets = ['原幣']
Overdue_path_sg = kd.get_latest_excel(Overdue_file, Overdue_keyword_sg)
Overdue_data_sg = pd.read_excel(Overdue_path_sg, sheet_name=Overdue_sheets, header=4, usecols=['客戶', '合計未收'])
Overdue_clean_sg = Overdue_data_sg['原幣']
Overdue_clean_sg = Overdue_clean_sg[Overdue_clean_sg['客戶'].notna() & (Overdue_clean_sg['客戶'].astype(str).str.strip() != "")]




df5 = pd.concat([ Overdue_clean_sg])

df5 = df5.rename(columns={'客戶': '客戶編號'})
df5['合計未收'] = df5['合計未收'].astype('int')
df5['當月逾期金額(依會計每月20號的逾期表)'] = df5['合計未收'] / 1.09
df5.loc[df5['合計未收'] < 0, '當月逾期金額(依會計每月20號的逾期表)'] = 0
df5['當月逾期金額(依會計每月20號的逾期表)'] = df5['當月逾期金額(依會計每月20號的逾期表)'].astype('float') + 0.00000001
df5['當月逾期金額(依會計每月20號的逾期表)'] = df5['當月逾期金額(依會計每月20號的逾期表)'].round(0)
df5 = df5.drop(['合計未收'], axis=1)

逾期_merge = pd.merge(共額_merge, df5, on="客戶編號", how='left')
逾期_merge['當月逾期金額(依會計每月20號的逾期表)'] = 逾期_merge['當月逾期金額(依會計每月20號的逾期表)'].fillna(0)
逾期_merge['當月逾期金額(依會計每月20號的逾期表)'] = 逾期_merge['當月逾期金額(依會計每月20號的逾期表)'].astype('int')
two_years_ago = datetime.now() - relativedelta(years=2)
two_years_str = two_years_ago.strftime('%Y/%m/%d')

sap = kd.get_data_from_MSSQL(f'''
        SELECT buyer as 客戶編號, taxfree_basecurr as 未稅本位幣,  planned_shipping_date as 預計發貨日期 
        FROM sap_sales_data 
        WHERE  (buyer LIKE '%SG%' or buyer like '%MY%')
        AND planned_shipping_date >= '{two_years_str}'
        AND taxfree_basecurr > 0
    ''')

sap = (sap.groupby('客戶編號', as_index=False)['未稅本位幣'].sum().query("`未稅本位幣` > 0"))

銷貨_merge = pd.merge(逾期_merge, sap, on="客戶編號", how='left')
銷貨_merge['近2年有交易(現金可轉月結)'] = np.where(銷貨_merge['未稅本位幣']>0, 'V', 'X')
銷貨_merge['近2年有交易(現金可轉月結)'] = 銷貨_merge['近2年有交易(現金可轉月結)'].fillna(value="X")
not_monthly_sg = pd.read_excel(共額_path,sheet_name="SG",usecols=['不轉月結名單']).rename(columns={'不轉月結名單': '客戶編號'})
not_monthly_my = pd.read_excel(共額_path,sheet_name="MY",usecols=['不轉月結名單']).rename(columns={'不轉月結名單': '客戶編號'})
not_monthly_df = pd.concat([not_monthly_my, not_monthly_sg])
not_monthly_df[not_monthly_df['客戶編號'].notna() & (not_monthly_df['客戶編號'].astype(str).str.strip() != "")]


not_monthly_list = not_monthly_df['客戶編號'].dropna().tolist()

total_money = kd.get_data_from_MSSQL('''
        SELECT  [KUNNR]  客戶編號,sum([DMBTR])  '已兌現金額(稅後)'
            FROM [SAPdb].[dbo].[ZFI66]
            where (BUKRS like '%SG%' or BUKRS like '%MY%')  and ( VKORG like '%SG%' or  VKORG like '%MY%')
            group by [KUNNR]
        ''')
銷貨_merge = pd.merge(銷貨_merge, total_money, on="客戶編號", how='left')
銷貨_merge['已兌現金額(未稅)'] = (銷貨_merge['已兌現金額(稅後)'].fillna(0).div(1.09).add(1e-8).round(0))


merged_df = 銷貨_merge.copy()

group_sums = merged_df[merged_df['共用額度組別'].notna()].groupby('共用額度組別')['已兌現金額(未稅)'].sum().reset_index()
group_sums.columns = ['共用額度組別', '共用組合計']
merged_df = pd.merge(merged_df, group_sums, on='共用額度組別', how='left')
merged_df['已兌現金額(未稅)'] = merged_df.apply(lambda row: row['共用組合計'] if pd.notna(row['共用組合計']) else row['已兌現金額(未稅)'], axis=1)
merged_df.drop(columns=['共用組合計'], inplace=True)
merged_df['共用額度組別分類'] = merged_df['共用額度組別分類'].replace(['', 'nan', 'None'], np.nan)
merged_df['共用額度組別分類'] = merged_df.apply(lambda x: x['客戶編號'][:2] if pd.isna(x['共用額度組別分類']) else x['共用額度組別分類'],axis=1)
級別及額度_file = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\✨規範.資料"
級別及額度_keyword = "月結申請暨交易額度"
級別及額度_path = kd.get_latest_excel(級別及額度_file, 級別及額度_keyword)

level_df = pd.read_excel(級別及額度_path, header=1)  # 這裡 header=1 因為標題行在第2列（從0算起）
sg_df = level_df.iloc[:, 0:3].copy()   # A, B, C 欄
sg_df.columns = ['級別', '累計兌現金額', '交易額度']  # 可依實際調整欄名
sg_df = sg_df.dropna(how='all')        # 去掉全空列
sg_df['累計兌現金額'] = pd.to_numeric(sg_df['累計兌現金額'], errors='coerce').astype(pd.Int64Dtype()).fillna(0)
sg_df['交易額度'] = pd.to_numeric(sg_df['交易額度'], errors='coerce').astype(pd.Int64Dtype())
my_df = level_df.iloc[:, 4:7].copy()   # E, F, G 欄
my_df.columns = ['級別', '累計兌現金額', '交易額度']
my_df = my_df.dropna(how='all')
my_df['累計兌現金額'] = pd.to_numeric(my_df['累計兌現金額'], errors='coerce').astype(pd.Int64Dtype()).fillna(0)
my_df['交易額度'] = pd.to_numeric(my_df['交易額度'], errors='coerce').astype(pd.Int64Dtype())

def get_credit_limit(row):
    used_amount = row['已兌現金額(未稅)']
    allocation_ratio = row['額度分配比例']
    category = row['共用額度組別分類']  # 新增這個欄位的判斷
    if category == 'MY':
        level_table = my_df
    else:
        level_table = sg_df

    credit_limit = 0
    for _, level_row in level_table.iterrows():
        if pd.notna(used_amount) and used_amount >= level_row['累計兌現金額']:
            credit_limit = level_row['交易額度']
        else:
            break
    if pd.isna(allocation_ratio):
        return round(credit_limit, 0)
    else:
        return round(credit_limit * allocation_ratio, 0)
merged_df['交易額度'] = merged_df.apply(get_credit_limit, axis=1).fillna(0).astype(int)
merged_df.loc[merged_df['交易額度'] == 0, '交易額度'] = 1





merged_df['額度變化數值'] =  merged_df['交易額度'] - merged_df['前月交易額度']


def classify_final(row):
    if row['信用管制說明'] == '呆賬管制':
        return '額度不變(管制)', '額度不變'

    if pd.notna(row['呆帳客戶']) and str(row['呆帳客戶']).strip() != "":
        return '額度不變(呆賬)', '額度不變'

    if row['當月逾期金額(依會計每月20號的逾期表)'] > 0:
        return '額度不變(逾期)', '額度不變'

    if row['近2年有交易(現金可轉月結)'] == 'X':
        return '額度不變(凍結)', '額度不變'

    if row['客戶編號'] in not_monthly_list:
        return '額度不變(共額不轉)', '額度不變'
    diff = row['額度變化數值']

    if diff < 0:
        return '額度下降', '額度下降'
    elif diff > 0:
        return '額度上升', '額度上升'
    else:
        return '額度不變', '額度不變'

merged_df[['客戶分類', '額度變化']] = merged_df.apply(
    lambda row: pd.Series(classify_final(row)), axis=1
)
condition = merged_df['額度變化'].str.contains('額度不變', na=False)
merged_df.loc[condition, '交易額度'] = merged_df.loc[condition, '前月交易額度']
merged_df.loc[condition, '額度變化數值'] = 0




merged_df['新信用管制說明'] =  merged_df['信用管制說明']
merged_df.loc[(merged_df['信用管制說明'] == '現金客戶') & (merged_df['額度變化'] == '額度上升') | (merged_df['額度變化'] == '額度下降') , '新信用管制說明'] = '月結'
merged_df.loc[(merged_df['額度變化'] == '額度上升') & (merged_df['信用管制說明'] == '月結') &  (merged_df['新信用管制說明'] == '月結'), '客戶分類'] = '自動化(額度增加)'
merged_df.loc[(merged_df['額度變化'] == '額度上升') & (merged_df['信用管制說明'] == '現金客戶') &  (merged_df['新信用管制說明'] == '月結'), '客戶分類'] = '自動化(新月結客戶)'



merged_df.loc[(merged_df['額度變化'] == '自動化(新月結客戶)') & (merged_df['新信用管制說明'] == '現金客戶'), '新信用管制說明'] = '月結'
共用額度不轉月結_merge3 = merged_df




共用額度不轉月結_merge3.pivot_table(
    index='額度變化',    
    columns='客戶分類', 
    values='客戶編號',   
    aggfunc='count',           
    fill_value=0
)

共用額度不轉月結_merge3 = 共用額度不轉月結_merge3.drop_duplicates(subset=['客戶編號'], keep='last')

d1 = datetime.today()
d1 = d1.strftime("%Y/%m/01")
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '月結起始日'] = d1
共用額度不轉月結_merge3['月結起始日'] = pd.to_datetime(共用額度不轉月結_merge3['月結起始日'], errors='coerce', utc=True).dt.date

共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '新月結客結帳日/寄單日'] = '結帳日30日/寄單日5日'
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '票期(統一5040不可動)'] = '票期5040'
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '新月結客帳單寄送方式'] = '99 其他'
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '新月結客付款方式'] = '02 匯款'




priority = {"自動化(新月結客戶)": 0}   # 0 最前面，其餘自動排後
共用額度不轉月結_merge3 = 共用額度不轉月結_merge3.sort_values( by="客戶分類", key=lambda s: s.map(priority).fillna(1) )




import os
from datetime import datetime
today = datetime.now().strftime("%Y%m%d")
folder = fr'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\{today}'
os.makedirs(folder, exist_ok=True)
共用額度不轉月結_merge3.to_excel(os.path.join(folder, f"SG_總額度判定{today}.xlsx"), index=False)
target_classifications = ['自動化(新月結客戶)', '額度下降', '自動化(額度增加)']
filtered_df = merged_df.loc[
    merged_df['客戶分類'].isin(target_classifications)]

credit_change_columns = ['客戶編號', '信用控制範圍', '信用額度', '風險種類', '客戶信用群組', '前次信用額度(比對用)',
    'GOOGLE申請', '比對上次', '共用額度組別', '備註', '變更']
credit_change = pd.DataFrame(columns=credit_change_columns)

credit_change['客戶編號'] = filtered_df['客戶編號'].values
credit_change['信用額度'] = filtered_df['交易額度'].values
credit_change['前次信用額度(比對用)'] = filtered_df['前月交易額度'].values
credit_change['共用額度組別'] = filtered_df['共用額度組別'].values
credit_change['備註'] = filtered_df['客戶分類'].values
new_monthly_df = merged_df.loc[merged_df['客戶分類'] == '自動化(新月結客戶)']

new_monthly_columns = ['客戶代碼', '公司代碼', '銷售組織', '配銷通路', '部門', '搜尋條件 2',
    '結帳日', '寄單日', '帳單寄送方式', '付款方式', '信用管制', '付款條件碼_K', '付款條件碼_V']
new_monthly_table = pd.DataFrame(columns=new_monthly_columns)

new_monthly_table['客戶代碼'] = new_monthly_df['客戶編號'].values
new_monthly_table['搜尋條件 2'] = datetime.today().replace(day=1).strftime('%Y-%m-%d')
new_monthly_table['結帳日'] = '結帳日30日/寄單日5日'
new_monthly_table['寄單日'] = '票期5040'
new_monthly_table['帳單寄送方式'] = '99 其他'
new_monthly_table['付款方式'] = '02 匯款'
new_monthly_table['信用管制'] = '月結'
save_path = fr'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\{today}'
file_name = f"SG_sap大批匯入(給惠茹)-{today}.xlsx"  
full_path = os.path.join(save_path, file_name)  
with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
    new_monthly_table.to_excel(writer, sheet_name='改月結', index=False)
    credit_change.to_excel(writer, sheet_name='額度變更', index=False)

    workbook = writer.book
    ws1 = workbook.create_sheet(title='共用')
    ws1.append(['客戶編號', '共用額度組別', '備註'])
    ws2 = workbook.create_sheet(title='轉現金')
    ws2.append(['客戶編號', '信用管制', '信用額度'])
    workbook.save(full_path)

print(f"已經成功匯出到 {full_path}！")

columns_for_sales = [
    '客戶編號', '名稱1', '搜尋條件1', '大區', '區域', '類別', '呆帳客戶',
    '客戶分類', '新信用管制說明', '交易額度',
    
    '月結客戶維持現金客戶(業務確認)', '維持原因(業務填寫)',
    '收款聯絡人', '收款聯絡人電話', '個人LINE網址', '群组LINE網址', '收款備註',
    '業務送帳單/收款需填原因',

    '新月結客結帳日/寄單日', '票期(統一5040不可動)', '新月結客帳單寄送方式', '新月結客付款方式',
    '月結起始日', '小區'
]

final_sales_table = pd.DataFrame(columns=columns_for_sales)

source_columns = [
    '客戶編號', '名稱1', '搜尋條件1', '大區', '區域', '類別', '呆帳客戶',
    '客戶分類', '新信用管制說明', '交易額度',
    '新月結客結帳日/寄單日', '票期(統一5040不可動)', '新月結客帳單寄送方式', '新月結客付款方式',
    '月結起始日', '小區']

for col in source_columns:
    if col in filtered_df.columns:
        final_sales_table[col] = filtered_df[col].values


info_file = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度"
info_keyword = "ALL客戶收款清單聯絡資訊"
info_path = kd.get_latest_excel(info_file, info_keyword)
info_df = pd.read_excel(info_path)
info_df = info_df[['客戶', '收款連絡人', '收款連絡人電話', '收款人LINE ID', '群組LINE ID', '收款備註']]
info_df = info_df.drop_duplicates(subset='客戶', keep='first')
final_sales_table['客戶編號'] = final_sales_table['客戶編號'].astype(str).str.strip()
info_df['客戶'] = info_df['客戶'].astype(str).str.strip()
info_dict = info_df.set_index('客戶').to_dict(orient='index')
final_sales_table['收款聯絡人'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款連絡人', pd.NA))
final_sales_table['收款聯絡人電話'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款連絡人電話', pd.NA))
final_sales_table['個人LINE網址'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款人LINE ID', pd.NA))
final_sales_table['群组LINE網址'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('群組LINE ID', pd.NA))
final_sales_table['收款備註'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款備註', pd.NA))
final_sales_table['月結起始日'] = datetime.today().replace(day=1).strftime('%Y-%m-%d')



folder = fr'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\{today}'
os.makedirs(folder, exist_ok=True)
final_sales_table.to_excel(os.path.join(folder, f"SG_新交易額度客戶明細-業務{today}.xlsx"), index=False)





# ==================================================
# SOURCE FILE: 3_自動化月結_海外六國.py
# ==================================================


# -*- coding: utf-8 -*-


import os
import sys
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta


# 0) 匯入kd
custom_path = Path(r"C:\Users\TW0002.TPTWKD\Desktop\項目整理")
sys.path.append(str(custom_path))
import common as kd  # noqa: E402


# 1) 共用路徑
BAD_DEBT_DIR = r"Z:\18_各部門共享區\01_會計課\11.海外子公司共用\呆帳"
BAD_DEBT_KEYWORD = "子公司歷年呆帳客戶名單"  
BAD_DEBT_SHEET = "歷年呆帳(勿改名稱!)"
BAD_DEBT_COL = "客代"

SHARED_CREDIT_DIR = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度"
SHARED_CREDIT_KEYWORD = "共用額度登記"

OVERDUE_DIR = r"Z:\18_各部門共享區\06_海外業務部\$$每月逾期貨款\逾期帳款報表-每月更新"
OVERDUE_SHEET = "原幣"
OVERDUE_HEADER = 4
OVERDUE_USECOLS = ["客戶", "合計未收"]

COLLECTION_INFO_DIR = SHARED_CREDIT_DIR
COLLECTION_INFO_KEYWORD = "ALL客戶收款清單聯絡資訊"

# 2) 國家設定）
# FROM [SAPdb].[dbo].[ZSD31B]
@dataclass
class CountryConfig:
    code: str
    vkorg: str
    buyer_prefix: List[str]
    tax_rate: float
    base_path: str
    has_bad_debt: bool
    overdue_keyword: str
    level_dir: str
    level_keyword: str
    shared_credit_sheet: Optional[str] = None
    not_monthly_sheets: Optional[List[str]] = None


COUNTRY_CONFIGS: Dict[str, CountryConfig] = {
    "HK": CountryConfig(
        code="HK",
        vkorg="HK00",
        buyer_prefix=["HK"],
        tax_rate=0.0,
        base_path=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\3.  HK 香港",
        has_bad_debt=True,  
        overdue_keyword="月逾期帳款-HK(",
        level_dir=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\3.  HK 香港",
        level_keyword="月結額度",
        shared_credit_sheet="HK",
        not_monthly_sheets=["HK"],
    ),
    "JP": CountryConfig(
        code="JP",
        vkorg="JP00",
        buyer_prefix=["JP"],
        tax_rate=0.10,
        base_path=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\4.  JP 日本",
        has_bad_debt=True,
        overdue_keyword="月逾期帳款-JP(",
        level_dir=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\4.  JP 日本",
        level_keyword="月結額度",
        shared_credit_sheet="JP",
        not_monthly_sheets=["JP"],
    ),
    "PH": CountryConfig(
        code="PH",
        vkorg="PH00",
        buyer_prefix=["PH"],
        tax_rate=0.12,
        base_path=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\5.  PH 菲律賓",
        has_bad_debt=True,
        overdue_keyword="月逾期帳款-PH(",
        level_dir=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\5.  PH 菲律賓",
        level_keyword="月結額度",
        shared_credit_sheet="PH",
        not_monthly_sheets=["PH"],
    ),
    "ID": CountryConfig(
        code="ID",
        vkorg="ID00",
        buyer_prefix=["ID"],
        tax_rate=0.11,
        base_path=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\6.  ID 印尼",
        has_bad_debt=True,
        overdue_keyword="月逾期帳款-ID(",
        level_dir=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\6.  ID 印尼",
        level_keyword="月結額度",
        shared_credit_sheet="ID",
        not_monthly_sheets=["ID"],
    ),
    "VN": CountryConfig(
        code="VN",
        vkorg="VN00",
        buyer_prefix=["VN"],
        tax_rate=0.08,
        base_path=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\7.  VN 越南",
        has_bad_debt=True,
        overdue_keyword="月逾期帳款-VN(",
        level_dir=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\7.  VN 越南",
        level_keyword="月結額度",
        shared_credit_sheet="VN",
        not_monthly_sheets=["VN"],
    ),
    "TH": CountryConfig(
        code="TH",
        vkorg="TH00",
        buyer_prefix=["TH"],
        tax_rate=0.07,
        base_path=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\8. TH 泰國",
        has_bad_debt=True,
        overdue_keyword="月逾期帳款-TH(",
        level_dir=r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\8. TH 泰國",
        level_keyword="月結額度",
        shared_credit_sheet="TH",
        not_monthly_sheets=["TH"],
    ),
}


# 3) 工具函數（共用）
def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def safe_strip_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip()


def build_prefix_like_sql(col: str, prefixes: List[str]) -> str:
    return " OR ".join([f"{col} LIKE '{p}%'" for p in prefixes])


# 4) Data Loaders
def load_zsd31(cfg: CountryConfig) -> pd.DataFrame:
    prefix_sql = build_prefix_like_sql("KUNNR", cfg.buyer_prefix)

    df = kd.get_data_from_MSSQL(f"""
        SELECT
            KUNNR   AS 客戶編號,
            NAME1   AS 名稱1,
            SORT1   AS 搜尋條件1,
            VTEX8   AS 信用管制說明,
            COUNTRY AS 國家碼,
            VTEX07  AS 大群組說明,
            VTEX08  AS 小群組說明,
            KATR9   AS 區域,
            KLIMK   AS 交易額度,
            KONDA   AS 客戶價格群組
        FROM [SAPdb].[dbo].[ZSD31B]
        WHERE vkorg = '{cfg.vkorg}'
          AND KTOKD = 'YB01'
          AND ({prefix_sql})
    """)

    df = df[
        ["客戶編號", "名稱1", "搜尋條件1", "信用管制說明", "國家碼", "大群組說明", "小群組說明", "區域", "交易額度", "客戶價格群組"]
    ].rename(
        columns={
            "大群組說明": "大區",
            "小群組說明": "小區",
            "交易額度": "前月交易額度",
            "客戶價格群組": "類別",
        }
    )

    # 排除 Z 類客戶
    df = df[~df["類別"].astype(str).str.contains("z", case=False, na=False)].copy()

    df["客戶編號"] = safe_strip_series(df["客戶編號"])
    df["前月交易額度"] = pd.to_numeric(df["前月交易額度"], errors="coerce").fillna(0).astype(int)

    return df


def load_bad_debt(cfg: CountryConfig) -> pd.DataFrame:
    if not cfg.has_bad_debt:
        return pd.DataFrame(columns=["客戶編號"])

    path = kd.get_latest_excel(BAD_DEBT_DIR, BAD_DEBT_KEYWORD)
    df = pd.read_excel(path, sheet_name=BAD_DEBT_SHEET, usecols=[BAD_DEBT_COL]).drop_duplicates()
    df = df.rename(columns={BAD_DEBT_COL: "客戶編號"})
    df["客戶編號"] = safe_strip_series(df["客戶編號"])
    df["呆帳客戶"] = df["客戶編號"]
    return df[["客戶編號", "呆帳客戶"]]


def load_shared_credit_registry(cfg: CountryConfig) -> pd.DataFrame:
    """
    共用額度登記表：
    - 若指定 sheet 不存在 → 回傳空 DataFrame（代表該國沒有共用額度）
    - 不視為錯誤
    """
    path = kd.get_latest_excel(SHARED_CREDIT_DIR, SHARED_CREDIT_KEYWORD)
    sheet = cfg.shared_credit_sheet or cfg.code

    try:
        df = pd.read_excel(
            path,
            sheet_name=sheet,
            usecols=["客戶編號", "共用組數", "額度分配比例"],
        )
    except ValueError:
        # sheet 不存在，直接跳過
        print(f"⚠ 共用額度登記表中沒有 {sheet} 分頁，已跳過")
        return pd.DataFrame(columns=["客戶編號", "共用額度組別", "額度分配比例"])

    df = df.rename(columns={"共用組數": "共用額度組別"})
    df["客戶編號"] = df["客戶編號"].astype(str).str.strip()
    df["共用額度組別"] = pd.to_numeric(df["共用額度組別"], errors="coerce").astype("Int64")

    return df



def load_not_monthly_list(cfg: CountryConfig) -> List[str]:
    """
    共用額度登記表裡的不轉月結名單
    這裡改成 cfg.not_monthly_sheets（通常只要該國）
    """
    path = kd.get_latest_excel(SHARED_CREDIT_DIR, SHARED_CREDIT_KEYWORD)
    sheets = cfg.not_monthly_sheets or [cfg.code]

    dfs = []
    for sh in sheets:
        try:
            tmp = pd.read_excel(path, sheet_name=sh, usecols=["不轉月結名單"]).rename(columns={"不轉月結名單": "客戶編號"})
            dfs.append(tmp)
        except Exception:
            # 該國沒有此欄/此 sheet 就略過
            continue

    if not dfs:
        return []

    df = pd.concat(dfs, ignore_index=True)
    df["客戶編號"] = safe_strip_series(df["客戶編號"])
    df = df[df["客戶編號"].notna() & (df["客戶編號"].astype(str).str.strip() != "")]
    return df["客戶編號"].dropna().tolist()

def load_overdue(cfg: CountryConfig) -> pd.DataFrame:
    """
    逾期帳款：
    - 找不到檔案 → 視為沒有逾期，回傳空資料（全部當 0）
    - 不視為錯誤
    """

    try:
        path = get_latest_excel_or_fail(
            OVERDUE_DIR,
            cfg.overdue_keyword,
            country=cfg.code,
            usage="逾期帳款"
        )
    except RuntimeError:
        print(f"⚠ [{cfg.code}] 無逾期帳款資料，已視為 0")
        return pd.DataFrame(
            columns=["客戶編號", "當月逾期金額(依會計每月20號的逾期表)"]
        )

    data = pd.read_excel(
        path,
        sheet_name=OVERDUE_SHEET,
        header=OVERDUE_HEADER,
        usecols=OVERDUE_USECOLS
    )

    df = data.rename(columns={"客戶": "客戶編號"}).copy()
    df = df[df["客戶編號"].notna() & (df["客戶編號"].astype(str).str.strip() != "")]
    df["客戶編號"] = safe_strip_series(df["客戶編號"])

    df["合計未收"] = pd.to_numeric(df["合計未收"], errors="coerce").fillna(0)
    df["當月逾期金額(依會計每月20號的逾期表)"] = (
        df["合計未收"] / (1 + cfg.tax_rate) + 1e-8
    ).round(0)

    df.loc[df["合計未收"] < 0, "當月逾期金額(依會計每月20號的逾期表)"] = 0
    df["當月逾期金額(依會計每月20號的逾期表)"] = (
        df["當月逾期金額(依會計每月20號的逾期表)"]
        .fillna(0)
        .astype(int)
    )

    return df[["客戶編號", "當月逾期金額(依會計每月20號的逾期表)"]]


def load_sales_recent(cfg: CountryConfig, years: int = 2) -> pd.DataFrame:
    start_dt = datetime.now() - relativedelta(years=years)
    start_str = start_dt.strftime("%Y/%m/%d")

    buyer_like_sql = build_prefix_like_sql("buyer", cfg.buyer_prefix)

    sap = kd.get_data_from_MSSQL(f"""
        SELECT
            buyer AS 客戶編號,
            taxfree_basecurr AS 未稅本位幣,
            planned_shipping_date AS 預計發貨日期
        FROM sap_sales_data
        WHERE ({buyer_like_sql})
          AND planned_shipping_date >= '{start_str}'
          AND taxfree_basecurr > 0
    """)

    sap["客戶編號"] = safe_strip_series(sap["客戶編號"])
    sap["未稅本位幣"] = pd.to_numeric(sap["未稅本位幣"], errors="coerce").fillna(0)

    sap_sum = sap.groupby("客戶編號", as_index=False)["未稅本位幣"].sum()
    sap_sum = sap_sum.query("`未稅本位幣` > 0").copy()
    return sap_sum


def load_redeemed_amount(cfg: CountryConfig) -> pd.DataFrame:
    """
    ZFI66 已兌現金額：
    這裡改成用 cfg.buyer_prefix 做 KUNNR 對應（更直接），BUKRS/VKORG 條件也可以保留。
    """
    kunnr_like_sql = build_prefix_like_sql("KUNNR", cfg.buyer_prefix)

    df = kd.get_data_from_MSSQL(f"""
        SELECT
            KUNNR AS 客戶編號,
            SUM(DMBTR) AS [已兌現金額(稅後)]
        FROM [SAPdb].[dbo].[ZFI66]
        WHERE ({kunnr_like_sql})
        GROUP BY KUNNR
    """)

    df["客戶編號"] = safe_strip_series(df["客戶編號"])
    df["已兌現金額(稅後)"] = pd.to_numeric(df["已兌現金額(稅後)"], errors="coerce").fillna(0)

    # 轉未稅：依國家稅率
    df["已兌現金額(未稅)"] = (df["已兌現金額(稅後)"].div(1 + cfg.tax_rate).add(1e-8).round(0))
    df["已兌現金額(未稅)"] = pd.to_numeric(df["已兌現金額(未稅)"], errors="coerce").fillna(0)

    return df[["客戶編號", "已兌現金額(稅後)", "已兌現金額(未稅)"]]


def load_level_tables(cfg: CountryConfig) -> pd.DataFrame:
    """
    讀「月結額度」規範表（業務邏輯）
    - 檔案存在性由 get_latest_excel_or_fail 保證
    """

    path = get_latest_excel_or_fail(
        cfg.level_dir,
        cfg.level_keyword,
        country=cfg.code,
        usage="月結額度規範"
    )

    level_df = pd.read_excel(path, header=1)

    # 取 A,B,C 欄
    t = level_df.iloc[:, 0:3].copy()
    t.columns = ["級別", "累計兌現金額", "交易額度"]
    t = t.dropna(how="all")

    t["累計兌現金額"] = (
        pd.to_numeric(t["累計兌現金額"], errors="coerce")
        .fillna(0)
        .astype("Int64")
    )
    t["交易額度"] = (
        pd.to_numeric(t["交易額度"], errors="coerce")
        .astype("Int64")
    )

    return t



# 5) Strategy（底層可替換）
class BaseStrategy:
    """
    預設策略（共用，穩定版）
    - 可直接整串覆蓋
    - 不會因 HK / 欄位缺失 / 空資料炸掉
    - 未來各國分歧 → 繼承後 override 單一 method
    """

    SALES_YEARS = 2  # 預設兩年銷貨

    
    # 共用額度組別分類
    
    def classify_shared_group_category(
        self,
        shared_df: pd.DataFrame,
        cfg: CountryConfig
    ) -> pd.DataFrame:

        if shared_df is None or shared_df.empty:
            return pd.DataFrame(
                columns=["客戶編號", "共用額度組別", "額度分配比例", "共用額度組別分類"]
            )

        df = shared_df.copy()

        if "共用額度組別分類" not in df.columns:
            df["共用額度組別分類"] = np.nan

        df["共用額度組別分類"] = df["共用額度組別分類"].replace(
            ["", "nan", "None"], np.nan
        )

        # 預設：用客戶編號前兩碼
        df["共用額度組別分類"] = df.apply(
            lambda r: r["客戶編號"][:2]
            if pd.isna(r["共用額度組別分類"])
            else r["共用額度組別分類"],
            axis=1,
        )

        return df

    
    # 近 N 年是否有交易
    
    def apply_recent_sales_flag(
        self,
        df: pd.DataFrame,
        sales_sum: pd.DataFrame
    ) -> pd.DataFrame:

        out = df.merge(sales_sum, on="客戶編號", how="left")

        out["近2年有交易(現金可轉月結)"] = np.where(
            out["未稅本位幣"] > 0,
            "V",
            "X"
        )

        out["近2年有交易(現金可轉月結)"] = out[
            "近2年有交易(現金可轉月結)"
        ].fillna("X")

        return out

    
    # 共用額度：已兌現金額組合計
    
    def apply_shared_group_redeemed_rollup(
        self,
        df: pd.DataFrame
    ) -> pd.DataFrame:

        if df is None or df.empty:
            return df

        if "共用額度組別" not in df.columns:
            return df

        out = df.copy()

        valid = out["共用額度組別"].notna()
        if not valid.any():
            return out

        group_sum = (
            out.loc[valid]
            .groupby("共用額度組別")["已兌現金額(未稅)"]
            .sum()
            .reset_index()
            .rename(columns={"已兌現金額(未稅)": "共用組合計"})
        )

        out = out.merge(group_sum, on="共用額度組別", how="left")

        out["已兌現金額(未稅)"] = np.where(
            out["共用組合計"].notna(),
            out["共用組合計"],
            out["已兌現金額(未稅)"]
        )

        out.drop(columns=["共用組合計"], inplace=True)

        return out

    
    # 級別 → 交易額度
    
    def assign_credit_limit(
        self,
        df: pd.DataFrame,
        level_table: pd.DataFrame
    ) -> pd.DataFrame:

        out = df.copy()

        def calc_limit(row):
            used = row.get("已兌現金額(未稅)", 0)
            ratio = row.get("額度分配比例", np.nan)

            credit = 0
            for _, lv in level_table.iterrows():
                if used >= lv["累計兌現金額"]:
                    credit = lv["交易額度"]
                else:
                    break

            if pd.isna(ratio):
                return int(round(credit, 0))

            return int(round(credit * ratio, 0))

        out["交易額度"] = out.apply(calc_limit, axis=1)
        out["交易額度"] = out["交易額度"].fillna(0).astype(int)
        out.loc[out["交易額度"] == 0, "交易額度"] = 1

        return out

    
    # 額度變化 / 客戶分類（核心穩定版）
    
    def classify_customer(
        self,
        df: pd.DataFrame,
        not_monthly_list: list
    ) -> pd.DataFrame:

        out = df.copy()
        out["額度變化數值"] = out["交易額度"] - out["前月交易額度"]

        def rule(row):
            credit_ctrl = row.get("信用管制說明")
            bad_debt = row.get("呆帳客戶")
            overdue = row.get("當月逾期金額(依會計每月20號的逾期表)", 0)
            recent = row.get("近2年有交易(現金可轉月結)")
            cust_id = row.get("客戶編號")
            diff = row.get("額度變化數值", 0)

            # ---- 額度不變（優先） ----
            if credit_ctrl == "呆賬管制":
                return ("額度不變(管制)", "額度不變")

            if pd.notna(bad_debt) and str(bad_debt).strip():
                return ("額度不變(呆賬)", "額度不變")

            if overdue > 0:
                return ("額度不變(逾期)", "額度不變")

            if recent == "X":
                return ("額度不變(凍結)", "額度不變")

            if cust_id in not_monthly_list:
                return ("額度不變(共額不轉)", "額度不變")

            # ---- 額度變動 ----
            if diff < 0:
                return ("額度下降", "額度下降")

            if diff > 0:
                return ("額度上升", "額度上升")

            return ("額度不變", "額度不變")

        result = out.apply(rule, axis=1)

        out[["客戶分類", "額度變化"]] = pd.DataFrame(
            result.tolist(),
            index=out.index
        )

        # 額度不變 → 回填前月
        mask = out["額度變化"].str.contains("額度不變", na=False)
        out.loc[mask, "交易額度"] = out.loc[mask, "前月交易額度"]
        out.loc[mask, "額度變化數值"] = 0

        return out

    
    # 自動化(新月結 / 額度增加)
    
    def apply_automation(self, df: pd.DataFrame) -> pd.DataFrame:

        out = df.copy()
        out["新信用管制說明"] = out["信用管制說明"]

        # 現金客戶且額度變動 → 月結
        mask_cash = (
            (out["信用管制說明"] == "現金客戶") &
            (out["額度變化"].isin(["額度上升", "額度下降"]))
        )
        out.loc[mask_cash, "新信用管制說明"] = "月結"

        # 原月結，額度上升
        mask_inc = (
            (out["信用管制說明"] == "月結") &
            (out["額度變化"] == "額度上升")
        )
        out.loc[mask_inc, "客戶分類"] = "自動化(額度增加)"

        # 現金 → 月結（新月結）
        mask_new = (
            (out["信用管制說明"] == "現金客戶") &
            (out["額度變化"] == "額度上升")
        )
        out.loc[mask_new, "客戶分類"] = "自動化(新月結客戶)"

        return out


# 若未來某國要特例，照這樣做：
# class JPStrategy(BaseStrategy):
#     SALES_YEARS = 3
#     def classify_customer(...): override


STRATEGY_REGISTRY: Dict[str, BaseStrategy] = {
    "HK": BaseStrategy(),
    "JP": BaseStrategy(),
    "PH": BaseStrategy(),
    "ID": BaseStrategy(),
    "VN": BaseStrategy(),
    "TH": BaseStrategy(),
}


# 6) Export（共用）
def export_main_outputs(df: pd.DataFrame, cfg: CountryConfig, today_yyyymmdd: str) -> str:
    """
    產出「總額度判定」excel
    """
    folder = os.path.join(cfg.base_path, today_yyyymmdd)
    ensure_dir(folder)

    out_path = os.path.join(folder, f"{cfg.code}_總額度判定{today_yyyymmdd}.xlsx")
    df.to_excel(out_path, index=False)
    return out_path


def export_for_huiru(df: pd.DataFrame, cfg: CountryConfig, today_yyyymmdd: str) -> str:
    """
    - Sheet: 改月結 / 額度變更 / 共用 / 轉現金
    """
    folder = os.path.join(cfg.base_path, today_yyyymmdd)
    ensure_dir(folder)

    target_classifications = ["自動化(新月結客戶)", "額度下降", "自動化(額度增加)"]
    filtered_df = df[df["客戶分類"].isin(target_classifications)].copy()

    # 額度變更
    credit_change_columns = [
        "客戶編號",
        "信用控制範圍",
        "信用額度",
        "風險種類",
        "客戶信用群組",
        "前次信用額度(比對用)",
        "GOOGLE申請",
        "比對上次",
        "共用額度組別",
        "備註",
        "變更",
    ]
    credit_change = pd.DataFrame(columns=credit_change_columns)
    credit_change["客戶編號"] = filtered_df["客戶編號"].values
    credit_change["信用額度"] = filtered_df["交易額度"].values
    credit_change["前次信用額度(比對用)"] = filtered_df["前月交易額度"].values
    credit_change["共用額度組別"] = filtered_df.get("共用額度組別", pd.Series([pd.NA] * len(filtered_df))).values
    credit_change["備註"] = filtered_df["客戶分類"].values

    # 改月結
    new_monthly_df = df[df["客戶分類"] == "自動化(新月結客戶)"].copy()
    new_monthly_columns = [
        "客戶代碼",
        "公司代碼",
        "銷售組織",
        "配銷通路",
        "部門",
        "搜尋條件 2",
        "結帳日",
        "寄單日",
        "帳單寄送方式",
        "付款方式",
        "信用管制",
        "付款條件碼_K",
        "付款條件碼_V",
    ]
    new_monthly_table = pd.DataFrame(columns=new_monthly_columns)
    new_monthly_table["客戶代碼"] = new_monthly_df["客戶編號"].values
    new_monthly_table["搜尋條件 2"] = datetime.today().replace(day=1).strftime("%Y-%m-%d")
    new_monthly_table["結帳日"] = "結帳日30日/寄單日5日"
    new_monthly_table["寄單日"] = "票期5040"
    new_monthly_table["帳單寄送方式"] = "99 其他"
    new_monthly_table["付款方式"] = "02 匯款"
    new_monthly_table["信用管制"] = "月結"

    file_name = f"{cfg.code}_sap大批匯入(給惠茹)-{today_yyyymmdd}.xlsx"
    full_path = os.path.join(folder, file_name)

    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        new_monthly_table.to_excel(writer, sheet_name="改月結", index=False)
        credit_change.to_excel(writer, sheet_name="額度變更", index=False)

        workbook = writer.book
        ws1 = workbook.create_sheet(title="共用")
        ws1.append(["客戶編號", "共用額度組別", "備註"])
        ws2 = workbook.create_sheet(title="轉現金")
        ws2.append(["客戶編號", "信用管制", "信用額度"])
        workbook.save(full_path)

    return full_path


def export_for_sales(df: pd.DataFrame, cfg: CountryConfig, today_yyyymmdd: str) -> str:

    folder = os.path.join(cfg.base_path, today_yyyymmdd)
    ensure_dir(folder)

    target_classifications = ["自動化(新月結客戶)", "額度下降", "自動化(額度增加)"]
    filtered_df = df[df["客戶分類"].isin(target_classifications)].copy()

    columns_for_sales = [
        "客戶編號",
        "名稱1",
        "搜尋條件1",
        "大區",
        "區域",
        "類別",
        "呆帳客戶",
        "客戶分類",
        "新信用管制說明",
        "交易額度",
        "月結客戶維持現金客戶(業務確認)",
        "維持原因(業務填寫)",
        "收款聯絡人",
        "收款聯絡人電話",
        "個人LINE網址",
        "群组LINE網址",
        "收款備註",
        "業務送帳單/收款需填原因",
        "新月結客結帳日/寄單日",
        "票期(統一5040不可動)",
        "新月結客帳單寄送方式",
        "新月結客付款方式",
        "月結起始日",
        "小區",
    ]
    final_sales_table = pd.DataFrame(columns=columns_for_sales)

    source_columns = [
        "客戶編號",
        "名稱1",
        "搜尋條件1",
        "大區",
        "區域",
        "類別",
        "呆帳客戶",
        "客戶分類",
        "新信用管制說明",
        "交易額度",
        "新月結客結帳日/寄單日",
        "票期(統一5040不可動)",
        "新月結客帳單寄送方式",
        "新月結客付款方式",
        "月結起始日",
        "小區",
    ]

    for col in source_columns:
        if col in filtered_df.columns:
            final_sales_table[col] = filtered_df[col].values

    # 收款聯絡資訊補齊
    info_path = kd.get_latest_excel(COLLECTION_INFO_DIR, COLLECTION_INFO_KEYWORD)
    info_df = pd.read_excel(info_path)
    info_df = info_df[["客戶", "收款連絡人", "收款連絡人電話", "收款人LINE ID", "群組LINE ID", "收款備註"]].drop_duplicates(
        subset="客戶", keep="first"
    )

    final_sales_table["客戶編號"] = safe_strip_series(final_sales_table["客戶編號"])
    info_df["客戶"] = safe_strip_series(info_df["客戶"])

    info_dict = info_df.set_index("客戶").to_dict(orient="index")

    final_sales_table["收款聯絡人"] = final_sales_table["客戶編號"].map(lambda x: info_dict.get(x, {}).get("收款連絡人", pd.NA))
    final_sales_table["收款聯絡人電話"] = final_sales_table["客戶編號"].map(lambda x: info_dict.get(x, {}).get("收款連絡人電話", pd.NA))
    final_sales_table["個人LINE網址"] = final_sales_table["客戶編號"].map(lambda x: info_dict.get(x, {}).get("收款人LINE ID", pd.NA))
    final_sales_table["群组LINE網址"] = final_sales_table["客戶編號"].map(lambda x: info_dict.get(x, {}).get("群組LINE ID", pd.NA))
    final_sales_table["收款備註"] = final_sales_table["客戶編號"].map(lambda x: info_dict.get(x, {}).get("收款備註", pd.NA))

    final_sales_table["月結起始日"] = datetime.today().replace(day=1).strftime("%Y-%m-%d")

    out_path = os.path.join(folder, f"{cfg.code}_新交易額度客戶明細-業務{today_yyyymmdd}.xlsx")
    final_sales_table.to_excel(out_path, index=False)
    return out_path

def get_latest_excel_or_fail(folder, keyword, *, country, usage):

    print(f" [{country}] 取得 {usage}")
    print(f"    路徑: {folder}")
    print(f"    關鍵字: {keyword}")

    path = kd.get_latest_excel(folder, keyword)

    if path is None:
        raise RuntimeError(
            f"\n [{country}] 缺少必要檔案：{usage}\n"
            f"   路徑: {folder}\n"
            f"   關鍵字: {keyword}\n"
        )

    return path

# 7) Engine（主流程：只負責調度）
def run_country(cfg: CountryConfig, strategy: BaseStrategy) -> Dict[str, str]:
    """
    回傳輸出檔路徑 dict
    """
    today_yyyymmdd = datetime.now().strftime("%Y%m%d")

    # 1) ZSD31
    zsd31 = load_zsd31(cfg)

    # 2) 呆帳
    bad_debt = load_bad_debt(cfg)
    df = zsd31.merge(bad_debt, on="客戶編號", how="left")

    # 3) 共用額度登記
    shared = load_shared_credit_registry(cfg)
    shared = strategy.classify_shared_group_category(shared, cfg)
    df = df.merge(shared, on="客戶編號", how="left")
    df = df.drop_duplicates(subset=["客戶編號"], keep="first")

    # 4) 逾期
    overdue = load_overdue(cfg)
    df = df.merge(overdue, on="客戶編號", how="left")
    df["當月逾期金額(依會計每月20號的逾期表)"] = df["當月逾期金額(依會計每月20號的逾期表)"].fillna(0).astype(int)

    # 5) 兩年銷貨
    sales_sum = load_sales_recent(cfg, years=strategy.SALES_YEARS)
    df = strategy.apply_recent_sales_flag(df, sales_sum)

    # 6) 已兌現金額
    redeemed = load_redeemed_amount(cfg)
    df = df.merge(redeemed, on="客戶編號", how="left")
    df["已兌現金額(稅後)"] = df["已兌現金額(稅後)"].fillna(0)
    df["已兌現金額(未稅)"] = df["已兌現金額(未稅)"].fillna(0)

    # 7) 共用組合計（把已兌現金額改成組合計）
    df = strategy.apply_shared_group_redeemed_rollup(df)

    # 8) 規範表 -> 交易額度
    level_table = load_level_tables(cfg)
    df = strategy.assign_credit_limit(df, level_table)

    # 9) 不轉月結名單
    not_monthly_list = load_not_monthly_list(cfg)

    # 10) 額度變化 / 客戶分類
    df = strategy.classify_customer(df, not_monthly_list)

    # 11) 自動化(新月結/額度增加)
    df = strategy.apply_automation(df)

    # 12) 後處理：新月結客戶固定欄位
    df = df.drop_duplicates(subset=["客戶編號"], keep="last").copy()

    d1 = datetime.today().strftime("%Y/%m/01")
    df.loc[df["客戶分類"] == "自動化(新月結客戶)", "月結起始日"] = d1
    df["月結起始日"] = pd.to_datetime(df["月結起始日"], errors="coerce", utc=True).dt.date

    df.loc[df["客戶分類"] == "自動化(新月結客戶)", "新月結客結帳日/寄單日"] = "結帳日30日/寄單日5日"
    df.loc[df["客戶分類"] == "自動化(新月結客戶)", "票期(統一5040不可動)"] = "票期5040"
    df.loc[df["客戶分類"] == "自動化(新月結客戶)", "新月結客帳單寄送方式"] = "99 其他"
    df.loc[df["客戶分類"] == "自動化(新月結客戶)", "新月結客付款方式"] = "02 匯款"

    # 13) 優先排序：自動化(新月結客戶) 放最上
    priority = {"自動化(新月結客戶)": 0}
    df = df.sort_values(by="客戶分類", key=lambda s: s.map(priority).fillna(1))

    # 14) 輸出
    out_main = export_main_outputs(df, cfg, today_yyyymmdd)
    out_huiru = export_for_huiru(df, cfg, today_yyyymmdd)
    out_sales = export_for_sales(df, cfg, today_yyyymmdd)

    return {"main": out_main, "huiru": out_huiru, "sales": out_sales}


def main():
    targets = ["HK", "JP", "PH", "ID", "VN", "TH"]
    # targets = ["JP"]
    results = {}

    for code in targets:
        cfg = COUNTRY_CONFIGS[code]
        strategy = STRATEGY_REGISTRY[code]

        print(f"執行 {code} 自動化月結")
        try:
            out = run_country(cfg, strategy)
            results[code] = out
            print(f" {code} 完成")
            print(f"   - main : {out['main']}")
            print(f"   - huiru: {out['huiru']}")
            print(f"   - sales: {out['sales']}")
        except Exception as e:
            print(f" {code} 失敗：{e}")

    print("\n===== SUMMARY =====")
    for code, out in results.items():
        print(code, out)


if __name__ == "__main__":
    main()


