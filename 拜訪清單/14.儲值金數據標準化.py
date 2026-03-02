
import pandas as pd
import numpy as np
import pyodbc
import json
import requests
from datetime import datetime, timedelta, date
import pymysql
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from math import ceil
import os
import common  as kd
from pathlib import Path
from openpyxl import load_workbook


path_old = fr"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\儲值金\儲值金名單\{str((datetime.today()+relativedelta(months=-1)).date().strftime("%Y-%m"))}\\"
path_old =  kd.get_latest_excel(path_old,'儲值金名單(模板)')

template_path = path_old 
dest_folder = fr"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\★每月業務電拜訪★\共用CRM資訊\Temp\電拜訪清單資料\儲值金\儲值金名單\{datetime.today().strftime('%Y-%m')}\\"
os.makedirs(dest_folder, exist_ok=True)

# month_cn = f"{datetime.today().month}月"
# dest_path = os.path.join(dest_folder, f"{month_cn}儲值金名單(模板).xlsx")

# wb = load_workbook(template_path, data_only=False, keep_vba=False)

# for ws in wb.worksheets:
#     max_row, max_col = ws.max_row, ws.max_column
#     if max_row <= 1:
#         continue  
#     for r in range(2, max_row + 1):
#         for c in range(1, max_col + 1):
#             cell = ws.cell(row=r, column=c)
#             cell.value = None
#             cell.hyperlink = None
#             cell.comment = None
# wb.save(dest_path)
# print(f"已建立：{dest_path}")



# 獲取儲值金達10萬的名單
path_over10 =  kd.get_latest_excel(dest_folder,'CD類儲值金客戶')
over10 = pd.read_excel(path_over10,  dtype=str)
over10 = over10[['公司代號', '資料區域群組名稱', '公司型態', '專案到期日']]
over10['主旨'] = '(經營)儲值金邀約-達10萬'

# 獲取儲值金未達10萬的名單
path_below10 =  kd.get_latest_excel(dest_folder,'CD類儲值金未滿10萬客戶')
below10 = pd.read_excel(path_below10,  dtype=str)
below10 = below10[['公司代號', '資料區域群組名稱', '公司型態', '專案到期日']]
below10['主旨'] = '(經營)儲值金邀約-當月3案場以上'

# 合併兩個DataFrame
K_invite = pd.concat([over10, below10], ignore_index=True).drop_duplicates(['公司代號'])

# 獲取所有客戶資料
account = kd.get_data_from_CRM("""SELECT accountCode__c,accountName 公司名稱, dimDepart.departName 資料區域名稱 FROM account  where dimDepart.departName like '%TW%'  """)


# 合併客戶資料
K_invite_total = pd.merge(K_invite, account, left_on='公司代號', right_on='accountCode__c', how='left')

K_invite_total_ok = K_invite_total.loc[
    K_invite_total["資料區域群組名稱"].eq(K_invite_total["資料區域名稱"])
    & ~K_invite_total["資料區域群組名稱"].astype(str).str.contains("Z", case=False, na=False)
    & ~K_invite_total["資料區域名稱"].astype(str).str.contains("Z", case=False, na=False)]
   
K_invite_total_ok.to_excel(dest_folder + "儲值金數據留存_" + str(datetime.today().strftime("%Y.%m.%d")) + ".xlsx",index=False)


K_invite_total_ok = (K_invite_total_ok [['公司代號','資料區域名稱','專案到期日','公司名稱']]
      .assign(簡稱=lambda d: d["公司名稱"].fillna("").str[:2],處理人員="", 處理進度="",異動時間="")
    [["公司代號","簡稱","資料區域名稱","專案到期日","處理人員","處理進度","異動時間","公司名稱"]])

K_invite_total_ok.to_clipboard(index=False, header=False, excel=True)
print("已將數據複製到剪貼簿，大人! 請去企微文檔中建立新的表格並將數據 粘貼到Excel中, 記得選擇保持原樣。")
# 點擊上個月sheet, 選擇創建副本,點擊確定, 修改名稱與到期日