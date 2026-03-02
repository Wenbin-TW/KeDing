
import os
import sys
import time
import logging
import traceback
from pathlib import Path
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import requests
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

custom_path = Path(r"C:\Users\TW0002.TPTWKD\Desktop\項目整理")
sys.path.append(str(custom_path))
import common as kd

user_df = kd.get_data_from_CRM ('select id, name, dimDepart from user')
user_df[user_df['name'] == '易家婕']


sample_date = int(datetime(2025,1, 1).timestamp() * 1000)
型錄發放日期 = int(datetime(2025,12, 29).timestamp() * 1000)
型錄寄送截止日期 = int(datetime(2026,1, 28).timestamp() * 1000)
專案型錄寄送截止日期 = int(datetime(2025,12, 29).timestamp() * 1000)
today = pd.to_datetime(datetime.today().strftime('%Y-%m-%d'))
year_ago_one = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_one = pd.to_datetime((datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_two = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_three = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_six = pd.to_datetime((datetime.today() - relativedelta(months=6) + relativedelta(days=1)).date()).timestamp()*1000
three_days_ago = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=-3)).timestamp() * 1000)
today_begin = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=0)).timestamp() * 1000)
today_end = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)).timestamp() * 1000)
five_days_ago = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=-5)).timestamp() * 1000)
seven_days_ago = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=-7)).timestamp() * 1000)
today_str = datetime.now().strftime("%Y-%m-%d")
Tasks_df = kd.get_data_from_CRM(f'''
    select
        id,
        name,
        createdAt,
        customItem42__c,
        customItem10__c,
        customItem3__c,
        approvalStatus,
        customItem191__c,
        entityType,
        customItem8__c,
        customItem42__c.name 客戶關係聯絡人代號,
        customItem120__c,
        customItem121__c,
        customItem42__c.contactPhone__c__c 客戶手機號,
        customItem42__c.contactCode__c__c 客戶代號
    from customEntity14__c
    where entityType in ('2904963933786093','3028348436713387')
      and createdAt >= {month_ago_three}
      and customItem120__c >= {today_begin}
      and customItem120__c <  {today_end}
''')
def is_empty_like(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return True
    if isinstance(x, (list, tuple, set, dict)):
        return len(x) == 0
    s = str(x).strip().lower()
    return s in {'', 'none', 'null', 'nan', '[]'}
Tasks_df['pool_type'] = 'OTHER'
Tasks_df.loc[
    (Tasks_df['customItem10__c'] == '3414568030034443')
    & Tasks_df['customItem121__c'].astype(str).str.contains("第一次電訪", na=False)
    & Tasks_df['customItem3__c'].astype(str).str.contains("拒絕客戶", na=False),
    'pool_type'
] = 'A'
Tasks_df.loc[
    (Tasks_df['customItem10__c'] == '3414568030034443')
    & Tasks_df['customItem121__c'].astype(str).str.contains("第一次電訪", na=False)
    & ~Tasks_df['customItem3__c'].astype(str).str.contains("拒絕客戶", na=False),
    'pool_type'
] = 'B'
Tasks_df.loc[
    (Tasks_df['customItem10__c'] == '3414568030034443')
    & Tasks_df['customItem121__c'].astype(str).str.contains("未接1", na=False)
    & Tasks_df['customItem191__c'].apply(is_empty_like),
    'pool_type'
] = 'C'
Tasks_df.loc[
    (Tasks_df['customItem10__c'] == '3884504824139673')
    & Tasks_df['customItem121__c'].astype(str).str.contains("第一次電訪", na=False),
    'pool_type'
] = 'D'
cnt = Tasks_df['pool_type'].value_counts()

A = cnt.get('A', 0)
B = cnt.get('B', 0)
C = cnt.get('C', 0)
D = cnt.get('D', 0)
total = A + B + C + D

table_md = (
    "| 分類 | 條件 | 筆數 |\n"
    "| :---- | :----------------------------- | -----: |\n"
    f"| A類 | 第一次電訪 & 拒絕客戶 | **{A}** |\n"
    f"| B類 | 第一次電訪 & 非拒絕客戶 | **{B}** |\n"
    f"| C類 | 未接1 & GC為空 | **{C}** |\n"
    f"| D類 | 二面 & 第一次電訪 | **{D}** |\n"
    f"| **合計** |  | **{total}** |\n"
)

markdown_text = (
    f"# K大名單日報（{today_str}）\n"
    f"> 篩選條件：GC 外撥\n\n"
    + table_md
)

payload = {
    "msgtype": "markdown_v2",
    "markdown_v2": {
        "content": markdown_text
    }
}
webhook_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=5613aaa0-2e6a-4365-99b2-01dc1e21189a"

resp = requests.post(webhook_url, json=payload, timeout=10)

try:
    rj = resp.json()
    if rj.get("errcode") == 0:
        print("播報成功")
    else:
        print(f"播報失敗：{rj}")
except Exception as e:
    print(f"回傳解析失敗：{e}")



def setup_logging():
    log_dir = r"C:\Users\TW0002.TPTWKD\Desktop\Wenbin\log"
    if not os.path.exists(log_dir):
        try:
            os.makedirs(log_dir)
        except Exception as e:
            print(f"無法建立 Log 資料夾: {e}")
            return None
    today_str = datetime.now().strftime("%Y-%m-%d")
    log_filename = f"K大播報{today_str}.log"
    log_path = os.path.join(log_dir, log_filename)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_path, encoding='utf-8'), # 寫入檔案
            logging.StreamHandler(sys.stdout)                # 同時印在控制台
        ]
    )
    return log_path


current_log_path = setup_logging()
logging.info("==========================================")
logging.info(f"程式啟動，Log 檔案路徑: {current_log_path}")
logging.info("==========================================")
webhook_url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=20ff0bf1-b6e0-49ae-ae48-806a1b5ff0ef"
try:
    logging.info("正在獲取基礎資料 (CRM電訪人員名單)...")
    start_t = time.time()
    sales_total = kd.get_data_from_CRM(f'''
            select customItem1__c customItem10__c, customItem3__c 預估通數,customItem1__c.dimDepart dimDepart,customItem4__c,
            customItem2__c 電訪人員類型,customItem1__c.name 電訪人員
            from customEntity42__c
            where customItem5__c = 1
            ''')
    承攬 = sales_total.loc[sales_total['customItem4__c'].astype(str).str.contains("承攬")]
    contractor_ids = 承攬['customItem10__c'].tolist()
    duration = time.time() - start_t
    logging.info(f"基礎資料獲取成功，耗時 {duration:.2f} 秒，今日電訪人數: {len(contractor_ids)}")
except Exception as e:
    logging.error(f"基礎資料獲取失敗: {e}")
    logging.error(traceback.format_exc())
def generate_report(report_title, status_code, folder_name, extra_filter_func=None):
    """
    Args:
        report_title (str): 報表標題
        status_code (int): SQL查詢中的 customItem121__c 代碼
        folder_name (str): 存檔資料夾名稱
        extra_filter_func (function, optional): 額外篩選邏輯
    """
    logging.info(f"--- [START] 開始執行任務: {report_title} (Status: {status_code}) ---")
    today_begin = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=0)).timestamp() * 1000)
    today_end = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)).timestamp() * 1000)
    today_str = datetime.now().strftime("%Y-%m-%d")

    try:
        logging.info(f"[{report_title}] 正在執行 SQL 查詢...")
        sql_start = time.time()
        
        sql_query = f'''    
            select name 交辦編號, createdAt 創建日期, customItem10__c,customItem10__c.name 執行人, customItem3__c 工作主旨, approvalStatus 審批狀態,customItem11__c.accountName 公司名稱,
            entityType 業務類型,customItem8__c 執行狀態,customItem42__c.name 客戶關係聯絡人代號,customItem120__c 預計完成日期, customItem121__c 電訪狀態,
            customItem42__c.contactPhone__c__c 客戶手機號, customItem42__c.contactCode__c__c 客戶代號,customItem277__c 下次聯絡時間,customItem56__c 連絡人姓名
            from customEntity14__c 
            where entityType in ('2904963933786093','3028348436713387') and customItem121__c = {status_code} and customItem8__c = 1
        '''
        Tasks_df_total = kd.get_data_from_CRM(sql_query)
        sql_duration = time.time() - sql_start
        logging.info(f"[{report_title}] SQL 查詢完成，耗時 {sql_duration:.2f} 秒，原始筆數: {len(Tasks_df_total)}")
        Tasks_df_total = Tasks_df_total[~Tasks_df_total['customItem10__c'].isin(contractor_ids)]
        Tasks_df = Tasks_df_total.loc[~Tasks_df_total['公司名稱'].astype(str).str.contains("測試")]
        logging.info(f"[{report_title}] 基礎過濾完成 (排除承攬/測試)，剩餘筆數: {len(Tasks_df)}")
        if extra_filter_func:
            logging.info(f"[{report_title}] 執行額外篩選邏輯...")
            Tasks_df = extra_filter_func(Tasks_df)
            logging.info(f"[{report_title}] 額外篩選後筆數: {len(Tasks_df)}")
        logging.info(f"[{report_title}] 正在處理日期格式與欄位...")
        Tasks_df['業務類型'] = np.where(Tasks_df['業務類型'].astype(str) == '3028348436713387', '每日K大', '寄後電訪')
        Tasks_df = kd.convert_to_date(Tasks_df, '創建日期')
        Tasks_df = kd.convert_to_date(Tasks_df, '下次聯絡時間')
        
        cols = ['交辦編號','創建日期','下次聯絡時間','預計完成日期','執行人','連絡人姓名','客戶手機號','工作主旨','公司名稱','業務類型','客戶關係聯絡人代號','客戶代號']
        Tasks_df = Tasks_df[cols]
        Tasks_df_before = Tasks_df[(pd.to_numeric(Tasks_df["預計完成日期"], errors="coerce") < today_begin)].reset_index(drop=True)
        Tasks_df_today = Tasks_df[
            (pd.to_numeric(Tasks_df["預計完成日期"], errors="coerce") >= today_begin) &
            (pd.to_numeric(Tasks_df["預計完成日期"], errors="coerce") < today_end)].reset_index(drop=True)
        Tasks_df_before = kd.convert_to_date(Tasks_df_before, '預計完成日期')
        Tasks_df_today = kd.convert_to_date(Tasks_df_today, '預計完成日期')
        summary = pd.merge(
            Tasks_df_before.groupby('執行人').size().reset_index(name='未完成任務數'), 
            Tasks_df_today.groupby('執行人').size().reset_index(name='今日任務數'), 
            on='執行人', how='outer'
        ).fillna(0).astype({'未完成任務數':int, '今日任務數':int})

        logging.info(f"[{report_title}] 準備發送 Webhook 通知...")
        
        text_lines = [f"** {report_title} 交辦情況（{today_str}）**"]
        for _, row in summary.iterrows():
            text_lines.append(
                f"> <font color=\"info\">{row['執行人']}</font>｜昨日剩餘:<font color=\"red\">{int(row['未完成任務數'])}</font>｜今日待執行:<font color=\"blue\">{int(row['今日任務數'])}</font>"
            )
        markdown_msg = "\n".join(text_lines)
        payload = {"msgtype": "markdown", "markdown": {"content": markdown_msg}}
        
        try:
            resp = requests.post(webhook_url, json=payload, timeout=10)
            logging.info(f"[{report_title}] Webhook 發送狀態碼: {resp.status_code}")
        except Exception as e:
            logging.error(f"[{report_title}] Webhook 發送失敗: {e}")
        output_path = rf"Z:\18_各部門共享區\15_數據中心課\文斌\{folder_name}"
        if not os.path.exists(output_path):
             logging.warning(f"[{report_title}]  目標資料夾不存在: {output_path} (可能是 Z 槽未掛載)")
             try:
                 os.makedirs(output_path, exist_ok=True)
             except Exception as path_err:
                 logging.error(f"[{report_title}] 無法建立資料夾: {path_err}")

        file_name = f"{report_title}_{today_str}.xlsx"
        full_path = os.path.join(output_path, file_name)

        logging.info(f"[{report_title}] 正在寫入 Excel: {full_path}")
        with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
            Tasks_df_before.to_excel(writer, index=False, sheet_name="昨日未完成")
            Tasks_df_today.to_excel(writer, index=False, sheet_name="今日待執行")
            summary.to_excel(writer, index=False, sheet_name="執行人匯總")
        logging.info(f"[{report_title}] Excel 格式美化中...")
        wb = load_workbook(full_path)
        for sheet_name in ["昨日未完成", "今日待執行", "執行人匯總"]:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            max_row, max_col = ws.max_row, ws.max_column
            if max_row < 2: continue 
            
            table_ref = f"A1:{chr(64 + max_col)}{max_row}"
            table = Table(displayName=f"Table_{sheet_name}_{int(time.time())}", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_length:
                        max_length = len(val)
                ws.column_dimensions[column].width = min(max(max_length + 2, 15), 70)

        wb.save(full_path)
        logging.info(f"[{report_title}]  任務圓滿完成！檔案已儲存。")
    
    except Exception as e:
        logging.error(f"[{report_title}]  執行過程中發生嚴重錯誤！")
        logging.error(traceback.format_exc()) # 記錄完整錯誤堆疊
        raise e # 拋出錯誤
logging.info(">>> 準備執行 任務 1: 已邀約時間未定")


try:
    def task1_filter(df):
        df = df.loc[df['執行狀態'].astype(str).str.contains("等待")]
        df = df.loc[df['電訪狀態'].astype(str).str.contains("已邀約")]
        return df

    generate_report(
        report_title="已邀約時間未定",
        status_code=7,
        folder_name="已邀約時間未定",
        extra_filter_func=task1_filter
    )
    logging.info("任務 1 執行成功！")

except Exception as e:
    logging.error(f"任務 1 失敗: {e}")
logging.info("等待 20 秒讓系統緩衝...")
time.sleep(20) 
logging.info(">>> 準備執行 任務 2: K大視訊未上線")
try:
    logging.info("正在撈取 K大視訊 資料...")
    
    generate_report(
        report_title="K大視訊未上線",
        status_code=12,
        folder_name="K大視訊未上線",
        extra_filter_func=None
    )
    logging.info("任務 2 執行成功！")

except Exception as e:
    error_msg = traceback.format_exc()
    logging.error(f"任務 2 崩潰 (Critical Error):\n{error_msg}")
    err_payload = {
        "msgtype": "text", 
        "text": {"content": f" K大視訊報表執行失敗\n\n請檢查 Log: {current_log_path}\n原因：\n{str(e)}"}
    }
    try:
        requests.post(webhook_url, json=err_payload)
    except:
        pass

logging.info("=== 所有排程執行結束 ===")