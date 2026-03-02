

import os
import glob
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from datetime import datetime
import pymysql
from dateutil.relativedelta import relativedelta
from pathlib import Path
import sys
custom_path = Path(r"C:\Users\TW0002.TPTWKD\Desktop\項目整理")
sys.path.append(str(custom_path))
import common as kd
import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


業管_客資_df = kd.get_data_from_MSSQL('''
            SELECT iNo	流水編號 ,KUNNR	客戶編號,KNAME	客戶名稱,SORT1	搜尋條件1,STREET	街道1,BEZEI	'客戶群組1-說明'
            ,STR_SUPPL2	街道3,TEL_NUMBER	電話,VTEX8	信用管制說明
            ,VTEX3	稅分類說明,STCD1	統一編號,VTEX10	性能分類說明
            ,GFORM	專案價,TDLIN3	行銷註記,PANAME	員工姓名
            ,VTEX08	小群組說明,FAX_NUMBER	傳真,TDLIN1	銷售註記
            ,SORT2	搜尋條件2,ERDAT	建檔日期
            ,ERNAM	建立者,TDLIN2	出貨指示,COUNTRY	國家碼
            ,ZZCRM_CUSTOMER	銷售易公司代號
            ,ZZCRM_COSHORTNAME	銷售易公司名稱
            ,VKORG	銷售組織,VTEX1	銷售組織說明
            ,BZIRK	銷售地區,BZTXT	銷售地區說明
            ,VKBUR	銷售據點,BEZE1	銷售據點說明
            ,WAERS	幣別,KTOKD	客戶科目群組
            ,TXT30	客戶科目群組說明,NAME1	名稱1
            ,ZTERM	'付款條件碼(公司)',VTEX9	'付款條件說明(公司)'
            ,ZTERM_1	'付款條件碼(銷售)'
            ,KONDA	客戶價格群組,VTEX2	客戶價格群組說明
            ,VSBED	出貨條件,VTEXT1	出貨條件說明,INCO1	國貿條件
            ,BEZE2	國貿條件說明,INCO2	國貿條件2,PERNR	員工編號
            ,ADNAME	聯絡人,TEL_NO	聯絡人手機,KTGRD	科目指派群組
            ,VTEX6	科目指派群組說明,BRTXT	產業說明,KATR1	結帳日
            ,VTEX01	結帳日說明,KATR2	寄單日,VTEX02	寄單日說明
            ,VTEX03	放款日說明,KATR4	帳單寄送方式
            ,VTEX04	帳單寄送方式說明,KATR5	付款方式,VTEX05	付款方式說明
            ,KATR6	'地區(附加)',VTEX06	'地區說明(附加)',KATR7	大群組
            ,VTEX07	大群組說明,KATR8	小群組,KATR9	區域,BANKL	銀行代號
            ,BANKN	銀行帳號,KOINH	銀行戶名,REMARK	備註
            ,VTEXT2	專案價說明,GRUPP	客戶信用群組,KLIMK	交易額度
            ,POST_CODE1	郵遞區號,REGION	'地區(一般)'
            ,BEZE4	'地區說明(一般)',PERNR_ZM	主任秘書代號
            ,PERNR_ZM_NAME	主任秘書
            ,EMAIL	Email,ZZBANKAC	銀行專屬繳款帳號,BANKA	銀行名稱
            ,dCreateDate	資料建立日期
            FROM [SAPdb].[dbo].[ZSD31B]
            where vkorg= 'SG00' and KTOKD = 'YB01' and (KUNNR like 'SG%' or KUNNR like 'MY%')

            ''')


業管_客資_df = 業管_客資_df[['客戶編號', '名稱1', '搜尋條件1', '信用管制說明','國家碼','大群組說明', '小群組說明', '區域', '交易額度', '客戶價格群組'
                     ]].rename(columns={'大群組說明': '大區','小群組說明': '小區','交易額度': '前月交易額度','客戶價格群組': '類別'})
業管_客資_df = 業管_客資_df[~業管_客資_df['類別'].astype(str).str.contains('z', case=False, na=False)]

業管_客資_df['客戶編號'] = 業管_客資_df['客戶編號'].astype(str)
業管_客資_df['前月交易額度'] = pd.to_numeric(業管_客資_df['前月交易額度'], errors='coerce').fillna(0).astype(int)
ZSD31 = 業管_客資_df.copy()
呆帳_file = r"Z:\18_各部門共享區\01_會計課\11.海外子公司共用\呆帳"
呆帳_keyword = "SG呆帳明細"
呆帳_path = kd.get_latest_excel(呆帳_file, 呆帳_keyword)

呆帳_df = pd.read_excel(呆帳_path, sheet_name='呆帳明細', usecols=['客代']).drop_duplicates(subset=['客代'], keep='first')
呆帳_merge = pd.merge(ZSD31, 呆帳_df, left_on="客戶編號", right_on="客代", how='left').rename(columns={'客代': '呆帳客戶'})
共額_file = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度"
共額_keyword = "共用額度登記"
共額_path = kd.get_latest_excel(共額_file, 共額_keyword)

共額_sg = pd.read_excel(共額_path,sheet_name="SG",usecols=['客戶編號', '共用組數', '額度分配比例'])
共額_df = pd.concat([共額_sg], ignore_index=True)

共額_df = 共額_df.rename(columns={'共用組數': '共用額度組別'})
共額_df['客戶編號'] = 共額_df['客戶編號'].str.strip()
def classify_group(df):
    my_groups = (
        df[df['客戶編號'].str.startswith('MY')]
        ['共用額度組別']
        .unique()
    )
    df['共用額度組別分類'] = df['共用額度組別'].apply(
        lambda x: 'MY' if x in my_groups else 'SG'
    )
    return df

共額_df = classify_group(共額_df)


共額_df['共用額度組別'] = 共額_df['共用額度組別'].astype('Int64')  # 注意 Int64 是 pandas 的 nullable int 型別
共額_merge = pd.merge(呆帳_merge, 共額_df, on="客戶編號", how='left')
共額_merge = 共額_merge.drop_duplicates(subset=['客戶編號'], keep='first')
Overdue_file = r"Z:\18_各部門共享區\06_海外業務部\$$每月逾期貨款\逾期帳款報表-每月更新"
Overdue_keyword_sg = "月逾期帳款-SG("
Overdue_sheets = ['原幣']
Overdue_path_sg = kd.get_latest_excel(Overdue_file, Overdue_keyword_sg)
Overdue_data_sg = pd.read_excel(Overdue_path_sg, sheet_name=Overdue_sheets, header=4, usecols=['客戶', '合計未收'])
Overdue_clean_sg = Overdue_data_sg['原幣']
Overdue_clean_sg = Overdue_clean_sg[Overdue_clean_sg['客戶'].notna() & (Overdue_clean_sg['客戶'].astype(str).str.strip() != "")]




df5 = pd.concat([ Overdue_clean_sg])

df5 = df5.rename(columns={'客戶': '客戶編號'})
df5['合計未收'] = df5['合計未收'].astype('int')
df5['當月逾期金額(依會計每月20號的逾期表)'] = df5['合計未收'] / 1.09
df5.loc[df5['合計未收'] < 0, '當月逾期金額(依會計每月20號的逾期表)'] = 0
df5['當月逾期金額(依會計每月20號的逾期表)'] = df5['當月逾期金額(依會計每月20號的逾期表)'].astype('float') + 0.00000001
df5['當月逾期金額(依會計每月20號的逾期表)'] = df5['當月逾期金額(依會計每月20號的逾期表)'].round(0)
df5 = df5.drop(['合計未收'], axis=1)

逾期_merge = pd.merge(共額_merge, df5, on="客戶編號", how='left')
逾期_merge['當月逾期金額(依會計每月20號的逾期表)'] = 逾期_merge['當月逾期金額(依會計每月20號的逾期表)'].fillna(0)
逾期_merge['當月逾期金額(依會計每月20號的逾期表)'] = 逾期_merge['當月逾期金額(依會計每月20號的逾期表)'].astype('int')
two_years_ago = datetime.now() - relativedelta(years=2)
two_years_str = two_years_ago.strftime('%Y/%m/%d')

sap = kd.get_data_from_MSSQL(f'''
        SELECT buyer as 客戶編號, taxfree_basecurr as 未稅本位幣,  planned_shipping_date as 預計發貨日期 
        FROM sap_sales_data 
        WHERE  (buyer LIKE '%SG%' or buyer like '%MY%')
        AND planned_shipping_date >= '{two_years_str}'
        AND taxfree_basecurr > 0
    ''')

sap = (sap.groupby('客戶編號', as_index=False)['未稅本位幣'].sum().query("`未稅本位幣` > 0"))

銷貨_merge = pd.merge(逾期_merge, sap, on="客戶編號", how='left')
銷貨_merge['近2年有交易(現金可轉月結)'] = np.where(銷貨_merge['未稅本位幣']>0, 'V', 'X')
銷貨_merge['近2年有交易(現金可轉月結)'] = 銷貨_merge['近2年有交易(現金可轉月結)'].fillna(value="X")
not_monthly_sg = pd.read_excel(共額_path,sheet_name="SG",usecols=['不轉月結名單']).rename(columns={'不轉月結名單': '客戶編號'})
not_monthly_my = pd.read_excel(共額_path,sheet_name="MY",usecols=['不轉月結名單']).rename(columns={'不轉月結名單': '客戶編號'})
not_monthly_df = pd.concat([not_monthly_my, not_monthly_sg])
not_monthly_df[not_monthly_df['客戶編號'].notna() & (not_monthly_df['客戶編號'].astype(str).str.strip() != "")]


not_monthly_list = not_monthly_df['客戶編號'].dropna().tolist()

total_money = kd.get_data_from_MSSQL('''
        SELECT  [KUNNR]  客戶編號,sum([DMBTR])  '已兌現金額(稅後)'
            FROM [SAPdb].[dbo].[ZFI66]
            where (BUKRS like '%SG%' or BUKRS like '%MY%')  and ( VKORG like '%SG%' or  VKORG like '%MY%')
            group by [KUNNR]
        ''')
銷貨_merge = pd.merge(銷貨_merge, total_money, on="客戶編號", how='left')
銷貨_merge['已兌現金額(未稅)'] = (銷貨_merge['已兌現金額(稅後)'].fillna(0).div(1.09).add(1e-8).round(0))


merged_df = 銷貨_merge.copy()

group_sums = merged_df[merged_df['共用額度組別'].notna()].groupby('共用額度組別')['已兌現金額(未稅)'].sum().reset_index()
group_sums.columns = ['共用額度組別', '共用組合計']
merged_df = pd.merge(merged_df, group_sums, on='共用額度組別', how='left')
merged_df['已兌現金額(未稅)'] = merged_df.apply(lambda row: row['共用組合計'] if pd.notna(row['共用組合計']) else row['已兌現金額(未稅)'], axis=1)
merged_df.drop(columns=['共用組合計'], inplace=True)
merged_df['共用額度組別分類'] = merged_df['共用額度組別分類'].replace(['', 'nan', 'None'], np.nan)
merged_df['共用額度組別分類'] = merged_df.apply(lambda x: x['客戶編號'][:2] if pd.isna(x['共用額度組別分類']) else x['共用額度組別分類'],axis=1)
級別及額度_file = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\✨規範.資料"
級別及額度_keyword = "月結申請暨交易額度"
級別及額度_path = kd.get_latest_excel(級別及額度_file, 級別及額度_keyword)

level_df = pd.read_excel(級別及額度_path, header=1)  # 這裡 header=1 因為標題行在第2列（從0算起）
sg_df = level_df.iloc[:, 0:3].copy()   # A, B, C 欄
sg_df.columns = ['級別', '累計兌現金額', '交易額度']  # 可依實際調整欄名
sg_df = sg_df.dropna(how='all')        # 去掉全空列
sg_df['累計兌現金額'] = pd.to_numeric(sg_df['累計兌現金額'], errors='coerce').astype(pd.Int64Dtype()).fillna(0)
sg_df['交易額度'] = pd.to_numeric(sg_df['交易額度'], errors='coerce').astype(pd.Int64Dtype())
my_df = level_df.iloc[:, 4:7].copy()   # E, F, G 欄
my_df.columns = ['級別', '累計兌現金額', '交易額度']
my_df = my_df.dropna(how='all')
my_df['累計兌現金額'] = pd.to_numeric(my_df['累計兌現金額'], errors='coerce').astype(pd.Int64Dtype()).fillna(0)
my_df['交易額度'] = pd.to_numeric(my_df['交易額度'], errors='coerce').astype(pd.Int64Dtype())

def get_credit_limit(row):
    used_amount = row['已兌現金額(未稅)']
    allocation_ratio = row['額度分配比例']
    category = row['共用額度組別分類']  # 新增這個欄位的判斷
    if category == 'MY':
        level_table = my_df
    else:
        level_table = sg_df

    credit_limit = 0
    for _, level_row in level_table.iterrows():
        if pd.notna(used_amount) and used_amount >= level_row['累計兌現金額']:
            credit_limit = level_row['交易額度']
        else:
            break
    if pd.isna(allocation_ratio):
        return round(credit_limit, 0)
    else:
        return round(credit_limit * allocation_ratio, 0)
merged_df['交易額度'] = merged_df.apply(get_credit_limit, axis=1).fillna(0).astype(int)
merged_df.loc[merged_df['交易額度'] == 0, '交易額度'] = 1





merged_df['額度變化數值'] =  merged_df['交易額度'] - merged_df['前月交易額度']


def classify_final(row):
    if row['信用管制說明'] == '呆賬管制':
        return '額度不變(管制)', '額度不變'

    if pd.notna(row['呆帳客戶']) and str(row['呆帳客戶']).strip() != "":
        return '額度不變(呆賬)', '額度不變'

    if row['當月逾期金額(依會計每月20號的逾期表)'] > 0:
        return '額度不變(逾期)', '額度不變'

    if row['近2年有交易(現金可轉月結)'] == 'X':
        return '額度不變(凍結)', '額度不變'

    if row['客戶編號'] in not_monthly_list:
        return '額度不變(共額不轉)', '額度不變'
    diff = row['額度變化數值']

    if diff < 0:
        return '額度下降', '額度下降'
    elif diff > 0:
        return '額度上升', '額度上升'
    else:
        return '額度不變', '額度不變'

merged_df[['客戶分類', '額度變化']] = merged_df.apply(
    lambda row: pd.Series(classify_final(row)), axis=1
)
condition = merged_df['額度變化'].str.contains('額度不變', na=False)
merged_df.loc[condition, '交易額度'] = merged_df.loc[condition, '前月交易額度']
merged_df.loc[condition, '額度變化數值'] = 0




merged_df['新信用管制說明'] =  merged_df['信用管制說明']
merged_df.loc[(merged_df['信用管制說明'] == '現金客戶') & (merged_df['額度變化'] == '額度上升') | (merged_df['額度變化'] == '額度下降') , '新信用管制說明'] = '月結'
merged_df.loc[(merged_df['額度變化'] == '額度上升') & (merged_df['信用管制說明'] == '月結') &  (merged_df['新信用管制說明'] == '月結'), '客戶分類'] = '自動化(額度增加)'
merged_df.loc[(merged_df['額度變化'] == '額度上升') & (merged_df['信用管制說明'] == '現金客戶') &  (merged_df['新信用管制說明'] == '月結'), '客戶分類'] = '自動化(新月結客戶)'



merged_df.loc[(merged_df['額度變化'] == '自動化(新月結客戶)') & (merged_df['新信用管制說明'] == '現金客戶'), '新信用管制說明'] = '月結'
共用額度不轉月結_merge3 = merged_df




共用額度不轉月結_merge3.pivot_table(
    index='額度變化',    
    columns='客戶分類', 
    values='客戶編號',   
    aggfunc='count',           
    fill_value=0
)

共用額度不轉月結_merge3 = 共用額度不轉月結_merge3.drop_duplicates(subset=['客戶編號'], keep='last')

d1 = datetime.today()
d1 = d1.strftime("%Y/%m/01")
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '月結起始日'] = d1
共用額度不轉月結_merge3['月結起始日'] = pd.to_datetime(共用額度不轉月結_merge3['月結起始日'], errors='coerce', utc=True).dt.date

共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '新月結客結帳日/寄單日'] = '結帳日30日/寄單日5日'
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '票期(統一5040不可動)'] = '票期5040'
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '新月結客帳單寄送方式'] = '99 其他'
共用額度不轉月結_merge3.loc[(共用額度不轉月結_merge3['客戶分類'] == '自動化(新月結客戶)'), '新月結客付款方式'] = '02 匯款'




priority = {"自動化(新月結客戶)": 0}   # 0 最前面，其餘自動排後
共用額度不轉月結_merge3 = 共用額度不轉月結_merge3.sort_values( by="客戶分類", key=lambda s: s.map(priority).fillna(1) )




import os
from datetime import datetime
today = datetime.now().strftime("%Y%m%d")
folder = fr'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\{today}'
os.makedirs(folder, exist_ok=True)
共用額度不轉月結_merge3.to_excel(os.path.join(folder, f"SG_總額度判定{today}.xlsx"), index=False)
target_classifications = ['自動化(新月結客戶)', '額度下降', '自動化(額度增加)']
filtered_df = merged_df.loc[
    merged_df['客戶分類'].isin(target_classifications)]

credit_change_columns = ['客戶編號', '信用控制範圍', '信用額度', '風險種類', '客戶信用群組', '前次信用額度(比對用)',
    'GOOGLE申請', '比對上次', '共用額度組別', '備註', '變更']
credit_change = pd.DataFrame(columns=credit_change_columns)

credit_change['客戶編號'] = filtered_df['客戶編號'].values
credit_change['信用額度'] = filtered_df['交易額度'].values
credit_change['前次信用額度(比對用)'] = filtered_df['前月交易額度'].values
credit_change['共用額度組別'] = filtered_df['共用額度組別'].values
credit_change['備註'] = filtered_df['客戶分類'].values
new_monthly_df = merged_df.loc[merged_df['客戶分類'] == '自動化(新月結客戶)']

new_monthly_columns = ['客戶代碼', '公司代碼', '銷售組織', '配銷通路', '部門', '搜尋條件 2',
    '結帳日', '寄單日', '帳單寄送方式', '付款方式', '信用管制', '付款條件碼_K', '付款條件碼_V']
new_monthly_table = pd.DataFrame(columns=new_monthly_columns)

new_monthly_table['客戶代碼'] = new_monthly_df['客戶編號'].values
new_monthly_table['搜尋條件 2'] = datetime.today().replace(day=1).strftime('%Y-%m-%d')
new_monthly_table['結帳日'] = '結帳日30日/寄單日5日'
new_monthly_table['寄單日'] = '票期5040'
new_monthly_table['帳單寄送方式'] = '99 其他'
new_monthly_table['付款方式'] = '02 匯款'
new_monthly_table['信用管制'] = '月結'
save_path = fr'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\{today}'
file_name = f"SG_sap大批匯入(給惠茹)-{today}.xlsx"  
full_path = os.path.join(save_path, file_name)  
with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
    new_monthly_table.to_excel(writer, sheet_name='改月結', index=False)
    credit_change.to_excel(writer, sheet_name='額度變更', index=False)

    workbook = writer.book
    ws1 = workbook.create_sheet(title='共用')
    ws1.append(['客戶編號', '共用額度組別', '備註'])
    ws2 = workbook.create_sheet(title='轉現金')
    ws2.append(['客戶編號', '信用管制', '信用額度'])
    workbook.save(full_path)

print(f"已經成功匯出到 {full_path}！")

columns_for_sales = [
    '客戶編號', '名稱1', '搜尋條件1', '大區', '區域', '類別', '呆帳客戶',
    '客戶分類', '新信用管制說明', '交易額度',
    
    '月結客戶維持現金客戶(業務確認)', '維持原因(業務填寫)',
    '收款聯絡人', '收款聯絡人電話', '個人LINE網址', '群组LINE網址', '收款備註',
    '業務送帳單/收款需填原因',

    '新月結客結帳日/寄單日', '票期(統一5040不可動)', '新月結客帳單寄送方式', '新月結客付款方式',
    '月結起始日', '小區'
]

final_sales_table = pd.DataFrame(columns=columns_for_sales)

source_columns = [
    '客戶編號', '名稱1', '搜尋條件1', '大區', '區域', '類別', '呆帳客戶',
    '客戶分類', '新信用管制說明', '交易額度',
    '新月結客結帳日/寄單日', '票期(統一5040不可動)', '新月結客帳單寄送方式', '新月結客付款方式',
    '月結起始日', '小區']

for col in source_columns:
    if col in filtered_df.columns:
        final_sales_table[col] = filtered_df[col].values


info_file = r"Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度"
info_keyword = "ALL客戶收款清單聯絡資訊"
info_path = kd.get_latest_excel(info_file, info_keyword)
info_df = pd.read_excel(info_path)
info_df = info_df[['客戶', '收款連絡人', '收款連絡人電話', '收款人LINE ID', '群組LINE ID', '收款備註']]
info_df = info_df.drop_duplicates(subset='客戶', keep='first')
final_sales_table['客戶編號'] = final_sales_table['客戶編號'].astype(str).str.strip()
info_df['客戶'] = info_df['客戶'].astype(str).str.strip()
info_dict = info_df.set_index('客戶').to_dict(orient='index')
final_sales_table['收款聯絡人'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款連絡人', pd.NA))
final_sales_table['收款聯絡人電話'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款連絡人電話', pd.NA))
final_sales_table['個人LINE網址'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款人LINE ID', pd.NA))
final_sales_table['群组LINE網址'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('群組LINE ID', pd.NA))
final_sales_table['收款備註'] = final_sales_table['客戶編號'].map(lambda x: info_dict.get(x, {}).get('收款備註', pd.NA))
final_sales_table['月結起始日'] = datetime.today().replace(day=1).strftime('%Y-%m-%d')



folder = fr'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\9.自動化交易額度\★自動化月結\2.  SG 新加坡\{today}'
os.makedirs(folder, exist_ok=True)
final_sales_table.to_excel(os.path.join(folder, f"SG_新交易額度客戶明細-業務{today}.xlsx"), index=False)



