
import os
import sys
import re
import json
import ast
import base64
import hashlib
from pathlib import Path
from typing import Optional
from datetime import datetime, timedelta

import pytz
import numpy as np
import pandas as pd
import requests
import win32com.client as win32
from sqlalchemy import create_engine, text
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from PIL import ImageGrab, Image

custom_path = Path(r"C:\Users\TW0002.TPTWKD\Desktop\項目整理")
sys.path.append(str(custom_path))
import common as kd


yeaterday = pytz.timezone('Asia/Taipei').localize(datetime.now() - timedelta(days=1))
last_month_str = (datetime.today() - relativedelta(months=1)).strftime('%Y-%m')
sample_date = int(datetime(2025,9, 26).timestamp() * 1000)
today = pd.to_datetime(datetime.today().strftime('%Y-%m-%d'))
year_ago_one = pd.to_datetime((datetime.today() - relativedelta(years=1) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_one = pd.to_datetime((datetime.today() - relativedelta(months=1) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_two = pd.to_datetime((datetime.today() - relativedelta(months=2) + relativedelta(days=1)).date()).timestamp()*1000
month_ago_three = pd.to_datetime((datetime.today() - relativedelta(months=3) + relativedelta(days=1)).date()).timestamp()*1000
three_days_ago = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=-3)).timestamp() * 1000)
today_begin = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=0)).timestamp() * 1000)
today_end = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)).timestamp() * 1000)
five_days_ago = int((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=-50)).timestamp() * 1000)




def get_crm_export_records_all_platforms(
    location: str = "TW",
    account: str = "BI",
    user_id: Optional[int] = None,
    belong_id: Optional[int] = None,
    start_time: Optional[int] = None,
    end_time: Optional[int] = None,
    page_size: int = 200
) -> pd.DataFrame:

    if location in ["", "TW"]:
        token = kd.get_access_token(account=account)
        base_url = "https://api-p10.xiaoshouyi.com/rest/data/v2.0/export/record/actions/list"
    elif location == "ML":
        token = kd.get_access_token_ml(account=account)
        base_url = "https://api-scrm.xiaoshouyi.com/rest/data/v2.0/export/record/actions/list"
    else:
        raise ValueError("location 必須為 'TW' 或 'ML'")

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"}

    all_data = []
    for platform in [0, 1]:
        page_no = 1
        print(f"查詢 platform={platform} 中...")
        while True:
            payload = {
                "pageNo": page_no,
                "pageSize": page_size,
                "platform": platform}

            if belong_id:
                payload["belongId"] = belong_id
            if user_id:
                payload["userId"] = user_id
            if start_time:
                payload["startTime"] = start_time
            if end_time:
                payload["endTime"] = end_time

            try:
                response = requests.post(base_url, headers=headers, json=payload)
                response.raise_for_status()
                result = response.json()
            except Exception as e:
                print(f"platform={platform} 第 {page_no} 頁查詢失敗：{e}")
                break

            if result.get("code") != "200":
                print(f" 失敗：{result.get('msg', '未知錯誤')}")
                break

            records = result.get("data", {}).get("records", [])
            if not records:
                break

            for r in records:
                r["platform"] = platform
            all_data.extend(records)

            if len(records) < page_size:
                break
            page_no += 1

    df = pd.DataFrame(all_data)
    if not df.empty:
        df = kd.convert_to_date(df, "createdAt", "createdAt_date")
        df = kd.convert_to_date(df, "updatedAt", "updatedAt_date")
        print(f" 共取得 {len(df)} 筆導出紀錄（含前台與後台）")

    return df


import re


def extract_target_and_time(filename):
    if pd.isna(filename):
        return pd.Series([None, pd.NaT])

    match = re.search(r"\d+([^\d_]+)_(\d{14,17})_", str(filename))
    if not match:
        return pd.Series([None, pd.NaT])

    target = match.group(1)
    time_str = match.group(2)[:14]
    try:
        export_time = pd.to_datetime(time_str, format="%Y%m%d%H%M%S", errors="coerce")
    except Exception:
        export_time = pd.NaT

    return pd.Series([target, export_time])




def assign_limit_values_corrected(staff_info_df: pd.DataFrame, rules_df: pd.DataFrame) -> pd.DataFrame:
    df = staff_info_df.copy()
    df['final_department'] = np.where(
        df['gd4'].isnull() | (df['gd4'].str.strip() == ''),
        df['tmp_departname'],
        df['gd4'])

    match_cols_map = {
        'gd4': 'final_department', 
        'tmp_departname': 'tmp_departname', 
        'jobcodename': 'jobcodename', 
        'employeename': 'employeename'}
    for col in match_cols_map.values():
        if col in df.columns:
            df[col] = df[col].fillna('')

    df['crm_demand_level'] = '無需求'
    df['daily_max'] = 0
    df['monthly_max'] = 0
    df['_matched_'] = False
    print("匹配.")

    for index, rule in rules_df.iterrows():
        mask = ~df['_matched_']
        if not mask.any():
            print("所有員工均已匹配，提前結束。")
            break

        rule_conditions = []
        for rule_col, staff_col in match_cols_map.items():
            if pd.notna(rule[rule_col]):

                mask &= (df[staff_col] == rule[rule_col])
                rule_conditions.append(f"'{staff_col}'=='{rule[rule_col]}'")

        if mask.any():
            matched_count = mask.sum()
            print(f"  - #規則 {index+1} ({' & '.join(rule_conditions)}) -> 匹配到 {matched_count} 人")
            df.loc[mask, 'crm_demand_level'] = rule['CRM下載需求度']
            df.loc[mask, 'daily_max'] = rule['daily_max']
            df.loc[mask, 'monthly_max'] = rule['monthly_max']
            df.loc[mask, '_matched_'] = True
    df = df.drop(columns=['_matched_', 'final_department'])
    print(f" 已成功為 {len(df)} 筆員工資料匹配導出上限。")
    return df




target_list = [
    "K大預約表",
    "K大預約參會人",
    "K大顧問暨電訪人員名單",
    "SAP銷貨明細",
    "人員資訊",
    "上下游",
    "公司型態",
    "分公司資訊",
    "外勤業務 業務負責區域",
    "目標客戶標籤歷程",
    "交辦管理-台灣",
    "交辦管理-海外",
    "地址",
    "客戶",
    "客戶普查",
    "客戶關係連絡人",
    "客戶關係連絡人普查",
    "客訴處理回報",
    "拜訪記錄",
    "展示館預約 參訪記錄",
    "展館營業日 接待清單",
    "追蹤記錄",
    "專案時數申請",
    "專案時數明細",
    "產品庫存",
    "連絡人",
    "部門",
    "報價單",
    "發放申請",
    "發放明細",
    "鄉、鎮、市、區",
    "銷售線索",
    "銷售線索（新客）",
    "銷售機會",
    "縣、市",
    "簡訊發送",
    "關聯公司（主）",
    "關聯公司明細"
]


end_time = int(datetime.now().timestamp() * 1000)
start_time = int((datetime.now() - timedelta(days=5)).timestamp() * 1000)

export_df = get_crm_export_records_all_platforms(
    location="TW",
    start_time=start_time,
    end_time=end_time)

keep_cols = [ "id", "numberRecordsProcessed", "originFileName", "userIp","createdBy", "createdAt_date","requestFileUrl"]
export_df = export_df[[c for c in keep_cols if c in export_df.columns]].copy()

user_info = kd.get_data_from_CRM(''' select id createdBy,employeeCode employeeid from user ''')
user_info["employeeid"] = user_info["employeeid"].astype(str).str.upper()
export_df["createdBy"] = export_df["createdBy"].astype(str)
user_info["createdBy"] = user_info["createdBy"].astype(str)
merged_df = pd.merge(export_df, user_info, on="createdBy", how="left").drop(columns=["createdBy"])

staff_info = kd.get_data_from_MSSQL('''
SELECT 
    employeename,
    employeeid,
    CASE 
        WHEN gd4 LIKE '%虛擬%' THEN gd5 
        WHEN gd4 IS NULL OR LTRIM(RTRIM(gd4)) = '' THEN tmp_departname
        ELSE gd4 
    END AS gd4,
    tmp_departname,
    jobcodename
FROM [clean_data].[dbo].[hrs_staff_info_valid]
''')
staff_info["employeeid"] = staff_info["employeeid"].astype(str).str.upper()
staff_info["gd4"] = staff_info["gd4"].apply(lambda x: str(x).split("_")[-1] if "_" in str(x) else x)
merged_df_dep = pd.merge(merged_df, staff_info, on="employeeid", how="inner")

merged_df_dep[["target", "export_time"]] = merged_df_dep["originFileName"].apply(extract_target_and_time)
merged_df_dep = merged_df_dep[merged_df_dep["target"].isin(target_list)].copy()

kd.write_to_sql(merged_df_dep, db_name="clean_data", table_name="crm_export_records", if_exists="update", dedup_keys=["id"])







rules_df = pd.read_excel(r"Z:\07_資訊單位\02_數據中心課\文斌\CRM下載播報\CRM_下載需求度表.xlsx")
staff_info_limited = assign_limit_values_corrected(staff_info, rules_df)
download_info = kd.get_data_from_MSSQL(f'''
        SELECT *
        FROM [clean_data].[dbo].[crm_export_records]
        WHERE TRY_CONVERT(date, createdAt_date) >= DATEADD(MONTH, -6, CAST(GETDATE() AS date))
''')





download_info["createdAt_date"] = pd.to_datetime(download_info["createdAt_date"], errors="coerce")
today = pd.Timestamp.today().normalize()
yesterday = today - timedelta(days=1)
month_ago_6 = today - pd.DateOffset(months=6)
day_ago_30 = today - timedelta(days=30)

df = download_info[download_info["createdAt_date"] >= month_ago_6].copy()

df_yesterday = df[df["createdAt_date"] == yesterday]
summary_yesterday = (
    df_yesterday.groupby(["employeeid", "employeename", "gd4"])
    .agg(download_count_yesterday=("id", "count"),
        download_success_yesterday=("numberRecordsProcessed", "sum") ) .reset_index())


df_30d = df[df["createdAt_date"] >= day_ago_30].copy()
daily_agg = (
    df_30d.groupby(["employeeid", "employeename", "gd4", "createdAt_date"])
    .agg(daily_download_count=("id", "count"), daily_download_success=("numberRecordsProcessed", "sum"))
    .reset_index())


daily_avg = (
    daily_agg[daily_agg["daily_download_count"] > 0]
    .groupby(["employeeid", "employeename", "gd4"])
    .agg(avg_daily_download_count=("daily_download_count", "mean"),
        avg_daily_download_success=("daily_download_success", "mean")).reset_index())


df["month"] = df["createdAt_date"].dt.to_period("M")
monthly_agg = (
    df.groupby(["employeeid", "employeename", "gd4", "month"])
    .agg(monthly_download_success=("numberRecordsProcessed", "sum")).reset_index())


download_30d = (
    df_30d.groupby(["employeeid", "employeename", "gd4"])
    .agg( total_download_30d=("id", "count"),
        total_success_30d=("numberRecordsProcessed", "sum")).reset_index())



monthly_avg = (
    monthly_agg[monthly_agg["monthly_download_success"] > 0]
    .groupby(["employeeid", "employeename", "gd4"])
    .agg( total_download_6m=("monthly_download_success", "sum"),
        avg_monthly_download=("monthly_download_success", "mean")).reset_index())


summary = (
    summary_yesterday
    .merge(daily_avg, on=["employeeid", "employeename", "gd4"], how="outer")
    .merge(download_30d, on=["employeeid", "employeename", "gd4"], how="outer")
    .merge(monthly_avg, on=["employeeid", "employeename", "gd4"], how="outer"))


summary = summary.fillna(0)
summary = summary.sort_values(["gd4", "employeename"])
print(f" 統計完成，共 {len(summary)} 位員工")

daily_report = pd.merge(summary,staff_info_limited[['employeeid','daily_max','monthly_max']], on = 'employeeid', how = 'left' )


def exceed_level(row):
    res = []
    daily_ratio = (row["download_success_yesterday"] / row["daily_max"]) if row["daily_max"] else 0
    monthly_ratio = (row["total_success_30d"] / row["monthly_max"]) if row["monthly_max"] else 0
    daily_ratio_pct = round(daily_ratio * 100, 1)
    monthly_ratio_pct = round(monthly_ratio * 100, 1)

    if daily_ratio >= 1:
        res.append(f"日閾值({daily_ratio_pct}%)")

    if monthly_ratio >= 1:
        res.append(f"月閾值({monthly_ratio_pct}%)")

    return "、".join(res) if res else "正常"

daily_report["exceed_limit"] = daily_report.apply(exceed_level, axis=1)
daily_report["is_exceed"] = np.where(daily_report["exceed_limit"] != "正常", 1, 0)

def extract_daily_pct(text):
    match = re.search(r"日閾值\(([\d\.]+)%\)", text)
    return float(match.group(1)) if match else 0

def extract_monthly_pct(text):
    match = re.search(r"月閾值\(([\d\.]+)%\)", text)
    return float(match.group(1)) if match else 0

daily_report["daily_ratio_pct"] = daily_report["exceed_limit"].apply(extract_daily_pct)
daily_report["monthly_ratio_pct"] = daily_report["exceed_limit"].apply(extract_monthly_pct)


exceed_df = daily_report.query("is_exceed == 1").copy()


exceed_df = exceed_df[
    ["employeename", "gd4", "download_success_yesterday", "total_success_30d",
     "exceed_limit", "daily_ratio_pct", "monthly_ratio_pct"]
].rename(columns={
    "employeename": "員工姓名",
    "gd4": "部門",
    "download_success_yesterday": "昨日下載成功數",
    "total_success_30d": "近30天下載成功數",
    "exceed_limit": "超標類型"})


exceed_df = exceed_df.sort_values( ["部門", "daily_ratio_pct", "monthly_ratio_pct"],
    ascending=[True, False, False]).reset_index(drop=True)


exceed_df = exceed_df.drop(columns=["daily_ratio_pct", "monthly_ratio_pct"])


save_dir = r"Z:\07_資訊單位\02_數據中心課\文斌\CRM下載播報"
file_name = f"CRM下載數量{datetime.today().strftime('%Y_%m_%d')}.xlsx"
save_path = os.path.join(save_dir, file_name)


with pd.ExcelWriter(save_path, engine="xlsxwriter") as writer:
    daily_report.to_excel(writer, sheet_name="日彙總報表", index=False)
    download_info.to_excel(writer, sheet_name="原始下載明細", index=False)
    exceed_df.to_excel(writer, sheet_name="播報數據", index=False)


wb = load_workbook(save_path)
ws = wb["播報數據"]

column_widths = [15, 15, 20, 20, 35]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = width

center_align = Alignment(horizontal="center", vertical="center")

fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    exceed_text = str(row[4].value)
    fill = fill_blue
    numbers = [float(x) for x in re.findall(r"(\d+(?:\.\d+)?)%", exceed_text)]
    if numbers:
        max_percent = max(numbers)
        if max_percent >= 150:
            fill = fill_red
        elif max_percent >= 120:
            fill = fill_orange
    for cell in row[:5]:
        cell.alignment = center_align
        cell.fill = fill


wb.save(save_path)
wb.close()






excel_path =save_path
sheet_name = "播報數據"
df = pd.read_excel(excel_path, sheet_name=sheet_name)
if df.empty:
    print("播報數據為空，跳過截圖與上傳。")
    raise SystemExit
img_path = excel_path.replace(".xlsx", "_excel截圖.png")
try:
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(excel_path)
    ws = wb.Sheets(sheet_name)
    for shp in ws.Shapes:
        shp.Delete()
    rng = ws.UsedRange
    rng.CopyPicture(Appearance=1, Format=2)
    ws.Paste()
    shape = ws.Shapes(ws.Shapes.Count)
    shape.Export(img_path, 2)
    print(f"截圖已生成：{img_path}")
    wb.Close(SaveChanges=False)
    excel.Quit()

except Exception as e:
    print("截圖失敗：", e)


webhook_key = "e67628a0-0a36-4768-9c6c-3a8699011aba"
send_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={webhook_key}"
upload_url_file = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={webhook_key}&type=file"


try:
    with open(img_path, "rb") as f:
        img_bytes = f.read()
        img_base64 = base64.b64encode(img_bytes).decode('utf-8')
        img_md5 = hashlib.md5(img_bytes).hexdigest()
    msg_img = {"msgtype": "image",
        "image": {"base64": img_base64,
            "md5": img_md5 }}
    res = requests.post(send_url, data=json.dumps(msg_img))
    if res.json().get("errcode") == 0:
        print("播報圖片已成功發送。")
    else:
        print("圖片發送失敗：", res.text)
except Exception as e:
    print("圖片發送錯誤：", e)


try:
    with open(excel_path, "rb") as f:
        res = requests.post(upload_url_file, files={"media": f})
        upload_res = res.json()
        if upload_res.get("errcode") == 0:
            media_id_excel = upload_res["media_id"]
            msg_file = {"msgtype": "file", "file": {"media_id": media_id_excel}}
            res = requests.post(send_url, data=json.dumps(msg_file))
            if res.json().get("errcode") == 0:
                print("Excel 檔案已成功發送。")
            else:
                print("Excel 發送失敗：", res.text)
        else:
            print("Excel 上傳失敗：", upload_res)
except Exception as e:
    print("Excel 上傳錯誤：", e)







 